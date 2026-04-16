"""
Dashboard de Análise de Dados — Locvix
ERP: GestãoClick via API REST
Gera relatório Excel + Dashboard HTML interativo
"""

import os
import sys
import json
import hashlib
import threading
import concurrent.futures
import html as html_mod
from datetime import datetime, timedelta, date
import math
import time
import base64
import tempfile

# Fix Windows console encoding to support Unicode / emoji in print()
if sys.platform == "win32":
    import io
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "buffer"):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import requests
import pandas as pd
from openpyxl import load_workbook

# reportlab — geração de PDF de orçamentos (opcional: degradação suave se ausente)
try:
    from io import BytesIO as _BytesIO
    from reportlab.lib.pagesizes import A4 as _A4
    from reportlab.lib import colors as _rlcolors
    from reportlab.lib.units import mm as _mm
    from reportlab.platypus import (SimpleDocTemplate as _SDT, Table as _RLTable,
                                     TableStyle as _RLTableStyle, Paragraph as _Para,
                                     Spacer as _Spacer, Image as _RLImg,
                                     KeepTogether as _KeepTogether)
    from reportlab.lib.styles import ParagraphStyle as _PS
    from reportlab.lib.enums import TA_CENTER as _TAC, TA_RIGHT as _TAR, TA_LEFT as _TAL
    _REPORTLAB_OK = True
except ImportError:
    _REPORTLAB_OK = False
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ══════════════════════════════════════════════════════════════════
#  CONFIGURAÇÕES — preencha as credenciais ou use variáveis de ambiente
# ══════════════════════════════════════════════════════════════════
# As credenciais ficam dentro do ERP GestãoClick:
#   Configurações → Integrações → API → Access Token / Secret Token
GCK_ACCESS_TOKEN  = os.getenv("GCK_ACCESS_TOKEN",  "SEU_ACCESS_TOKEN_AQUI")
GCK_SECRET_TOKEN  = os.getenv("GCK_SECRET_TOKEN",  "SEU_SECRET_TOKEN_AQUI")

# URL base da API GestãoClick
GCK_BASE_URL = "https://api.beteltecnologia.net/api"

# ── Lojas cadastradas no GestãoClick ─────────────────────────────────────────
LOJA_GJ_ID  = "521831"   # G & J (loja padrão)
LOJA_WA_ID  = "65731"    # W & A
# LOJA_FILTRO: None = padrão da API | "521831" | "65731" | "ambas"
LOJA_FILTRO: str | None = None

# ── Credenciais Dixiponto (Ponto Colaborador) ──────────────────────────────
DIXI_EMAIL   = os.getenv("DIXI_EMAIL",   "")
DIXI_SENHA   = os.getenv("DIXI_SENHA",   "")
DIXI_UNIDADE = os.getenv("DIXI_UNIDADE", "locacao-guindastes")

# ── Credenciais Supabase (LocvixApp — Horas) ─────────────────────────────────
SUPABASE_URL  = os.getenv("SUPABASE_URL",  "https://fjgugglxqyhlyxwzvdts.supabase.co")
SUPABASE_ANON = os.getenv("SUPABASE_ANON", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZqZ3VnZ2x4cXlobHl4d3p2ZHRzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjAzODI4NzQsImV4cCI6MjA3NTk1ODg3NH0.YGQIfxghu4yK58iVoklI1YkwwH6aoprZ06LUsWzcYPk")

# Período padrão (últimos 12 meses)
_hoje       = datetime.now()
_12m_atras  = _hoje.replace(year=_hoje.year - 1) if _hoje.month > 0 else _hoje
DATA_INI    = os.getenv("GCK_DATA_INI", (_hoje - __import__("datetime").timedelta(days=365)).strftime("%d/%m/%Y"))
DATA_FIM    = os.getenv("GCK_DATA_FIM", _hoje.strftime("%d/%m/%Y"))

# Pastas de saída
_agora       = datetime.now().strftime("%Y%m%d_%H%M%S")
_BASE_SAIDA  = r"Z:\codigos\Locvix GPT" if os.path.isdir(r"Z:\codigos\Locvix GPT") else os.path.dirname(os.path.abspath(__file__))
SAIDA_EXCEL  = os.path.join(_BASE_SAIDA, f"Relatorio_Locvix_{_agora}.xlsx")
SAIDA_HTML   = os.path.join(_BASE_SAIDA, f"Dashboard_Locvix_{_agora}.html")

# Cache em disco (pasta _cache_gck/ junto ao script)
_BASE_DIR   = _BASE_SAIDA
_CACHE_DIR  = os.path.join(
    tempfile.gettempdir() if os.name != "nt" else _BASE_DIR,
    "_cache_gck"
)
os.makedirs(_CACHE_DIR, exist_ok=True)

# TTL de cache (em segundos)
_TTL_VENDAS    = 1800   # 30 min
_TTL_OUTROS    = 86400  # 24 h
_TTL_SUPABASE  = 300    # 5 min para dados operacionais do app

# Callback de progresso (usado pelo Streamlit ou similar)
_progresso = None

# Flag para forçar refresh do ponto (ignorar cache em disco)
_SKIP_PONTO_CACHE = False
_SKIP_CACHE       = False   # quando True, ignora todo cache em disco (Atualizar Dados)

def _prog(pct: float, msg: str = ""):
    if callable(_progresso):
        try:
            _progresso(min(float(pct), 1.0), msg)
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════
#  CLIENTE HTTP — GestãoClick API
# ══════════════════════════════════════════════════════════════════
class GCKClient:
    """
    Cliente para a API REST do GestãoClick.
    Autenticação via headers access-token + secret-access-token.
    Suporte a paginação automática.
    """

    def __init__(self, access_token: str, secret_token: str, base_url: str = GCK_BASE_URL):
        self.base_url = base_url.rstrip("/")
        self.session  = requests.Session()
        self.session.headers.update({
            "access-token":        access_token,
            "secret-access-token": secret_token,
            "Content-Type":        "application/json",
            "Accept":              "application/json",
        })
        self._rate_lock = threading.Lock()
        self._last_call = 0.0

    def _throttle(self, intervalo: float = 0.35):
        with self._rate_lock:
            espera = intervalo - (time.time() - self._last_call)
            if espera > 0:
                time.sleep(espera)
            self._last_call = time.time()

    def get(self, endpoint: str, params: dict | None = None, tentativas: int = 5) -> dict | None:
        """GET com retry e rate-limit."""
        url = f"{self.base_url}/{endpoint.lstrip('/')}"
        for t in range(tentativas):
            self._throttle()
            try:
                r = self.session.get(url, params=params or {}, timeout=60)
                if r.status_code == 429:
                    espera = 30 * (t + 1)
                    print(f"  [AVISO] Rate limit (429), aguardando {espera}s...")
                    time.sleep(espera)
                    continue
                r.raise_for_status()
                return r.json()
            except requests.exceptions.HTTPError as e:
                print(f"  [ERRO] HTTP {e.response.status_code} em {url}")
                if e.response.status_code in (401, 403):
                    print("  ⚠ Credenciais inválidas ou sem permissão de API.")
                    return None
                if e.response.status_code == 404:
                    print(f"  [INFO] Endpoint '{endpoint}' não disponível neste plano (404).")
                    return None
                time.sleep(2 * (t + 1))
            except Exception as e:
                print(f"  [AVISO] {endpoint} tentativa {t+1}: {e}")
                time.sleep(2 * (t + 1))
        return None

    def post(self, endpoint: str, body: dict, loja_id: str | None = None) -> dict | None:
        """POST com retry e rate-limit."""
        url = f"{self.base_url}/{endpoint.lstrip('/')}"
        if loja_id:
            body = {**body, "loja_id": loja_id}
        for t in range(5):
            self._throttle()
            try:
                r = self.session.post(url, json=body, timeout=60)
                if r.status_code == 429:
                    espera = 30 * (t + 1)
                    print(f"  [AVISO] Rate limit (429), aguardando {espera}s...")
                    time.sleep(espera)
                    continue
                r.raise_for_status()
                return r.json()
            except requests.exceptions.HTTPError as e:
                print(f"  [ERRO] POST HTTP {e.response.status_code} em {url}: {e.response.text[:200]}")
                return {"code": e.response.status_code, "status": "error",
                        "errors": e.response.text[:500]}
            except Exception as e:
                print(f"  [AVISO] POST {endpoint} tentativa {t+1}: {e}")
                time.sleep(2 * (t + 1))
        return None

    def put(self, endpoint: str, body: dict, loja_id: str | None = None) -> dict | None:
        """PUT com retry e rate-limit."""
        url = f"{self.base_url}/{endpoint.lstrip('/')}"
        if loja_id:
            body = {**body, "loja_id": loja_id}
        for t in range(5):
            self._throttle()
            try:
                r = self.session.put(url, json=body, timeout=60)
                if r.status_code == 429:
                    espera = 30 * (t + 1)
                    print(f"  [AVISO] Rate limit (429), aguardando {espera}s...")
                    time.sleep(espera)
                    continue
                r.raise_for_status()
                return r.json()
            except requests.exceptions.HTTPError as e:
                print(f"  [ERRO] PUT HTTP {e.response.status_code} em {url}: {e.response.text[:200]}")
                return {"code": e.response.status_code, "status": "error",
                        "errors": e.response.text[:500]}
            except Exception as e:
                print(f"  [AVISO] PUT {endpoint} tentativa {t+1}: {e}")
                time.sleep(2 * (t + 1))
        return None

    def delete(self, endpoint: str) -> dict | None:
        """DELETE com retry e rate-limit."""
        url = f"{self.base_url}/{endpoint.lstrip('/')}"
        for t in range(3):
            self._throttle()
            try:
                r = self.session.delete(url, timeout=60)
                if r.status_code == 429:
                    time.sleep(30 * (t + 1))
                    continue
                if r.status_code == 204:
                    return {"code": 204, "status": "success"}
                r.raise_for_status()
                return r.json()
            except requests.exceptions.HTTPError as e:
                print(f"  [ERRO] DELETE HTTP {e.response.status_code} em {url}: {e.response.text[:200]}")
                return {"code": e.response.status_code, "status": "error",
                        "errors": e.response.text[:500]}
            except Exception as e:
                print(f"  [AVISO] DELETE {endpoint} tentativa {t+1}: {e}")
                time.sleep(2 * (t + 1))
        return None

    def paginar(self, endpoint: str, params: dict | None = None,
                campo_dados: str = "data", limite: int = 100) -> list[dict]:
        """
        Busca todas as páginas de um endpoint paginado.
        Parâmetros de paginação: pagina / limite.
        """
        todos: list[dict] = []
        params = dict(params or {})
        params["limite"] = limite
        pag = 1
        tot_pags = 1

        while pag <= tot_pags:
            params["pagina"] = pag
            resp = self.get(endpoint, params)
            if resp is None:
                break
            items = resp.get(campo_dados, resp.get("data", []))
            if isinstance(items, list):
                todos.extend(items)
            meta = resp.get("meta", {})
            tot_pags = meta.get("last_page", meta.get("ultima_pagina", meta.get("total_paginas", 1)))
            if tot_pags is None:
                tot_pags = 1
            total    = meta.get("total", meta.get("total_registros", "?"))
            if pag == 1:
                print(f"  {endpoint}: {total} registros ({tot_pags} pág.)")
            pag += 1

        return todos



# ══════════════════════════════════════════════════════════════════
#  DIXIPONTO — CLIENTE API (web.dixiponto.com.br)
# ══════════════════════════════════════════════════════════════════
class DixiPontoClient:
    """Cliente para a API interna do Dixiponto (web.dixiponto.com.br)."""
    _BASE = "https://webapiponto.dixiponto.com.br:8899"
    _AUTH = "YW5ndWxhcjpAbmd1bEByMA=="  # angular:@ngul@r0

    def __init__(self, email: str, senha: str, unidade: str):
        self.email   = email
        self.senha   = senha
        self.unidade = unidade
        self._tokens: dict = {}

    def _hdrs(self) -> dict:
        return {
            "Content-Type":      "application/json",
            "Origin":            "https://web.dixiponto.com.br",
            "X-User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "X-Language":        "pt-BR",
            "X-Resolution":      "1920x1080",
            "X-Client-Platform": "web",
        }

    def login(self) -> dict:
        r = requests.post(f"{self._BASE}/login_", json={
            "usuario": self.email, "senha": self.senha,
            "unidade": self.unidade, "suporte": False,
        }, headers=self._hdrs(), timeout=15)
        r.raise_for_status()
        data = r.json()["data"]
        if not data.get("loginSucesso"):
            raise ValueError(f"Dixiponto login falhou: {data.get('motivo', '?')}")
        return data

    def _token(self, entidade: str) -> str:
        if entidade not in self._tokens:
            r = requests.post(f"{self._BASE}/oauth/token", data={
                "username": json.dumps({
                    "login": self.email, "entidade": entidade,
                    "unidade": self.unidade, "suporte": False,
                }),
                "password":   self.senha,
                "grant_type": "password",
            }, headers={
                "Authorization":  f"Basic {self._AUTH}",
                "Content-Type":   "application/x-www-form-urlencoded",
                "X-Client-Platform": "web",
                "Origin": "https://web.dixiponto.com.br",
            }, timeout=15)
            r.raise_for_status()
            self._tokens[entidade] = r.json()["access_token"]
        return self._tokens[entidade]

    def _get(self, path: str, entidade: str, params: dict | None = None):
        r = requests.get(f"{self._BASE}{path}", params=params or {}, headers={
            "Content-Type":      "application/json; charset=UTF-8",
            "Accept":            "application/json",
            "Authorization":     f"bearer {self._token(entidade)}",
            "Entidade":          entidade,
            "Unidade":           self.unidade,
            "X-Client-Platform": "web",
            "Origin":            "https://web.dixiponto.com.br",
        }, timeout=30)
        r.raise_for_status()
        return r.json() if r.text else None

    def _all_pages(self, path: str, entidade: str, params: dict | None = None) -> list:
        params = dict(params or {})
        params.setdefault("size", 200)
        out: list = []
        for page in range(200):
            params["page"] = page
            data = self._get(path, entidade, params)
            if not data:
                break
            if isinstance(data, list):
                out.extend(data)
                break
            content = data.get("content", [])
            out.extend(content)
            if data.get("last", True) or not content:
                break
        return out

    def get_funcionarios(self) -> list:
        data = self._get("/funcionario_", "Funcionario")
        if isinstance(data, list):
            return data
        return data.get("content", []) if isinstance(data, dict) else []

    def get_marcacoes(self, d_ini: date, d_fim: date) -> list:
        return self._all_pages("/marcacao_", "Marcacao", {
            "dataInicio": d_ini.isoformat(),
            "dataFim":    d_fim.isoformat(),
        })

    @staticmethod
    def parse_data(v: int) -> str:
        s = str(v)
        return f"{s[:4]}-{s[4:6]}-{s[6:]}" if len(s) == 8 else str(v)

    @staticmethod
    def parse_hora(v: int) -> str:
        # API retorna hora em minutos desde meia-noite (ex: 423 = 07:03)
        return f"{v // 60:02d}:{v % 60:02d}"


# Instância global (inicializada no main)
_client: GCKClient | None = None

def _gck() -> GCKClient:
    global _client
    if _client is None:
        _client = GCKClient(GCK_ACCESS_TOKEN, GCK_SECRET_TOKEN)
    return _client


_empresa_cache: dict = {}

# Fallbacks por loja_id conhecida
_EMPRESA_DEFAULTS: dict[str, dict] = {
    LOJA_GJ_ID: {
        "nome":    "LOCVIX LOCAÇÕES LTDA",
        "cnpj":    "29.007.819/0001-96",
        "cidade":  "Serra", "uf": "ES",
        "telefone":"(27) 3065-2627",
        "email":   "contato@locvix.com.br",
    },
    LOJA_WA_ID: {
        "nome":    "W & A LOCAÇÕES LTDA",
        "cnpj":    "30.993.597/0001-83",
        "cidade":  "Serra", "uf": "ES",
        "telefone":"(27) 3065-2627",
        "email":   "contato@locvix.com.br",
    },
}
_EMPRESA_DEFAULT_GERAL = _EMPRESA_DEFAULTS[LOJA_GJ_ID]

def _get_empresa_info(loja_id: str | None = None) -> dict:
    """Busca dados da loja no GestãoClick e cacheia por loja_id.
    Fallback para valores fixos se o endpoint não estiver disponível."""
    global _empresa_cache
    key = str(loja_id or LOJA_GJ_ID)
    if key in _empresa_cache:
        return _empresa_cache[key]
    defaults = _EMPRESA_DEFAULTS.get(key, _EMPRESA_DEFAULT_GERAL)
    try:
        gck = _gck()
        resp = gck.get(f"lojas/{key}")
        data = (resp.get("data", {}) if resp else {}) or {}
        if data:
            end = (data.get("enderecos") or [{}])
            end0 = end[0].get("endereco", end[0]) if end else {}
            info = {
                "nome":     (data.get("razao_social") or data.get("nome") or defaults["nome"]),
                "cnpj":     (data.get("cnpj") or data.get("cpf_cnpj") or defaults["cnpj"]),
                "cidade":   (end0.get("nome_cidade") or end0.get("cidade") or data.get("cidade") or defaults["cidade"]),
                "uf":       (end0.get("estado") or end0.get("uf") or data.get("uf") or defaults["uf"]),
                "telefone": (data.get("telefone") or data.get("fone") or defaults["telefone"]),
                "email":    (data.get("email") or defaults["email"]),
            }
            _empresa_cache[key] = info
            return info
    except Exception:
        pass
    _empresa_cache[key] = defaults
    return defaults


def _paginar_lojas(endpoint: str, params: dict | None = None) -> list[dict]:
    """Busca dados respeitando LOJA_FILTRO global.
    None = padrão API | 'ambas' = W&A + G&J merged | id = loja específica.
    """
    p = dict(params or {})
    if LOJA_FILTRO == "ambas":
        r1 = _gck().paginar(endpoint, {**p, "loja_id": LOJA_GJ_ID})
        r2 = _gck().paginar(endpoint, {**p, "loja_id": LOJA_WA_ID})
        for row in r1: row["_loja"] = "GJ"
        for row in r2: row["_loja"] = "WA"
        return r1 + r2
    elif LOJA_FILTRO:
        return _gck().paginar(endpoint, {**p, "loja_id": LOJA_FILTRO})
    else:
        return _gck().paginar(endpoint, p)


def _chave_loja(chave: str) -> str:
    """Adiciona sufixo da loja à chave de cache para evitar mistura de dados."""
    if LOJA_FILTRO:
        return f"{chave}|loja:{LOJA_FILTRO}"
    return chave


# ══════════════════════════════════════════════════════════════════
#  CACHE EM DISCO
# ══════════════════════════════════════════════════════════════════
_CACHE_SCHEMA = "4"  # incremente quando o esquema de dados mudar (invalida cache antigo)

def _cache_path(chave: str) -> str:
    h = hashlib.md5(f"{_CACHE_SCHEMA}|{chave}".encode()).hexdigest()
    return os.path.join(_CACHE_DIR, f"{h}.json")

def _cache_load(chave: str, ttl: int) -> list | dict | None:
    if _SKIP_CACHE:
        return None
    p = _cache_path(chave)
    if not os.path.exists(p):
        return None
    if time.time() - os.path.getmtime(p) > ttl:
        return None
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def _cache_save(chave: str, data) -> None:
    try:
        with open(_cache_path(chave), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, default=str)
    except Exception as e:
        print(f"  [AVISO] Não foi possível salvar cache: {e}")


# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — VENDAS
# ══════════════════════════════════════════════════════════════════
def buscar_vendas(data_ini: str, data_fim: str) -> list[dict]:
    """
    Retorna lista de itens de venda do GestãoClick no período.
    Tenta os endpoints: vendas, pedidos (fallback).
    Campos normalizados: id, data, cliente, status, produto, qtd, v_bruto, v_desc, v_liq, vendedor, categoria
    """
    _prog(0.05, "Buscando vendas...")
    chave = _chave_loja(f"vendas|{data_ini}|{data_fim}")
    cached = _cache_load(chave, _TTL_VENDAS)
    if cached is not None:
        print(f"  ✔ Vendas (cache): {len(cached)} registros")
        _prog(0.30, "Vendas carregadas do cache")
        return cached

    # Converte datas para o formato esperado pela API (YYYY-MM-DD)
    def br_to_iso(d: str) -> str:
        try:
            return datetime.strptime(d, "%d/%m/%Y").strftime("%Y-%m-%d")
        except Exception:
            return d

    params = {
        "data_inicio": br_to_iso(data_ini),
        "data_fim":    br_to_iso(data_fim),
    }

    # Tenta endpoint /vendas; se falhar, tenta /pedidos
    endpoint_usado = "vendas"
    for endpoint in ["vendas", "pedidos"]:
        raw = _paginar_lojas(endpoint, params)
        if raw:
            endpoint_usado = endpoint
            break

    if not raw:
        print("  [AVISO] Nenhuma venda encontrada no período ou credenciais inválidas.")
        return []

    # GestãoClick às vezes omite produtos/servicos na listagem paginada
    _sem_itens = [v for v in raw if not v.get("produtos") and not v.get("servicos") and not v.get("itens")]
    if _sem_itens and len(_sem_itens) == len(raw):
        print(f"  ℹ Listagem sem itens detalhados — buscando cada venda individualmente ({len(raw)})...")
        gck = _gck()
        _detalhadas = []
        for i, v in enumerate(raw, 1):
            vid = v.get("id")
            if not vid:
                _detalhadas.append(v)
                continue
            loja_tag = v.get("_loja", "")
            detalhe = gck.get(f"{endpoint_usado}/{vid}")
            if detalhe:
                det = detalhe.get("data", detalhe)
                if isinstance(det, dict):
                    det["_loja"] = loja_tag
                    _detalhadas.append(det)
                    continue
            _detalhadas.append(v)
            if i % 20 == 0:
                print(f"    {i}/{len(raw)} vendas detalhadas...")
        raw = _detalhadas

    registros: list[dict] = []
    for v in raw:
        # Normaliza campos (GestãoClick pode variar o nome dos campos entre planos)
        data_raw  = v.get("data_emissao") or v.get("data_venda") or v.get("data") or ""
        try:
            data_dt = datetime.strptime(data_raw[:10], "%Y-%m-%d")
        except Exception:
            data_dt = None

        cliente   = (v.get("nome_cliente") or v.get("cliente_nome") or v.get("cliente") or "")
        status    = (v.get("nome_situacao") or v.get("status") or v.get("situacao") or "").upper()
        vendedor  = (v.get("nome_vendedor") or v.get("vendedor_nome") or v.get("vendedor") or "Sem Vendedor")
        # centro de custo da venda (nível cabeçalho)
        cc_venda  = (v.get("nome_centro_custo") or v.get("centro_custo") or v.get("centro_de_custo") or "").upper().strip()
        cat       = (v.get("categoria") or v.get("grupo") or cc_venda or "SEM CATEGORIA").upper()
        nf        = str(v.get("numero_nf") or v.get("numero") or v.get("id") or "")
        id_venda  = str(v.get("id") or "")

        # Itens da venda — GestãoClick retorna produtos e serviços em listas separadas
        # Formato: produtos=[{"produto": {...}}], servicos=[{"servico": {...}}]
        itens_raw = []
        for it in (v.get("produtos") or v.get("itens") or []):
            itens_raw.append(it.get("produto", it))
        for it in (v.get("servicos") or []):
            itens_raw.append(it.get("servico", it))
        if itens_raw:
            for it in itens_raw:
                cod   = str(it.get("produto_id") or it.get("servico_id") or it.get("codigo") or "")
                desc  = str(it.get("nome_produto") or it.get("nome_servico") or it.get("descricao") or it.get("nome") or "")
                unid  = str(it.get("sigla_unidade") or it.get("unidade") or "UN")
                # centro de custo: tenta no item primeiro, depois usa o da venda
                cc_item = (it.get("nome_centro_custo") or it.get("centro_custo") or it.get("centro_de_custo") or "").upper().strip()
                cc_final = cc_item or cc_venda or "SEM CENTRO DE CUSTO"
                try:    qtd    = float(it.get("quantidade") or 1)
                except: qtd    = 1.0
                try:    v_unit = float(it.get("valor_venda") or it.get("valor_unitario") or it.get("preco") or 0)
                except: v_unit = 0.0
                try:    v_bruto= float(it.get("valor_total") or qtd * v_unit)
                except: v_bruto= qtd * v_unit
                try:    v_desc = float(it.get("desconto_valor") or it.get("desconto") or 0)
                except: v_desc = 0.0
                registros.append({
                    "ID":            id_venda,
                    "NF":            nf,
                    "Data":          data_dt,
                    "Cliente":       cliente,
                    "Status":        status,
                    "Vendedor":      vendedor,
                    "Categoria":     cat,
                    "Centro Custo":  cc_final,
                    "Cod. Produto":  cod,
                    "Produto":       desc,
                    "Unidade":       unid,
                    "Qtd":           qtd,
                    "Vlr Unitário":  v_unit,
                    "Vlr Bruto":     v_bruto,
                    "Desconto":      v_desc,
                    "Vlr Líquido":   v_bruto - v_desc,
                    "Loja":          v.get("_loja", ""),
                })
        else:
            # Venda sem detalhamento de itens — registra a venda como um todo
            try:    v_bruto = float(v.get("valor_total") or v.get("total") or 0)
            except: v_bruto = 0.0
            try:    v_desc  = float(v.get("desconto") or v.get("valor_desconto") or 0)
            except: v_desc  = 0.0
            registros.append({
                "ID":            id_venda,
                "NF":            nf,
                "Data":          data_dt,
                "Cliente":       cliente,
                "Status":        status,
                "Vendedor":      vendedor,
                "Categoria":     cat,
                "Centro Custo":  cc_venda or "SEM CENTRO DE CUSTO",
                "Cod. Produto":  "",
                "Produto":       "",
                "Unidade":       "",
                "Qtd":           1.0,
                "Vlr Unitário":  v_bruto - v_desc,
                "Vlr Bruto":     v_bruto,
                "Desconto":      v_desc,
                "Vlr Líquido":   v_bruto - v_desc,
                "Loja":          v.get("_loja", ""),
            })

    _cache_save(chave, [
        {**r, "Data": r["Data"].isoformat() if isinstance(r["Data"], datetime) else r["Data"]}
        for r in registros
    ])

    # Converte datas de volta
    for r in registros:
        if isinstance(r["Data"], str) and r["Data"]:
            try:
                r["Data"] = datetime.fromisoformat(r["Data"][:19])
            except Exception:
                pass

    print(f"  ✔ {len(registros)} itens de venda")
    _prog(0.30, f"Vendas: {len(registros)} itens carregados")
    return registros


# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — VENDAS (EXCEL MANUAL)
# ══════════════════════════════════════════════════════════════════
def buscar_vendas_excel(data_ini: str, data_fim: str) -> list[dict]:
    """
    Lê os arquivos Excel manuais de Contas a Receber (2025 e 2026) e retorna
    a lista de recebimentos no mesmo formato de buscar_vendas().

    Estrutura da planilha:
      Linha 7 = cabeçalho: (None, DATA, CLIENTE, FORMA, TIPO, SITUAÇÃO, ATRASADO,
                             DESCONTADO, SIMPLES, DESCONTADO, SIMPLES, ...)
      Linhas 8+ = dados: banco de valores nas colunas H-Q (índices 7-16).
    """
    import os as _os
    import openpyxl as _xl

    _prog(0.05, "Lendo Excel de vendas...")

    def br_to_date(d: str):
        try:
            return datetime.strptime(d, "%d/%m/%Y")
        except Exception:
            return None

    d_ini_dt = br_to_date(data_ini)
    d_fim_dt = br_to_date(data_fim)

    base_dir = _os.path.dirname(_os.path.abspath(__file__))
    arquivos = [
        _os.path.join(base_dir, "CONTROLE CONTAS A RECEBER - 2025.xlsx"),
        _os.path.join(base_dir, "CONTROLE CONTAS A RECEBER - 2026.xlsx"),
    ]

    ABAS_IGNORAR = {"ATRASADOS", "Planilha1"}

    registros: list[dict] = []
    for caminho_xlsx in arquivos:
        if not _os.path.exists(caminho_xlsx):
            print(f"  [AVISO] Arquivo não encontrado: {caminho_xlsx}")
            continue
        try:
            wb = _xl.load_workbook(caminho_xlsx, data_only=True, read_only=True)
        except Exception as e:
            print(f"  [AVISO] Erro ao abrir {caminho_xlsx}: {e}")
            continue

        for aba_nome in wb.sheetnames:
            if aba_nome in ABAS_IGNORAR:
                continue
            ws = wb[aba_nome]
            for row_idx, row in enumerate(ws.iter_rows(min_row=8, values_only=True), start=8):
                data_cel   = row[1]   # col B
                cliente    = row[2]   # col C
                forma      = row[3]   # col D
                tipo_nf    = row[4]   # col E  (NF 0104, RE 001-12, etc.)
                situacao   = row[5]   # col F  (PAGO, A PAGAR, ATRASADO)

                # Ignora linhas de totais ou vazias
                if not cliente or not isinstance(data_cel, datetime):
                    continue

                # Soma colunas H-Q (índices 7-16) = todos os bancos DESCONTADO/SIMPLES
                valor = sum(
                    v for v in row[7:17]
                    if isinstance(v, (int, float))
                )
                if valor <= 0:
                    continue

                # Filtro de período pela data do recebimento
                if d_ini_dt and data_cel < d_ini_dt:
                    continue
                if d_fim_dt and data_cel > d_fim_dt:
                    continue

                status_str  = str(situacao or "").strip().upper() or "PAGO"
                cliente_str = str(cliente).strip().upper()
                nf_str      = str(tipo_nf).strip() if tipo_nf else ""
                forma_str   = str(forma).strip().upper() if forma else ""

                registros.append({
                    "ID":           f"XLS|{aba_nome}|{row_idx}",
                    "NF":           nf_str,
                    "Data":         data_cel,
                    "Cliente":      cliente_str,
                    "Status":       status_str,
                    "Vendedor":     "Sem Vendedor",
                    "Categoria":    "SEM CATEGORIA",
                    "Centro Custo": aba_nome.upper().strip() or "SEM CENTRO DE CUSTO",
                    "Cod. Produto": "",
                    "Produto":      forma_str,
                    "Unidade":      "",
                    "Qtd":          1.0,
                    "Vlr Unitário": valor,
                    "Vlr Bruto":    valor,
                    "Desconto":     0.0,
                    "Vlr Líquido":  valor,
                })
        wb.close()

    print(f"  ✔ Excel: {len(registros)} recebimentos no período")
    _prog(0.30, f"Vendas (Excel): {len(registros)} registros carregados")
    return registros


# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — FINANCEIRO
# ══════════════════════════════════════════════════════════════════
def buscar_financeiro(data_ini: str, data_fim: str) -> dict:
    """
    Retorna {'receber': [...], 'pagar': [...]}
    Campos: id, descricao, cliente/fornecedor, valor, data_vencimento, data_pagamento, status, categoria
    """
    _prog(0.32, "Buscando financeiro...")
    resultado = {}

    def br_to_iso(d: str) -> str:
        try:
            return datetime.strptime(d, "%d/%m/%Y").strftime("%Y-%m-%d")
        except Exception:
            return d

    params = {"data_inicio": br_to_iso(data_ini), "data_fim": br_to_iso(data_fim)}

    # Apenas Contas a Pagar (recebimentos sem dados neste plano)
    endpoints_map = [
        ("pagar", "pagamentos"),
    ]

    for tipo, endpoint in endpoints_map:
        chave  = _chave_loja(f"{endpoint}|{data_ini}|{data_fim}")
        cached = _cache_load(chave, _TTL_VENDAS)
        if cached is not None:
            resultado[tipo] = cached
            print(f"  ✔ {tipo} (cache): {len(cached)} registros")
            continue

        raw = _paginar_lojas(endpoint, params)
        normalizado = []
        for c in raw:
            def _float(k, _c=c):
                try:    return float(_c.get(k) or 0)
                except: return 0.0
            def _data(k, _c=c):
                v = _c.get(k, "")
                try:    return datetime.strptime(str(v)[:10], "%Y-%m-%d")
                except: return None

            # /pagamentos usa campos diferentes de /contas_pagar
            pessoa = (c.get("nome_fornecedor") or c.get("nome_cliente")
                      or c.get("nome_transportadora") or c.get("nome_funcionario")
                      or c.get("cliente_nome") or c.get("fornecedor_nome") or "")
            categoria = (c.get("nome_plano_conta") or c.get("categoria")
                         or c.get("plano_conta") or "OUTROS")
            liquidado = str(c.get("liquidado", "0")) == "1"
            status = "PAGO" if liquidado else "A PAGAR"
            valor = _float("valor_total") or _float("valor")
            pago  = valor if liquidado else 0.0

            normalizado.append({
                "ID":           str(c.get("id") or ""),
                "Descrição":    c.get("descricao") or c.get("historico") or "",
                "Pessoa":       pessoa,
                "Valor":        valor,
                "Valor Pago":   pago,
                "Saldo":        0.0 if liquidado else valor,
                "Vencimento":   _data("data_vencimento"),
                "Pagamento":    _data("data_liquidacao") or _data("data_pagamento"),
                "Status":       status,
                "Categoria":    categoria.upper(),
                "Centro Custo": (c.get("nome_centro_custo") or c.get("centro_custo") or "").upper(),
                "Loja":         c.get("_loja", ""),
            })

        _cache_save(chave, [
            {**r,
             "Vencimento": r["Vencimento"].isoformat() if isinstance(r["Vencimento"], datetime) else r["Vencimento"],
             "Pagamento":  r["Pagamento"].isoformat()  if isinstance(r["Pagamento"],  datetime) else r["Pagamento"],
            } for r in normalizado
        ])
        resultado[tipo] = normalizado

    _prog(0.45, "Financeiro carregado")
    return resultado


# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — CLIENTES
# ══════════════════════════════════════════════════════════════════
def buscar_clientes() -> list[dict]:
    """Lista de clientes cadastrados."""
    _prog(0.47, "Buscando clientes...")
    chave  = _chave_loja("clientes")
    cached = _cache_load(chave, _TTL_OUTROS)
    if cached:
        print(f"  ✔ Clientes (cache): {len(cached)}")
        return cached
    raw = _paginar_lojas("clientes")
    normalizado = []
    for c in raw:
        enderecos = c.get("enderecos") or []
        end0 = enderecos[0].get("endereco", {}) if enderecos and isinstance(enderecos[0], dict) else {}
        normalizado.append({
            "id":      str(c.get("id") or ""),
            "ID":      str(c.get("id") or ""),
            "nome":    c.get("nome") or c.get("razao_social") or "",
            "Nome":    c.get("nome") or c.get("razao_social") or "",
            "Cidade":  end0.get("nome_cidade") or c.get("cidade") or "",
            "UF":      end0.get("estado") or c.get("uf") or "",
            "CNPJ":    c.get("cnpj") or c.get("cpf") or c.get("cpf_cnpj") or "",
            "Tipo":    (c.get("tipo_pessoa") or c.get("tipo") or "PF").upper(),
            "Grupo":   (c.get("grupo") or c.get("categoria") or "").upper(),
            "Ativo":   bool(c.get("ativo", True)),
        })
    _cache_save(chave, normalizado)
    print(f"  ✔ {len(normalizado)} clientes")
    _prog(0.52, "Clientes carregados")
    return normalizado


# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — PRODUTOS / ESTOQUE
# ══════════════════════════════════════════════════════════════════
def buscar_produtos() -> list[dict]:
    """Catálogo de produtos com estoque."""
    _prog(0.53, "Buscando produtos...")
    chave  = _chave_loja("produtos")
    cached = _cache_load(chave, _TTL_OUTROS)
    if cached:
        print(f"  ✔ Produtos (cache): {len(cached)}")
        return cached
    raw = _paginar_lojas("produtos")
    normalizado = []
    for p in raw:
        def _float(k):
            try:    return float(p.get(k) or 0)
            except: return 0.0
        normalizado.append({
            "Código":     str(p.get("codigo_interno") or p.get("codigo") or p.get("id") or ""),
            "Descrição":  p.get("nome") or p.get("descricao") or "",
            "Categoria":  (p.get("nome_grupo") or p.get("categoria") or p.get("grupo") or "SEM CATEGORIA").upper(),
            "Marca":      (p.get("marca") or "").upper(),
            "Unidade":    p.get("unidade") or "UN",
            "Estoque":    _float("estoque") or _float("quantidade_estoque"),
            "Preco Venda":_float("valor_venda") or _float("preco_venda"),
            "Preco Custo":_float("valor_custo") or _float("preco_custo"),
            "Ativo":      bool(p.get("ativo", True)),
        })
    _cache_save(chave, normalizado)
    print(f"  ✔ {len(normalizado)} produtos")
    _prog(0.60, "Produtos carregados")
    return normalizado


# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — CONTRATOS
# ══════════════════════════════════════════════════════════════════
def buscar_contratos() -> list[dict]:
    """Contratos ativos (recorrência)."""
    _prog(0.61, "Buscando contratos...")
    chave  = _chave_loja("contratos")
    cached = _cache_load(chave, _TTL_OUTROS)
    if cached:
        print(f"  ✔ Contratos (cache): {len(cached)}")
        return cached
    raw = _paginar_lojas("contratos")
    normalizado = []
    for c in raw:
        def _float(k):
            try:    return float(c.get(k) or 0)
            except: return 0.0
        def _data(k):
            v = c.get(k, "")
            try:    return datetime.strptime(str(v)[:10], "%Y-%m-%d")
            except: return None
        normalizado.append({
            "ID":          str(c.get("id") or ""),
            "Número":      str(c.get("numero") or c.get("id") or ""),
            "Cliente":     c.get("cliente_nome") or c.get("nome") or "",
            "Descricao":   c.get("descricao") or c.get("objeto") or "",
            "Valor":       _float("valor") or _float("valor_total"),
            "Periodicidade":(c.get("periodicidade") or c.get("recorrencia") or "MENSAL").upper(),
            "Inicio":      _data("data_inicio"),
            "Fim":         _data("data_fim") or _data("data_vencimento"),
            "Status":      (c.get("status") or c.get("situacao") or "").upper(),
        })
    _cache_save(chave, [
        {**r,
         "Inicio": r["Inicio"].isoformat() if isinstance(r["Inicio"], datetime) else r["Inicio"],
         "Fim":    r["Fim"].isoformat()    if isinstance(r["Fim"],    datetime) else r["Fim"],
        } for r in normalizado
    ])
    print(f"  ✔ {len(normalizado)} contratos")
    _prog(0.68, "Contratos carregados")
    return normalizado


# ══════════════════════════════════════════════════════════════════
#  GERAÇÃO DE PDF — ORÇAMENTO  (layout profissional)
# ══════════════════════════════════════════════════════════════════
def _gerar_pdf_orc_bytes(d: dict, cli_data: dict) -> bytes | None:
    """
    Gera PDF profissional do orçamento (layout corporativo Locvix).
    d        = dados de GET /orcamentos/{id}
    cli_data = dados de GET /clientes/{id}
    Retorna bytes do PDF ou None se reportlab indisponível.
    """
    if not _REPORTLAB_OK:
        return None

    # ── helpers ────────────────────────────────────────────────────
    def _brl(v):
        try:
            return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        except Exception:
            return "0,00"

    def _fdate(dt):
        if not dt: return ""
        p = str(dt).split("-")
        return f"{p[2]}/{p[1]}/{p[0]}" if len(p) == 3 else str(dt)

    # ── paleta corporativa Locvix (laranja + azul escuro) ───────────
    NAVY     = _rlcolors.HexColor("#1e3a5f")   # azul escuro — cabeçalho principal
    NAVY_T   = _rlcolors.HexColor("#e87722")   # laranja Locvix — col headers / info bar
    NAVY_CUI = _rlcolors.HexColor("#f59e42")   # laranja claro — "aos cuidados"
    ROW_ALT  = _rlcolors.HexColor("#fff7ed")   # laranja pastel muito claro — linha par
    ROW_TOT  = _rlcolors.HexColor("#fed7aa")   # laranja pastel — linha TOTAL
    BDR      = _rlcolors.HexColor("#fdba74")   # laranja bordas
    PRETO    = _rlcolors.black
    BRANCO   = _rlcolors.white
    CINZA_L  = _rlcolors.HexColor("#475569")   # texto label (cinza escuro)

    CW = _A4[0] - 30*_mm   # largura útil do conteúdo

    # ── estilos de parágrafo ───────────────────────────────────────
    # Cabeçalho da empresa
    st_emp  = _PS("em", fontSize=15, fontName="Helvetica-Bold", textColor=BRANCO, alignment=_TAC, spaceAfter=8)
    st_esub = _PS("es", fontSize=8,  fontName="Helvetica",      textColor=BRANCO, alignment=_TAC, leading=11, spaceBefore=4)
    st_pnum = _PS("pn", fontSize=11, fontName="Helvetica-Bold", textColor=BRANCO, alignment=_TAC, spaceBefore=4)
    # Barra de info (preto em laranja)
    st_bl   = _PS("bl", fontSize=8,  fontName="Helvetica-Bold", textColor=PRETO, spaceAfter=1)
    st_bv   = _PS("bv", fontSize=8,  fontName="Helvetica",      textColor=PRETO)
    # Seção (preto em laranja)
    st_sec  = _PS("sc", fontSize=9,  fontName="Helvetica-Bold", textColor=PRETO)
    # Campos do cliente
    st_lbl  = _PS("lb", fontSize=8,  fontName="Helvetica-Bold", textColor=CINZA_L)
    st_val  = _PS("vl", fontSize=8,  fontName="Helvetica",      textColor=PRETO)
    # Cabeçalho de colunas (preto em laranja)
    st_th   = _PS("th", fontSize=8,  fontName="Helvetica-Bold", textColor=PRETO)
    st_th_c = _PS("tc", fontSize=8,  fontName="Helvetica-Bold", textColor=PRETO,  alignment=_TAC)
    st_th_r = _PS("tr", fontSize=8,  fontName="Helvetica-Bold", textColor=PRETO,  alignment=_TAR)
    # Dados das linhas
    st_td   = _PS("td", fontSize=8,  fontName="Helvetica",      textColor=PRETO)
    st_td_c = _PS("dc", fontSize=8,  fontName="Helvetica",      textColor=PRETO,  alignment=_TAC)
    st_td_r = _PS("dr", fontSize=8,  fontName="Helvetica",      textColor=PRETO,  alignment=_TAR)
    # Linha TOTAL nas tabelas (preto no fundo laranja pastel)
    st_tt   = _PS("tt", fontSize=8,  fontName="Helvetica-Bold", textColor=PRETO,  alignment=_TAR)
    st_tr   = _PS("tor",fontSize=8,  fontName="Helvetica-Bold", textColor=PRETO,  alignment=_TAR)
    # Subtotais (caixa de totais)
    st_sl   = _PS("sl", fontSize=9,  fontName="Helvetica",      textColor=PRETO,  alignment=_TAR)
    st_sv   = _PS("sv", fontSize=9,  fontName="Helvetica",      textColor=PRETO,  alignment=_TAR)
    # Linha TOTAL GERAL (azul escuro, branco) — mantém destaque forte
    st_fl   = _PS("fl", fontSize=10, fontName="Helvetica-Bold", textColor=BRANCO, alignment=_TAR)
    st_fv   = _PS("fv", fontSize=11, fontName="Helvetica-Bold", textColor=BRANCO, alignment=_TAR)
    # Termos
    st_intr = _PS("in", fontSize=9,  fontName="Helvetica",      textColor=PRETO,  leading=13)

    # ── helpers de estilo de tabela ────────────────────────────────
    def _bst():
        return [
            ("BOX",           (0,0),(-1,-1), 0.4, BDR),
            ("INNERGRID",     (0,0),(-1,-1), 0.25, BDR),
            ("TOPPADDING",    (0,0),(-1,-1), 4),
            ("BOTTOMPADDING", (0,0),(-1,-1), 4),
            ("LEFTPADDING",   (0,0),(-1,-1), 5),
            ("RIGHTPADDING",  (0,0),(-1,-1), 5),
        ]

    def _sec_hdr(txt):
        t = _RLTable([[_Para(txt, st_sec)]], colWidths=[CW])
        t.setStyle(_RLTableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), NAVY_T),
            ("BOX",           (0,0),(-1,-1), 0, NAVY_T),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
            ("LEFTPADDING",   (0,0),(-1,-1), 7),
            ("RIGHTPADDING",  (0,0),(-1,-1), 7),
        ]))
        return t

    # ── buffer / documento ─────────────────────────────────────────
    buf = _BytesIO()
    doc = _SDT(buf, pagesize=_A4,
               leftMargin=15*_mm, rightMargin=15*_mm,
               topMargin=12*_mm, bottomMargin=12*_mm)
    els = []

    # ══ CABEÇALHO DA EMPRESA ══════════════════════════════════════
    emp = _get_empresa_info(d.get("loja_id") or d.get("id_loja"))
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ARTE locvix.png")
    _sub = (f"CNPJ: {emp['cnpj']}  \u2022  {emp['cidade']}/{emp['uf']}"
            f"  \u2022  {emp['telefone']}  \u2022  {emp['email']}")
    info_cell = [
        _Para(emp["nome"], st_emp),
        _Para(_sub, st_esub),
        _Para(f"PROPOSTA COMERCIAL  N\u00ba  {d.get('codigo','')}", st_pnum),
    ]
    if os.path.exists(logo_path):
        logo_w = 50*_mm
        hdr = _RLTable([[_RLImg(logo_path, width=44*_mm, height=27.5*_mm), info_cell]],
                       colWidths=[logo_w, CW - logo_w])
        hdr.setStyle(_RLTableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), NAVY),
            ("ALIGN",         (0,0),(0,0), "CENTER"),
            ("ALIGN",         (1,0),(1,0), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("BOX",           (0,0),(-1,-1), 0.5, NAVY),
            ("TOPPADDING",    (0,0),(-1,-1), 8),
            ("BOTTOMPADDING", (0,0),(-1,-1), 8),
            ("LEFTPADDING",   (0,0),(-1,-1), 6),
            ("RIGHTPADDING",  (0,0),(-1,-1), 6),
        ]))
    else:
        hdr = _RLTable([[info_cell]], colWidths=[CW])
        hdr.setStyle(_RLTableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), NAVY),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("BOX",           (0,0),(-1,-1), 0.5, NAVY),
            ("TOPPADDING",    (0,0),(-1,-1), 8),
            ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ]))
    els.append(hdr)
    els.append(_Spacer(1, 2*_mm))

    # ══ BARRA DE INFORMAÇÕES ══════════════════════════════════════
    data_orc = _fdate(d.get("data",""))
    cuidados = (d.get("aos_cuidados_de","") or "").strip()

    bar_data = [[
        [_Para("Data:", st_bl), _Para(data_orc, st_bv)],
    ]]
    bar_t = _RLTable(bar_data, colWidths=[CW])
    bar_t.setStyle(_RLTableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), NAVY_T),
        ("BOX",           (0,0),(-1,-1), 0.4, BDR),
        ("INNERGRID",     (0,0),(-1,-1), 0.3, BDR),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
        ("RIGHTPADDING",  (0,0),(-1,-1), 6),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
    ]))
    els.append(bar_t)

    if cuidados:
        cuid_t = _RLTable([[
            _Para("Aos cuidados de:", st_bl),
            _Para(cuidados, st_bv),
        ]], colWidths=[38*_mm, CW-38*_mm])
        cuid_t.setStyle(_RLTableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), NAVY_CUI),
            ("BOX",           (0,0),(-1,-1), 0.4, BDR),
            ("TOPPADDING",    (0,0),(-1,-1), 3),
            ("BOTTOMPADDING", (0,0),(-1,-1), 3),
            ("LEFTPADDING",   (0,0),(-1,-1), 7),
            ("RIGHTPADDING",  (0,0),(-1,-1), 6),
        ]))
        els.append(cuid_t)

    els.append(_Spacer(1, 3*_mm))

    # ══ DADOS DO CLIENTE ══════════════════════════════════════════
    els.append(_sec_hdr("\u258c  DADOS DO CLIENTE"))
    enderecos = cli_data.get("enderecos", [])
    end = enderecos[0].get("endereco", enderecos[0]) if enderecos else {}
    logr    = end.get("logradouro",""); numero = end.get("numero",""); bairro = end.get("bairro","")
    end_str = f"{logr}, {numero}" + (f" \u2014 {bairro}" if bairro else "")
    razao   = (cli_data.get("razao_social","") or cli_data.get("nome","") or
               d.get("nome_cliente","") or d.get("cliente",""))
    nome_fn = cli_data.get("nome_fantasia","") or cli_data.get("nome","") or razao
    cnpj    = (cli_data.get("cnpj","") or cli_data.get("cpf","") or
               cli_data.get("cpf_cnpj",""))
    tel     = cli_data.get("telefone","") or cli_data.get("fone","") or cli_data.get("celular","")
    email   = cli_data.get("email","") or cli_data.get("email_nfe","")
    cep     = end.get("cep","") or cli_data.get("cep","")
    nome_cid= end.get("nome_cidade","") or end.get("cidade","") or cli_data.get("cidade","")
    estado  = end.get("estado","") or end.get("uf","") or cli_data.get("uf","")
    cid_uf  = f"{nome_cid}/{estado}"
    cl = 26*_mm; cv = CW/2 - cl
    rows_cli = [
        [_Para("Razão Social:",  st_lbl), _Para(razao,             st_val), _Para("Nome Fantasia:", st_lbl), _Para(nome_fn,   st_val)],
        [_Para("CNPJ / CPF:",   st_lbl), _Para(cnpj,              st_val), _Para("Endereço:",      st_lbl), _Para(end_str,   st_val)],
        [_Para("CEP:",           st_lbl), _Para(cep,              st_val), _Para("Cidade / UF:",   st_lbl), _Para(cid_uf,    st_val)],
        [_Para("Telefone:",      st_lbl), _Para(tel,               st_val), _Para("E-mail:",         st_lbl), _Para(email,     st_val)],
    ]
    cli_t = _RLTable(rows_cli, colWidths=[cl, cv, cl, cv])
    cli_t.setStyle(_RLTableStyle(_bst() + [
        ("BACKGROUND", (0,1),(-1,1), ROW_ALT),
        ("BACKGROUND", (0,3),(-1,3), ROW_ALT),
    ]))
    els.append(cli_t)
    els.append(_Spacer(1, 3*_mm))

    # ══ SERVIÇOS ══════════════════════════════════════════════════
    servicos = d.get("servicos", [])
    if servicos:
        els.append(_sec_hdr("\u258c  SERVI\u00c7OS"))
        c0=16*_mm; c1=28*_mm; c3=18*_mm; c4=27*_mm; c5=29*_mm
        c2 = CW - c0 - c1 - c3 - c4 - c5
        cws = [c0, c1, c2, c3, c4, c5]
        rows = [[
            _Para("ITEM",        st_th_c),
            _Para("C\u00d3DIGO", st_th_c),
            _Para("DESCRI\u00c7\u00c3O", st_th),
            _Para("QTD.",        st_th_c),
            _Para("VR. UNIT.",   st_th_r),
            _Para("SUBTOTAL",    st_th_r),
        ]]
        ts_q = 0.0; ts_v = 0.0
        for i, sv in enumerate(servicos, 1):
            s  = sv.get("servico", sv)
            q  = float(s.get("quantidade", 0) or 0)
            vt = float(s.get("valor_total", 0) or 0)
            vu = float(s.get("valor_venda", 0) or 0)
            ts_q += q; ts_v += vt
            rows.append([
                _Para(str(i),                                    st_td_c),
                _Para(str(s.get("codigo_servico","") or ""),     st_td_c),
                _Para(s.get("nome_servico",""),                   st_td),
                _Para(_brl(q),                                   st_td_r),
                _Para(_brl(vu),                                  st_td_r),
                _Para(_brl(vt),                                  st_td_r),
            ])
        lr = len(rows)
        # TOTAL row: SPAN colunas 0-3 → "TOTAL" não quebra linha
        rows.append([_Para("TOTAL", st_tt), "", "", "", _Para(_brl(ts_q), st_tr), _Para(_brl(ts_v), st_tr)])
        tbl = _RLTable(rows, colWidths=cws)
        sty = _bst() + [
            ("BACKGROUND", (0,0),  (-1,0),   NAVY_T),
            ("BACKGROUND", (0,lr), (-1,lr),  ROW_TOT),
            ("SPAN",       (0,lr), (3,lr)),
            ("ALIGN",      (0,lr), (3,lr),   "RIGHT"),
            ("BOX",        (0,lr), (-1,lr),  0.5, _rlcolors.HexColor("#93c5fd")),
        ]
        for ri in range(1, lr):
            if ri % 2 == 0:
                sty.append(("BACKGROUND", (0,ri),(-1,ri), ROW_ALT))
        tbl.setStyle(_RLTableStyle(sty))
        els.append(tbl)
        st_obs = _PS("ob", fontSize=8, fontName="Helvetica-Oblique", textColor=CINZA_L)
        els.append(_Para("Obs.: Mínimo 10 horas/dia", st_obs))
        els.append(_Spacer(1, 3*_mm))

    # ══ PRODUTOS ══════════════════════════════════════════════════
    produtos = d.get("produtos", [])
    if produtos:
        els.append(_sec_hdr("\u258c  PRODUTOS"))
        c0=16*_mm; c1=22*_mm; cu=14*_mm; c3=18*_mm; c4=27*_mm; c5=29*_mm
        c2 = CW - c0 - c1 - cu - c3 - c4 - c5
        cwp = [c0, c1, c2, cu, c3, c4, c5]
        rows = [[
            _Para("ITEM",        st_th_c),
            _Para("C\u00d3DIGO", st_th_c),
            _Para("DESCRI\u00c7\u00c3O", st_th),
            _Para("UND.",        st_th_c),
            _Para("QTD.",        st_th_c),
            _Para("VR. UNIT.",   st_th_r),
            _Para("SUBTOTAL",    st_th_r),
        ]]
        tp_q = 0.0; tp_v = 0.0
        for i, pv in enumerate(produtos, 1):
            p  = pv.get("produto", pv)
            q  = float(p.get("quantidade", 0) or 0)
            vt = float(p.get("valor_total", 0) or 0)
            vu = float(p.get("valor_venda", 0) or 0)
            tp_q += q; tp_v += vt
            rows.append([
                _Para(str(i),                                    st_td_c),
                _Para(str(p.get("codigo_produto","") or ""),     st_td_c),
                _Para(p.get("nome_produto",""),                   st_td),
                _Para(str(p.get("sigla_unidade","") or ""),      st_td_c),
                _Para(_brl(q),                                   st_td_r),
                _Para(_brl(vu),                                  st_td_r),
                _Para(_brl(vt),                                  st_td_r),
            ])
        lr = len(rows)
        # TOTAL row: SPAN colunas 0-4
        rows.append([_Para("TOTAL", st_tt), "", "", "", "", _Para(_brl(tp_q), st_tr), _Para(_brl(tp_v), st_tr)])
        tbl = _RLTable(rows, colWidths=cwp)
        sty = _bst() + [
            ("BACKGROUND", (0,0),  (-1,0),  NAVY_T),
            ("BACKGROUND", (0,lr), (-1,lr), ROW_TOT),
            ("SPAN",       (0,lr), (4,lr)),
            ("ALIGN",      (0,lr), (4,lr),  "RIGHT"),
            ("BOX",        (0,lr), (-1,lr), 0.5, _rlcolors.HexColor("#93c5fd")),
        ]
        for ri in range(1, lr):
            if ri % 2 == 0:
                sty.append(("BACKGROUND", (0,ri),(-1,ri), ROW_ALT))
        tbl.setStyle(_RLTableStyle(sty))
        els.append(tbl)
        els.append(_Spacer(1, 3*_mm))

    # ══ CAIXA DE TOTAIS ═══════════════════════════════════════════
    vt_g  = float(d.get("valor_total",    0) or 0)
    vs_g  = float(d.get("valor_servicos", 0) or 0)
    vp_g  = float(d.get("valor_produtos", 0) or 0)
    vf_g  = float(d.get("valor_frete",    0) or 0)
    rows_t = []
    if vp_g > 0: rows_t.append([_Para("PRODUTOS",  st_sl), _Para(f"R$ {_brl(vp_g)}", st_sv)])
    if vs_g > 0: rows_t.append([_Para("SERVI\u00c7OS",  st_sl), _Para(f"R$ {_brl(vs_g)}", st_sv)])
    if vf_g > 0: rows_t.append([_Para("FRETE",     st_sl), _Para(f"R$ {_brl(vf_g)}", st_sv)])
    lr_t = len(rows_t)
    rows_t.append([_Para("TOTAL GERAL", st_fl), _Para(f"R$ {_brl(vt_g)}", st_fv)])
    tot_t = _RLTable(rows_t, colWidths=[44*_mm, 38*_mm], hAlign="RIGHT")
    tot_t.setStyle(_RLTableStyle([
        ("BOX",           (0,0), (-1,-1),    0.5, BDR),
        ("INNERGRID",     (0,0), (-1,-1),    0.3, BDR),
        ("TOPPADDING",    (0,0), (-1,-1),    5),
        ("BOTTOMPADDING", (0,0), (-1,-1),    5),
        ("LEFTPADDING",   (0,0), (-1,-1),    8),
        ("RIGHTPADDING",  (0,0), (-1,-1),    8),
        ("ALIGN",         (0,0), (-1,-1),    "RIGHT"),
        ("BACKGROUND",    (0,lr_t),(-1,lr_t), NAVY),
        ("BOX",           (0,lr_t),(-1,lr_t), 0.5, NAVY),
    ]))
    els.append(tot_t)
    els.append(_Spacer(1, 4*_mm))

    # ══ DADOS DO PAGAMENTO ════════════════════════════════════════
    pagamentos = d.get("pagamentos", [])
    els.append(_sec_hdr("\u258c  DADOS DO PAGAMENTO"))
    c_venc=30*_mm; c_vlr=32*_mm; c_obs=45*_mm
    c_forma = CW - c_venc - c_vlr - c_obs
    rows_pg = [[
        _Para("VENCIMENTO",          st_th_c),
        _Para("VALOR",               st_th_r),
        _Para("FORMA DE PAGAMENTO",  st_th),
        _Para("OBSERVAÇÃO",          st_th),
    ]]
    if pagamentos:
        for pv in pagamentos:
            pg = pv.get("pagamento", pv)
            rows_pg.append([
                _Para(_fdate(pg.get("data_vencimento","")),                 st_td_c),
                _Para(_brl(pg.get("valor",0)),                              st_td_r),
                _Para(str(pg.get("nome_forma_pagamento","") or ""),         st_td),
                _Para("Sujeito a alteração da data de vencimento e dos valores.",         st_td),
            ])
    else:
        # linha padrão quando a API não retorna parcelas
        total_geral = sum(
            float(sv.get("servico", sv).get("valor_total", 0) or 0)
            for sv in d.get("servicos", [])
        ) + sum(
            float(pv2.get("produto", pv2).get("valor_total", 0) or 0)
            for pv2 in d.get("produtos", [])
        )
        rows_pg.append([
            _Para("",                                              st_td_c),
            _Para(_brl(total_geral) if total_geral else "",        st_td_r),
            _Para("",                                              st_td),
            _Para("Sujeito a alteração da data de vencimento e dos valores.",    st_td),
        ])
    tpg = _RLTable(rows_pg, colWidths=[c_venc, c_vlr, c_forma, c_obs])
    sty_pg = _bst() + [("BACKGROUND", (0,0),(-1,0), NAVY_T)]
    for ri in range(1, len(rows_pg)):
        if ri % 2 == 0:
            sty_pg.append(("BACKGROUND", (0,ri),(-1,ri), ROW_ALT))
    tpg.setStyle(_RLTableStyle(sty_pg))
    els.append(tpg)
    els.append(_Spacer(1, 4*_mm))

    # ══ TERMOS E CONDIÇÕES ════════════════════════════════════════
    intro = (d.get("introducao","") or "").strip()
    if intro:
        import re as _re_pdf
        _SEC_HDRS = {
            "PERÍODO DE UTILIZAÇÃO",
            "1 PERÍODO DE UTILIZAÇÃO",
            "JORNADA DE TRABALHO E APONTAMENTO DE HORAS",
            "OBRIGAÇÕES DA CONTRATADA",
            "OBRIGAÇÕES DA CONTRATANTE",
            "IMPOSTOS",
            "MEDIÇÃO, FATURAMENTO E PAGAMENTO",
            "DISPOSIÇÕES FINAIS",
        }
        st_tc_title = _PS("tct", fontSize=10, fontName="Helvetica-Bold", textColor=PRETO, spaceBefore=4, spaceAfter=4)
        els.append(_Para("TERMOS E CONDIÇÕES", st_tc_title))
        els.append(_Spacer(1, 2*_mm))
        pending = []
        for raw in intro.split("\n"):
            # remove leading numbering like "2 " / "2. " / "2.1 "
            clean = _re_pdf.sub(r'^\d+[\.\s]+', '', raw.strip().upper()).strip()
            if clean in _SEC_HDRS:
                if pending:
                    block = "<br/>".join(l.replace("\t","&nbsp;&nbsp;&nbsp;&nbsp;") for l in pending)
                    els.append(_Para(block, st_intr))
                    els.append(_Spacer(1, 1*_mm))
                    pending = []
                els.append(_Spacer(1, 2*_mm))
                els.append(_sec_hdr(f"\u258c  {clean}"))
            else:
                pending.append(raw)
        if pending:
            block = "<br/>".join(l.replace("\t","&nbsp;&nbsp;&nbsp;&nbsp;") for l in pending)
            els.append(_Para(block, st_intr))
        els.append(_Spacer(1, 4*_mm))

    # ── Assinatura do cliente ──────────────────────────────────────
    st_ass = _PS("as", fontSize=9, fontName="Helvetica-Bold", textColor=PRETO,
                 alignment=_TAC)
    sig_w = 90*_mm
    sig_tbl = _RLTable(
        [[_Para("Assinatura do cliente", st_ass)]],
        colWidths=[sig_w]
    )
    sig_tbl.setStyle(_RLTableStyle([
        ("LINEABOVE",   (0,0), (-1,0), 0.8, PRETO),
        ("TOPPADDING",  (0,0), (-1,0), 4),
        ("BOTTOMPADDING",(0,0),(-1,0), 2),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
    ]))
    sig_outer = _RLTable(
        [[sig_tbl]],
        colWidths=[CW],
        style=[("ALIGN",(0,0),(-1,-1),"CENTER"),
               ("LEFTPADDING",(0,0),(-1,-1),0),
               ("RIGHTPADDING",(0,0),(-1,-1),0)]
    )
    els.append(_KeepTogether([_Spacer(1, 4*_mm), sig_outer]))

    doc.build(els)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════
#  LEITURA DE PROPOSTAS (PDF) — PERDIDAS E FECHADAS
# ══════════════════════════════════════════════════════════════════
def buscar_centros_custo() -> list[dict]:
    """Retorna lista de centros de custo cadastrados no GestãoClick."""
    try:
        raw = _paginar_lojas("centros_custos")
        return [{"id": str(c.get("id","")), "nome": c.get("nome","") or c.get("descricao","")} for c in raw]
    except Exception as e:
        print(f"  [AVISO] centros_custos: {e}")
        return []


def buscar_situacoes_orcamento() -> list[dict]:
    """Retorna situações de orçamento cadastradas."""
    try:
        raw = _paginar_lojas("situacoes_orcamentos")
        return [{"id": str(s.get("id","")), "nome": s.get("nome","") or s.get("descricao","")} for s in raw]
    except Exception as e:
        print(f"  [AVISO] situacoes_orcamentos: {e}")
        return []


def buscar_formas_pagamento() -> list[dict]:
    """Retorna formas de pagamento cadastradas."""
    try:
        raw = _paginar_lojas("formas_pagamentos")
        return [{"id": str(f.get("id","")), "nome": f.get("nome","") or f.get("descricao","")} for f in raw]
    except Exception as e:
        print(f"  [AVISO] formas_pagamentos: {e}")
        return []


def buscar_servicos() -> list[dict]:
    """Retorna serviços cadastrados no GestãoClick."""
    try:
        raw = _paginar_lojas("servicos")
        return [
            {
                "id":    str(s.get("id","")),
                "nome":  s.get("nome","") or s.get("descricao",""),
                "preco": float(s.get("valor_venda") or s.get("preco") or 0),
                "unidade": s.get("unidade") or s.get("sigla_unidade") or "H",
            }
            for s in raw
        ]
    except Exception as e:
        print(f"  [AVISO] servicos: {e}")
        return []


def buscar_vendedores() -> list[dict]:
    """Retorna lista de vendedores/usuários de AMBAS as lojas (sem duplicatas)."""
    try:
        gck = _gck()
        # Busca em cada loja separadamente para garantir todos os usuários
        todos: dict[str, dict] = {}
        for loja in [LOJA_GJ_ID, LOJA_WA_ID, None]:
            params = {"loja_id": loja} if loja else {}
            raw = gck.paginar("usuarios", params)
            for u in raw:
                uid = str(u.get("id", ""))
                nome = u.get("nome") or u.get("name") or ""
                if uid and nome and uid not in todos:
                    todos[uid] = {"id": uid, "nome": nome}
        return list(todos.values())
    except Exception as e:
        print(f"  [AVISO] usuarios: {e}")
        return []


def criar_orcamento_api(payload: dict, loja_id: str | None = None) -> dict:
    """
    Envia POST /orcamentos para o GestãoClick e retorna:
    {
        "ok":      bool,
        "id":      str,
        "codigo":  str,
        "msg":     str,
        "pdf_bytes": bytes | None,   # PDF gerado com layout corporativo Locvix
    }

    payload deve seguir a estrutura do APIB:
    {
        "cliente_id": int,
        "vendedor_id": str (opcional),
        "centro_custo_id": int (opcional),
        "situacao_id": str,
        "data": "YYYY-MM-DD",
        "validade": "30 dias",
        "aos_cuidados_de": str (opcional),
        "introducao": str (opcional),
        "observacoes": str (opcional),
        "servicos": [
            {"servico": {"id": str, "nome_servico": str, "quantidade": str,
                         "valor_venda": str, "detalhes": str,
                         "desconto_valor": "0", "desconto_porcentagem": "0"}}
        ],
        "condicao_pagamento": "a_vista" | "parcelado",
        "forma_pagamento_id": str (se parcelado),
        "numero_parcelas": int (se parcelado),
        "data_primeira_parcela": "YYYY-MM-DD" (se parcelado),
    }
    """
    global LOJA_FILTRO
    _loja_orig = LOJA_FILTRO
    if loja_id:
        LOJA_FILTRO = loja_id
    try:
        gck = _gck()
        resp = gck.post("orcamentos", payload)
    finally:
        LOJA_FILTRO = _loja_orig

    if not resp:
        return {"ok": False, "id": "", "codigo": "", "msg": "Sem resposta da API", "pdf_bytes": None}

    if resp.get("code") not in (200, 201) and resp.get("status") != "success":
        erros = resp.get("errors") or resp.get("message") or resp.get("data") or str(resp)
        return {"ok": False, "id": "", "codigo": "", "msg": str(erros), "pdf_bytes": None}

    det = resp.get("data", {}) or {}
    orc_id  = str(det.get("id", ""))
    orc_cod = str(det.get("codigo", ""))

    # Gera PDF com os dados retornados pela API
    pdf_bytes = None
    try:
        if loja_id:
            det["loja_id"] = str(loja_id)
        cli_id = str(det.get("cliente_id") or payload.get("cliente_id") or "")
        cli_data: dict = {}
        if cli_id:
            resp_cli = gck.get(f"clientes/{cli_id}")
            cli_data = (resp_cli.get("data", {}) if resp_cli else {}) or {}
        if not cli_data.get("nome") and not cli_data.get("razao_social"):
            cli_data["razao_social"] = det.get("nome_cliente", "")
        pdf_bytes = _gerar_pdf_orc_bytes(det, cli_data)
    except Exception as e:
        print(f"  [AVISO] PDF novo orçamento: {e}")

    return {
        "ok":        True,
        "id":        orc_id,
        "codigo":    orc_cod,
        "msg":       f"Orçamento nº {orc_cod} criado com sucesso!",
        "pdf_bytes": pdf_bytes,
    }


def deletar_orcamento_api(orc_id: str, loja_id: str | None = None) -> dict:
    """
    Envia DELETE /orcamentos/{id} para o GestãoClick.
    Retorna {"ok": bool, "msg": str}.
    """
    global LOJA_FILTRO
    _loja_orig = LOJA_FILTRO
    if loja_id:
        LOJA_FILTRO = loja_id
    try:
        gck = _gck()
        resp = gck.delete(f"orcamentos/{orc_id}")
    finally:
        LOJA_FILTRO = _loja_orig

    if resp is None:
        return {"ok": False, "msg": "Sem resposta da API"}
    code = resp.get("code", 0)
    if code in (200, 204) or resp.get("status") == "success":
        return {"ok": True, "msg": f"Orçamento {orc_id} excluído com sucesso."}
    erros = resp.get("errors") or resp.get("message") or str(resp)
    return {"ok": False, "msg": str(erros)}


def alterar_orcamento_api(orc_id: str, payload: dict, loja_id: str | None = None) -> dict:
    """
    Envia PUT /orcamentos/{id} para o GestãoClick.
    Retorna {"ok": bool, "id": str, "codigo": str, "msg": str, "pdf_bytes": bytes|None}
    """
    global LOJA_FILTRO
    _loja_orig = LOJA_FILTRO
    if loja_id:
        LOJA_FILTRO = loja_id
    try:
        gck = _gck()
        resp = gck.put(f"orcamentos/{orc_id}", payload)
    finally:
        LOJA_FILTRO = _loja_orig

    if not resp:
        return {"ok": False, "id": orc_id, "codigo": "", "msg": "Sem resposta da API", "pdf_bytes": None}

    if resp.get("code") not in (200, 201) and resp.get("status") not in ("success", "ok"):
        erros = resp.get("errors") or resp.get("message") or resp.get("data") or str(resp)
        return {"ok": False, "id": orc_id, "codigo": "", "msg": str(erros), "pdf_bytes": None}

    det = resp.get("data", {}) or {}
    orc_cod = str(det.get("codigo", ""))

    # Gera PDF com os dados retornados pela API
    pdf_bytes = None
    try:
        if loja_id:
            det.setdefault("loja_id", str(loja_id))
        cli_id = str(det.get("cliente_id") or payload.get("cliente_id") or "")
        cli_data: dict = {}
        if cli_id:
            resp_cli = gck.get(f"clientes/{cli_id}")
            cli_data = (resp_cli.get("data", {}) if resp_cli else {}) or {}
        if not cli_data.get("nome") and not cli_data.get("razao_social"):
            cli_data["razao_social"] = det.get("nome_cliente", "")
        pdf_bytes = _gerar_pdf_orc_bytes(det, cli_data)
    except Exception as e:
        print(f"  [AVISO] PDF alterar orçamento: {e}")

    return {
        "ok":        True,
        "id":        str(det.get("id", orc_id)),
        "codigo":    orc_cod,
        "msg":       f"Orçamento nº {orc_cod} alterado com sucesso!",
        "pdf_bytes": pdf_bytes,
    }


def buscar_orcamento_por_id(orc_id: str, loja_id: str | None = None) -> dict:
    """
    Busca dados resumidos de um orçamento pelo código/número visível ou ID interno.
    Tenta 3 estratégias: GET direto por ID, lista filtrada por numero, lista por codigo.
    Retorna {"ok": bool, "id": str, "codigo": str, "cliente": str, "valor": float,
             "data": str, "situacao": str, "msg": str}
    """
    def _parse(det: dict) -> dict:
        dt = det.get("data", "") or ""
        if dt and "-" in dt:
            p = dt.split("-")
            dt = f"{p[2]}/{p[1]}/{p[0]}" if len(p) == 3 else dt
        return {
            "ok":       True,
            "id":       str(det.get("id", orc_id)),
            "codigo":   str(det.get("codigo", "")),
            "cliente":  det.get("nome_cliente") or det.get("cliente") or "—",
            "valor":    round(float(det.get("valor_total", 0) or 0), 2),
            "data":     dt,
            "situacao": det.get("nome_situacao") or det.get("situacao") or "—",
            "msg":      "ok",
        }

    _lid = {"loja_id": loja_id} if loja_id else {}

    def _get_lista(extra_params: dict) -> list:
        """GET /orcamentos com limite=1 — muito mais rápido que paginar."""
        resp = _gck().get("orcamentos", {**_lid, "limite": 10, **extra_params})
        if not resp:
            return []
        items = resp.get("data", [])
        return items if isinstance(items, list) else []

    def _best(lista: list) -> dict | None:
        return next(
            (o for o in lista if str(o.get("codigo","")) == str(orc_id)
             or str(o.get("numero","")) == str(orc_id)
             or str(o.get("id","")) == str(orc_id)),
            lista[0] if lista else None,
        )

    # 1) Filtro por codigo (campo mais específico)
    lista = _get_lista({"codigo": orc_id})
    hit = _best(lista)
    if hit:
        return _parse(hit)

    # 2) Filtro por numero
    lista2 = _get_lista({"numero": orc_id})
    hit2 = _best(lista2)
    if hit2:
        return _parse(hit2)

    # 3) GET direto por ID interno (fallback)
    resp = _gck().get(f"orcamentos/{orc_id}", _lid or None)
    if resp:
        det = resp.get("data", {}) or {}
        if det:
            return _parse(det)

    return {"ok": False, "msg": f"Orçamento '{orc_id}' não encontrado."}


def buscar_orcamentos():
    """
    Busca orçamentos da API GestãoClick (GET /orcamentos).
    Para cada orçamento faz GET /orcamentos/{id} e GET /clientes/{id}
    para gerar o PDF e incluí-lo como base64 no campo 'pdf_b64'.
    Retorna lista com: codigo, data, cliente, vendedor, situacao, valor,
    centro_custo, loja, id, pdf_b64.
    """
    _prog(0.69, "Buscando orçamentos GestãoClick...")
    raw = _paginar_lojas("orcamentos")
    registros: list[dict] = []

    def _fetch_one(o: dict) -> dict:
        dt = o.get("data", "") or ""
        if dt and "-" in dt:
            p = dt.split("-")
            dt = f"{p[2]}/{p[1]}/{p[0]}" if len(p) == 3 else dt
        sit = (o.get("nome_situacao") or "Em aberto").strip()
        rec = {
            "id":           o.get("id", ""),
            "codigo":       o.get("codigo", ""),
            "data":         dt,
            "situacao":     sit,
            "cliente":      o.get("nome_cliente", ""),
            "vendedor":     o.get("nome_vendedor", ""),
            "centro_custo": o.get("nome_centro_custo", ""),
            "loja":         o.get("nome_loja", ""),
            "valor":        round(float(o.get("valor_total", 0) or 0), 2),
            "pdf_b64":      "",
        }
        # Gera PDF com dados detalhados
        try:
            orc_id = str(o.get("id",""))
            cli_id = str(o.get("cliente_id",""))
            gck = _gck()
            resp_det = gck.get(f"orcamentos/{orc_id}")
            det = (resp_det.get("data", {}) if resp_det else {}) or {}
            # garante que loja_id chegue ao PDF — usa o registro da lista como fonte
            _nome_loja_map = {"w & a": LOJA_WA_ID, "w&a": LOJA_WA_ID,
                              "g & j": LOJA_GJ_ID, "g&j": LOJA_GJ_ID}
            _nome_loja = (o.get("nome_loja") or "").strip().lower()
            loja_id_orc = (o.get("loja_id") or o.get("id_loja") or
                           (o.get("loja", {}) or {}).get("id") or
                           _nome_loja_map.get(_nome_loja) or LOJA_GJ_ID)
            det.setdefault("loja_id", str(loja_id_orc))
            cli_data: dict = {}
            if cli_id:
                resp_cli = gck.get(f"clientes/{cli_id}")
                cli_data = (resp_cli.get("data", {}) if resp_cli else {}) or {}
            # garante nome do cliente mesmo sem detalhe do cliente
            if not cli_data.get("nome") and not cli_data.get("razao_social"):
                cli_data["razao_social"] = o.get("nome_cliente", "")
            pdf_bytes = _gerar_pdf_orc_bytes(det, cli_data)
            if pdf_bytes:
                rec["pdf_b64"] = base64.b64encode(pdf_bytes).decode("ascii")
        except Exception as e:
            print(f"  ⚠ PDF orçamento {rec['codigo']}: {e}")
        return rec

    # Gera PDFs em paralelo (max 4 threads para não sobrecarregar a API)
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as ex:
        futs = [ex.submit(_fetch_one, o) for o in raw]
        registros = [f.result() for f in concurrent.futures.as_completed(futs)]

    # Reordena pelo código (a conclusão em paralelo pode ficar fora de ordem)
    registros.sort(key=lambda r: str(r.get("codigo","")).zfill(10))

    conc = sum(1 for r in registros if r["situacao"] == "Concretizado")
    print(f"  ✔ {len(registros)} orçamentos ({conc} concretizados)")
    return registros


# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — BOLETINS DE MEDIÇÃO
# ══════════════════════════════════════════════════════════════════
def buscar_medicoes() -> list[dict]:
    """
    Lê PDFs da pasta 'MEDIÇÃO' e extrai dados dos Boletins de Medição:
    número, equipamento, período, cliente, horas diurno/noturno/extra,
    valores financeiros por turno, desmobilização e total a pagar.
    """
    import os as _os
    import re as _re

    try:
        import pdfplumber as _pl
    except ImportError:
        print("  [AVISO] pdfplumber não instalado — módulo Medição indisponível")
        return []

    _prog(0.695, "Lendo boletins de medição PDF...")
    base_dir = _os.path.dirname(_os.path.abspath(__file__))
    pasta = _os.path.join(base_dir, "MEDIÇÃO")

    if not _os.path.exists(pasta):
        print(f"  [AVISO] Pasta MEDIÇÃO não encontrada: {pasta}")
        return []

    def _parse_val(txt: str) -> float:
        """Converte string com R$ e espaços para float."""
        if not txt:
            return 0.0
        limpo = _re.sub(r'[R$\s\.]', '', str(txt)).replace(',', '.')
        try:
            return float(limpo)
        except Exception:
            return 0.0

    registros: list[dict] = []

    for nome_arq in sorted(_os.listdir(pasta)):
        if not nome_arq.lower().endswith(".pdf"):
            continue
        caminho = _os.path.join(pasta, nome_arq)
        try:
            with _pl.open(caminho) as pdf:
                page = pdf.pages[0]
                tabelas = page.extract_tables()
        except Exception as e:
            print(f"  [AVISO] Erro ao ler {nome_arq}: {e}")
            continue

        if not tabelas:
            continue

        tbl = tabelas[0]  # O PDF inteiro é uma tabela grande

        # ── Número do boletim (linha 0, col 15)
        numero = ""
        for row in tbl[:3]:
            if row and len(row) > 15 and row[15]:
                val = str(row[15]).strip()
                if _re.match(r'\d+', val):
                    numero = val
                    break

        # ── Equipamento e Período (linha 1, col 1, dentro do texto multi-linha)
        equipamento = ""
        periodo = ""
        for row in tbl[:4]:
            if row and row[1]:
                celula = str(row[1])
                m = _re.search(r'EQUIPAMENTO:\s*(.+)', celula)
                if m:
                    equipamento = m.group(1).strip()
                m = _re.search(r'PERÍODO\s*:\s*(.+)', celula)
                if m:
                    periodo = m.group(1).strip()

        # ── Cliente: extraído do nome do arquivo
        # Padrão: "MEDIÇÃO [TIPO] CLIENTE - EQUIPAMENTO DATA..."
        # Ex:     "MEDIÇÃO DEFINITIVA REDE - EMPILHADEIRA 10T ..."
        cliente = ""
        nome_base = nome_arq.replace('.pdf', '')
        partes = _re.split(r'\s+-\s+', nome_base, maxsplit=1)
        if partes:
            # Remove o prefixo "MEDIÇÃO" e palavras modificadoras (DEFINITIVA, PARCIAL etc)
            _mods = {'DEFINITIVA','DEFINITIVO','PARCIAL','FINAL','COMPLEMENTAR',
                     'INICIAL','REVISAO','REVISÃO','LOCACAO','LOCAÇÃO'}
            palavras = partes[0].split()
            cli_partes = []
            for pw in palavras:
                pu = pw.upper().replace('Ç','C').replace('Ã','A').replace('Õ','O')
                if not cli_partes and (pu.startswith('MEDIC') or pu in _mods):
                    continue
                cli_partes.append(pw)
            cliente = ' '.join(cli_partes).strip()

        # ── Lançamentos diários: cols 1-7 diurno, 9-15 noturno
        dias_diurno: list[dict] = []
        dias_noturno: list[dict] = []
        horas_diurno_total = 0.0
        horas_noturno_total = 0.0
        horas_extra_diurno = 0.0
        horas_extra_noturno = 0.0

        DIAS_PT = {'SEG','TER','QUA','QUI','SEX','SÁB','SAB','DOM'}
        for row in tbl:
            if not row or len(row) < 16:
                continue
            data_d = str(row[1] or '').strip()
            sem_d  = str(row[2] or '').strip().upper()
            # Linha de dado diurno: col 1 tem data dd/mm/yyyy, col 2 tem dia da semana
            if _re.match(r'\d{2}/\d{2}/\d{4}', data_d) and sem_d in DIAS_PT:
                h_d = _parse_val(str(row[7] or ''))
                e_d = _parse_val(str(row[8] or ''))
                dias_diurno.append({
                    "data": data_d, "semana": sem_d,
                    "entrada": str(row[3] or '').strip(),
                    "saida_almoco": str(row[4] or '').strip(),
                    "retorno_almoco": str(row[5] or '').strip(),
                    "saida": str(row[6] or '').strip(),
                    "horas": h_d, "extra": e_d,
                })
                horas_diurno_total += h_d
                horas_extra_diurno += e_d
                # Turno noturno (colunas 9-15, mesma linha)
                if len(row) > 15:
                    h_n = _parse_val(str(row[15] or ''))
                    e_n = _parse_val(str(row[16] or '')) if len(row) > 16 else 0.0
                    if h_n > 0:
                        dias_noturno.append({
                            "data": str(row[9] or data_d).strip(), "semana": str(row[10] or sem_d).strip().upper(),
                            "entrada": str(row[11] or '').strip(),
                            "saida_jantar": str(row[12] or '').strip(),
                            "retorno_jantar": str(row[13] or '').strip(),
                            "saida": str(row[14] or '').strip(),
                            "horas": h_n, "extra": e_n,
                        })
                        horas_noturno_total += h_n
                        horas_extra_noturno += e_n

        # ── Totais de horas da linha de TOTAL HORAS MÊS
        for row in tbl:
            if not row:
                continue
            celula = str(row[1] or '')
            if 'TOTAL HORAS MÊS' in celula and 'TURNO DIA' in celula:
                if row[7]:
                    horas_diurno_total = _parse_val(str(row[7]))
                if row[8]:
                    horas_extra_diurno = _parse_val(str(row[8]))
                if row[15]:
                    horas_noturno_total = _parse_val(str(row[15]))
                if len(row) > 16 and row[16]:
                    horas_extra_noturno = _parse_val(str(row[16]))
                break

        horas_extra_total = horas_extra_diurno + horas_extra_noturno

        # ── Valores financeiros
        valor_hora_dia = 0.0
        valor_hora_not = 0.0
        valor_hora_ext = 0.0
        valor_diurno = 0.0
        valor_noturno = 0.0
        valor_extra = 0.0
        valor_desmobi = 0.0
        total_medicao = 0.0
        total_pagar = 0.0

        for row in tbl:
            if not row:
                continue
            col1 = str(row[1] or '')
            col15 = str(row[15] or '') if len(row) > 15 else ''
            col11 = str(row[11] or '') if len(row) > 11 else ''
            if 'TOTAL DE HORAS/MÊS - TURNO DIURNO' in col1:
                valor_hora_dia = _parse_val(col11)
                valor_diurno   = _parse_val(col15)
            elif 'TOTAL DE HORAS/MÊS - TURNO NOTURNO' in col1:
                valor_hora_not = _parse_val(col11)
                valor_noturno  = _parse_val(col15)
            elif 'HORAS EXTRA TOTAL' in col1:
                valor_hora_ext = _parse_val(col11)
                valor_extra    = _parse_val(col15)
            elif 'DESMOBILIZAÇÃO' in col1:
                valor_desmobi = _parse_val(col15)
            elif 'TOTAL MEDIÇÃO PERÍODO' in col1 or 'TOTAL MEDI' in col1:
                total_medicao = _parse_val(col15)
            elif 'TOTAL A PAGAR' in col1:
                total_pagar = _parse_val(col15)

        registros.append({
            "numero":            numero,
            "equipamento":       equipamento,
            "periodo":           periodo,
            "cliente":           cliente,
            "horas_diurno":      round(horas_diurno_total, 1),
            "horas_noturno":     round(horas_noturno_total, 1),
            "horas_extra":       round(horas_extra_total, 1),
            "valor_hora_dia":    round(valor_hora_dia, 2),
            "valor_hora_not":    round(valor_hora_not, 2),
            "valor_hora_ext":    round(valor_hora_ext, 2),
            "valor_diurno":      round(valor_diurno, 2),
            "valor_noturno":     round(valor_noturno, 2),
            "valor_extra":       round(valor_extra, 2),
            "valor_desmobi":     round(valor_desmobi, 2),
            "total_medicao":     round(total_medicao, 2),
            "total_pagar":       round(total_pagar, 2),
            "dias_diurno":       dias_diurno,
            "dias_noturno":      dias_noturno,
            "arquivo":           nome_arq,
        })

    print(f"  ✔ {len(registros)} boletins de medição lidos")
    return registros


# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — ORDENS DE SERVIÇO
# ══════════════════════════════════════════════════════════════════
def buscar_ordens_servico(data_ini: str, data_fim: str) -> list[dict]:
    """Ordens de serviço no período."""
    _prog(0.70, "Buscando ordens de serviço...")
    chave  = _chave_loja(f"os|{data_ini}|{data_fim}")
    cached = _cache_load(chave, _TTL_VENDAS)
    if cached:
        print(f"  ✔ OS (cache): {len(cached)}")
        _prog(0.78, "OS carregadas do cache")
        return cached

    def br_to_iso(d: str) -> str:
        try:    return datetime.strptime(d, "%d/%m/%Y").strftime("%Y-%m-%d")
        except: return d

    params = {"data_inicio": br_to_iso(data_ini), "data_fim": br_to_iso(data_fim)}
    raw = _paginar_lojas("ordens_servicos", params)

    normalizado = []
    for o in raw:
        def _float(k):
            try:    return float(o.get(k) or 0)
            except: return 0.0
        def _data(k):
            v = o.get(k, "")
            try:    return datetime.strptime(str(v)[:10], "%Y-%m-%d")
            except: return None

        normalizado.append({
            "ID":        str(o.get("id") or ""),
            "Número":    str(o.get("numero") or o.get("id") or ""),
            "Data":      _data("data_abertura") or _data("data"),
            "Cliente":   o.get("cliente_nome") or o.get("nome") or "",
            "Tecnico":   o.get("tecnico_nome") or o.get("tecnico") or o.get("responsavel") or "",
            "Descricao": o.get("descricao") or o.get("servico") or "",
            "Status":    (o.get("status") or o.get("situacao") or "").upper(),
            "Prioridade":(o.get("prioridade") or "NORMAL").upper(),
            "Valor":     _float("valor") or _float("valor_total"),
            "Fechamento":_data("data_fechamento") or _data("data_conclusao"),
        })

    _cache_save(chave, [
        {**r,
         "Data":      r["Data"].isoformat()      if isinstance(r["Data"],      datetime) else r["Data"],
         "Fechamento":r["Fechamento"].isoformat() if isinstance(r["Fechamento"], datetime) else r["Fechamento"],
        } for r in normalizado
    ])
    print(f"  ✔ {len(normalizado)} ordens de serviço")
    _prog(0.78, "OS carregadas")
    return normalizado



# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — PONTO COLABORADOR (Dixiponto)
# ══════════════════════════════════════════════════════════════════
def buscar_ponto(data_ini: str, data_fim: str) -> dict:
    """
    Busca marcações de ponto via Dixiponto API.
    Retorna: {'funcionarios': [...], 'marcacoes': [...]}
    """
    _prog(0.79, "Buscando dados de ponto...")
    try:
        chave  = f"ponto_v2|{data_ini}|{data_fim}"
        if not _SKIP_PONTO_CACHE:
            cached = _cache_load(chave, _TTL_OUTROS)
            if cached:
                n = len(cached.get("marcacoes", []))
                print(f"  ✔ Ponto (cache): {n} marcações")
                _prog(0.81, f"Ponto: {n} marcações (cache)")
                return cached

        d_ini_dt = datetime.strptime(data_ini, "%d/%m/%Y").date()
        d_fim_dt = datetime.strptime(data_fim, "%d/%m/%Y").date()
        d_ini_iso = d_ini_dt.isoformat()
        d_fim_iso = d_fim_dt.isoformat()

        client = DixiPontoClient(DIXI_EMAIL, DIXI_SENHA, DIXI_UNIDADE)
        client.login()

        funcionarios  = client.get_funcionarios()
        marcacoes_raw = client.get_marcacoes(d_ini_dt, d_fim_dt)

        marcacoes = []
        for m in marcacoes_raw:
            if m.get("considerar") != 1:
                continue
            rf   = m.get("registroFuncionario") or {}
            func = rf.get("funcionario") or {}
            tp   = m.get("tpOrigemMarcacao") or {}
            data_str = DixiPontoClient.parse_data(m.get("dataMarcacao", 0))
            # A API ignora os parâmetros de data e retorna todo o histórico;
            # filtramos aqui no Python para garantir somente o período solicitado.
            if data_str < d_ini_iso or data_str > d_fim_iso:
                continue
            marcacoes.append({
                "id":            m.get("idMarcacao"),
                "funcionario_id": func.get("idFuncionario"),
                "funcionario":   func.get("nome", "Desconhecido"),
                "data":          data_str,
                "hora":          DixiPontoClient.parse_hora(m.get("hora", 0)),
                "origem_id":     tp.get("idTpOrigemMarcacao"),
                "origem":        tp.get("descricao", ""),
                "descricao":     m.get("descricao", ""),
            })

        result = {
            "funcionarios": [
                {"id": f.get("idFuncionario"), "nome": f.get("nome", "")}
                for f in funcionarios
            ],
            "marcacoes": marcacoes,
        }
        _cache_save(chave, result)
        print(f"  ✔ Ponto: {len(funcionarios)} funcionários, {len(marcacoes)} marcações")
        _prog(0.81, f"Ponto: {len(marcacoes)} marcações carregadas")
        return result

    except Exception as e:
        print(f"  [AVISO] Ponto (Dixiponto): {e}")
        _prog(0.81, "Ponto: não disponível (Dixiponto offline)")
        return {"funcionarios": [], "marcacoes": []}


# ══════════════════════════════════════════════════════════════════
#  BUSCA DE DADOS — HORAS REGISTRADAS PELO APP (Supabase)
# ══════════════════════════════════════════════════════════════════
def buscar_horas_app(data_ini: str, data_fim: str) -> list[dict]:
    """
    Busca registros de horas do LocvixApp via Supabase REST API.
    Retorna lista de medições registradas pelo app mobile.
    Campos: id, data, hora_inicio, hora_fim, horas_trabalhadas, turno,
            equipamento, placa, cliente, operador, status, observacoes
    """
    _prog(0.815, "Buscando horas do app (Supabase)...")
    try:
        chave  = f"horas_app|{data_ini}|{data_fim}"
        d_ini_iso = datetime.strptime(data_ini, "%d/%m/%Y").date().isoformat()
        d_fim_iso = datetime.strptime(data_fim, "%d/%m/%Y").date().isoformat()
        inclui_hoje = d_fim_iso >= date.today().isoformat()

        cached = None if inclui_hoje else _cache_load(chave, _TTL_SUPABASE)
        if cached is not None:
            print(f"  ✔ Horas App (cache): {len(cached)} registros")
            _prog(0.82, f"Horas App: {len(cached)} registros (cache)")
            return cached

        if not SUPABASE_URL or not SUPABASE_ANON:
            return []

        hdrs = {
            "apikey":        SUPABASE_ANON,
            "Authorization": f"Bearer {SUPABASE_ANON}",
        }
        params = {
            "data":   f"gte.{d_ini_iso}",
            "order":  "data.asc,hora_inicio.asc",
            "select": "id,data,hora_inicio,hora_fim,horas_trabalhadas,turno,"
                      "equipamento,placa,marca,modelo,cliente,operador,"
                      "observacoes,status",
        }
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/medicoes_horas",
            headers=hdrs, params=params, timeout=15
        )
        resp.raise_for_status()
        raw = resp.json()

        # Supabase retorna todos os registros >= data_ini; filtramos até data_fim
        registros = [rec for rec in raw if rec.get("data", "") <= d_fim_iso]

        _cache_save(chave, registros)
        print(f"  ✔ Horas App: {len(registros)} registros")
        _prog(0.82, f"Horas App: {len(registros)} registros carregados")
        return registros

    except Exception as e:
        print(f"  [AVISO] Horas App (Supabase): {e}")
        _prog(0.82, "Horas App: não disponível")
        return []


# ══════════════════════════════════════════════════════════════════
#  MANUTENÇÃO PREVENTIVA (Supabase — tabela manutencoes_equipamentos)
# ══════════════════════════════════════════════════════════════════

def buscar_manutencoes() -> list[dict]:
    """
    Busca registros de manutenção preventiva de equipamentos via Supabase.
    Tabela: manutencoes_equipamentos
    Colunas: id, equipamento, ultima_manutencao, responsavel_email,
             intervalo_meses, tipo_servico, updated_at
    """
    try:
        if not SUPABASE_URL or not SUPABASE_ANON:
            return []
        hdrs = {
            "apikey":        SUPABASE_ANON,
            "Authorization": f"Bearer {SUPABASE_ANON}",
        }
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/manutencoes_equipamentos",
            headers=hdrs,
            params={
            "select": "id,equipamento,ultima_manutencao,responsavel_email,intervalo_meses,tipo_servico,updated_at",
                "order":  "equipamento.asc",
            },
            timeout=15,
        )
        resp.raise_for_status()
        registros = resp.json()
        print(f"  ✔ Manutenções: {len(registros)} equipamentos")
        return registros
    except Exception as e:
        print(f"  [AVISO] buscar_manutencoes: {e}")
        return []


def salvar_manutencao(equipamento: str, data_str: str,
                      email: str = "", intervalo_meses: int = 2,
                      tipo_servico: str = "") -> bool:
    """
    Salva ou atualiza um registro de manutenção no Supabase.
    data_str: formato 'YYYY-MM-DD'
    Retorna True em caso de sucesso.
    """
    try:
        if not SUPABASE_URL or not SUPABASE_ANON:
            return False
        hdrs = {
            "apikey":        SUPABASE_ANON,
            "Authorization": f"Bearer {SUPABASE_ANON}",
            "Content-Type":  "application/json",
        }
        base = f"{SUPABASE_URL}/rest/v1/manutencoes_equipamentos"
        equip = equipamento.strip()

        # Verifica se já existe registro para este equipamento
        r_get = requests.get(
            base, headers=hdrs,
            params={"equipamento": f"eq.{equip}", "select": "id"},
            timeout=10,
        )
        r_get.raise_for_status()
        existing = r_get.json()

        payload = {
            "ultima_manutencao": data_str,
            "responsavel_email": email.strip() if email.strip() else None,
            "intervalo_meses":   int(intervalo_meses),
          "tipo_servico":      tipo_servico.strip() if tipo_servico.strip() else None,
            "updated_at":        datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        }

        if existing:
            r_up = requests.patch(
                base, headers=hdrs,
                params={"equipamento": f"eq.{equip}"},
                json=payload, timeout=10,
            )
            r_up.raise_for_status()
        else:
            payload["equipamento"] = equip
            r_ins = requests.post(base, headers=hdrs, json=payload, timeout=10)
            r_ins.raise_for_status()

        return True
    except Exception as e:
        print(f"  [ERRO] salvar_manutencao: {e}")
        return False


# ══════════════════════════════════════════════════════════════════
#  GERADOR DE EXCEL
# ══════════════════════════════════════════════════════════════════
COR_HD   = "1F4E79"
COR_HD_F = "FFFFFF"
COR_TOTL = "D6E4F0"
COR_ZEBR = "EBF4FB"
COR_SUBT = "BDD7EE"

def _hdr(ws, row, c1, c2, texto):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=texto)
    cell.font  = Font(bold=True, color=COR_HD_F, size=12)
    cell.fill  = PatternFill("solid", fgColor=COR_HD)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _borda(ws, r1, r2, c1, c2):
    brd = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = brd

def _larguras(ws, d: dict):
    for col, w in d.items():
        ws.column_dimensions[col].width = w

def _sheet_vendas(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Vendas por Produto")
    grp = (df.groupby(["Cod. Produto","Produto","Categoria"])
            .agg(Qtd=("Qtd","sum"), V_Bruto=("Vlr Bruto","sum"),
                 Desc=("Desconto","sum"), V_Liq=("Vlr Líquido","sum"), NFs=("ID","nunique"))
            .reset_index().sort_values("V_Liq", ascending=False))
    grp["Part %"] = grp["V_Liq"] / grp["V_Liq"].sum() * 100

    cols  = ["Cod. Produto","Produto","Categoria","Qtd","V_Bruto","Desc","V_Liq","Part %","NFs"]
    hdrs  = ["Código","Produto","Categoria","Qtd Total","Vlr Bruto","Desconto","Vlr Líquido","Part. %","Qtd NFs"]
    _hdr(ws, 1, 1, len(cols), "RANKING DE VENDAS POR PRODUTO — LOCVIX")
    ws.row_dimensions[1].height = 25
    hf = PatternFill("solid", fgColor=COR_HD)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = hf; c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    zf = PatternFill("solid", fgColor=COR_ZEBR)
    for ri, row in enumerate(grp[cols].itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=round(val, 4) if isinstance(val, float) else val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0: cell.fill = zf
            if ci == 4:   cell.number_format = "#,##0.00"
            elif ci in (5,6,7): cell.number_format = "R$ #,##0.00"
            elif ci == 8:
                cell.number_format = "0.00%"
                cell.value = (val/100) if isinstance(val, float) else val
    total_row = len(grp) + 3
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    for ci, cl in [(4,"D"),(5,"E"),(6,"F"),(7,"G")]:
        c = ws.cell(row=total_row, column=ci,
                    value=f"=SUM({cl}3:{cl}{total_row-1})")
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=COR_TOTL)
        c.number_format = "#,##0.00" if ci == 4 else "R$ #,##0.00"
    _borda(ws, 2, total_row, 1, len(cols))
    _larguras(ws, {"A":14,"B":44,"C":22,"D":10,"E":14,"F":14,"G":14,"H":10,"I":8})
    ws.freeze_panes = "A3"
    return grp

def _sheet_clientes(wb, df_vendas: pd.DataFrame):
    ws = wb.create_sheet("Vendas por Cliente")
    grp = (df_vendas.groupby(["Cliente"])
            .agg(NFs=("ID","nunique"), V_Bruto=("Vlr Bruto","sum"),
                 Desc=("Desconto","sum"), V_Liq=("Vlr Líquido","sum"))
            .reset_index().sort_values("V_Liq", ascending=False))
    grp["Part %"] = grp["V_Liq"] / grp["V_Liq"].sum() * 100
    cols  = ["Cliente","NFs","V_Bruto","Desc","V_Liq","Part %"]
    hdrs  = ["Cliente","Nº NFs / Pedidos","Vlr Bruto","Desconto","Vlr Líquido","Part. %"]
    _hdr(ws, 1, 1, len(cols), "RANKING DE VENDAS POR CLIENTE — LOCVIX")
    ws.row_dimensions[1].height = 25
    hf = PatternFill("solid", fgColor=COR_HD)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = hf; c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    zf = PatternFill("solid", fgColor=COR_ZEBR)
    for ri, row in enumerate(grp[cols].itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=round(val, 4) if isinstance(val, float) else val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0: cell.fill = zf
            if ci in (3,4,5): cell.number_format = "R$ #,##0.00"
            elif ci == 6:
                cell.number_format = "0.00%"
                cell.value = (val/100) if isinstance(val, float) else val
    total_row = len(grp) + 3
    for ci, cl in [(3,"C"),(4,"D"),(5,"E")]:
        c = ws.cell(row=total_row, column=ci,
                    value=f"=SUM({cl}3:{cl}{total_row-1})")
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=COR_TOTL)
        c.number_format = "R$ #,##0.00"
    _borda(ws, 2, total_row, 1, len(cols))
    _larguras(ws, {"A":42,"B":14,"C":14,"D":14,"E":14,"F":10})
    ws.freeze_panes = "A3"

def _sheet_financeiro(wb, receber: list, pagar: list):
    for dados, nome, titulo in [
        (receber, "Contas a Receber", "CONTAS A RECEBER — LOCVIX"),
        (pagar,   "Contas a Pagar",   "CONTAS A PAGAR — LOCVIX"),
    ]:
        ws = wb.create_sheet(nome)
        if not dados:
            ws.cell(row=1, column=1, value=f"Sem dados de {nome}")
            continue
        cols  = ["Descrição","Pessoa","Valor","Valor Pago","Saldo","Vencimento","Status","Categoria"]
        _hdr(ws, 1, 1, len(cols), titulo)
        ws.row_dimensions[1].height = 25
        hf = PatternFill("solid", fgColor=COR_HD)
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.fill = hf; c.font = Font(bold=True, color="FFFFFF", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 28
        zf = PatternFill("solid", fgColor=COR_ZEBR)
        for ri, r in enumerate(dados, start=3):
            vals = [
                r.get("Descrição",""), r.get("Pessoa",""),
                r.get("Valor",0), r.get("Valor Pago",0), r.get("Saldo",0),
                r.get("Vencimento"), r.get("Status",""), r.get("Categoria",""),
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(vertical="center")
                if ri % 2 == 0: cell.fill = zf
                if ci in (3,4,5): cell.number_format = "R$ #,##0.00"
                elif ci == 6 and val: cell.number_format = "DD/MM/YYYY"
        _borda(ws, 2, len(dados)+2, 1, len(cols))
        _larguras(ws, {"A":30,"B":34,"C":14,"D":14,"E":14,"F":12,"G":16,"H":22})
        ws.freeze_panes = "A3"

def _sheet_os(wb, os_list: list):
    ws = wb.create_sheet("Ordens de Serviço")
    if not os_list:
        ws.cell(row=1, column=1, value="Sem dados de OS no período")
        return
    cols  = ["Número","Data","Cliente","Tecnico","Descricao","Status","Prioridade","Valor","Fechamento"]
    _hdr(ws, 1, 1, len(cols), "ORDENS DE SERVIÇO — LOCVIX")
    ws.row_dimensions[1].height = 25
    hf = PatternFill("solid", fgColor=COR_HD)
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = hf; c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    zf = PatternFill("solid", fgColor=COR_ZEBR)
    for ri, r in enumerate(os_list, start=3):
        vals = [r.get(k) for k in cols]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0: cell.fill = zf
            if ci == 8: cell.number_format = "R$ #,##0.00"
            elif ci in (2,9) and val: cell.number_format = "DD/MM/YYYY"
    _borda(ws, 2, len(os_list)+2, 1, len(cols))
    _larguras(ws, {"A":10,"B":12,"C":34,"D":22,"E":36,"F":16,"G":12,"H":12,"I":12})
    ws.freeze_panes = "A3"

def _sheet_contratos(wb, contratos: list):
    ws = wb.create_sheet("Contratos")
    if not contratos:
        ws.cell(row=1, column=1, value="Sem contratos cadastrados")
        return
    cols = ["Número","Cliente","Descricao","Valor","Periodicidade","Inicio","Fim","Status"]
    _hdr(ws, 1, 1, len(cols), "CONTRATOS / RECORRÊNCIA — LOCVIX")
    ws.row_dimensions[1].height = 25
    hf = PatternFill("solid", fgColor=COR_HD)
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = hf; c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    zf = PatternFill("solid", fgColor=COR_ZEBR)
    for ri, r in enumerate(contratos, start=3):
        vals = [r.get(k) for k in cols]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0: cell.fill = zf
            if ci == 4: cell.number_format = "R$ #,##0.00"
            elif ci in (6,7) and val: cell.number_format = "DD/MM/YYYY"
    _borda(ws, 2, len(contratos)+2, 1, len(cols))
    _larguras(ws, {"A":10,"B":34,"C":36,"D":14,"E":16,"F":12,"G":12,"H":16})
    ws.freeze_panes = "A3"

def gerar_excel(df_vendas: pd.DataFrame, receber: list, pagar: list,
                os_list: list, contratos: list, caminho: str) -> str:
    """Gera arquivo Excel completo com todas as abas."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove a aba padrão

    grp_prod = _sheet_vendas(wb, df_vendas)
    _sheet_clientes(wb, df_vendas)
    _sheet_financeiro(wb, receber, pagar)
    _sheet_os(wb, os_list)
    _sheet_contratos(wb, contratos)

    wb.save(caminho)
    print(f"  ✔ Excel salvo: {caminho}")
    return caminho


# ══════════════════════════════════════════════════════════════════
#  DASHBOARD HTML
# ══════════════════════════════════════════════════════════════════
def gerar_dashboard_html(
    df_vendas: pd.DataFrame,
    receber:   list,
    pagar:     list,
    pagar_all: list,
    os_list:   list,
    contratos: list,
    caminho:   str,
    data_ini:  str,
    data_fim:  str,
    ponto_data:   dict | None = None,
    orcamentos:   list | None = None,
    medicoes:     list | None = None,
    horas_app:    list | None = None,
    manutencoes:  list | None = None,
) -> str:
    """Gera dashboard HTML interativo completo para Locvix."""
    import json as _json

    # ── Datas ───────────────────────────────────────────────────────
    try:
        dt_min = df_vendas["Data"].dropna().min()
        dt_max = df_vendas["Data"].dropna().max()
    except Exception:
        dt_min = dt_max = datetime.now()
    dt_min_iso = dt_min.strftime("%Y-%m-%d") if isinstance(dt_min, datetime) else data_ini
    dt_max_iso = dt_max.strftime("%Y-%m-%d") if isinstance(dt_max, datetime) else data_fim

    # Datas ISO do período de ponto (independente do range de vendas)
    try:
        ponto_d_ini_iso = datetime.strptime(data_ini, "%d/%m/%Y").strftime("%Y-%m-%d")
        ponto_d_fim_iso = datetime.strptime(data_fim, "%d/%m/%Y").strftime("%Y-%m-%d")
    except Exception:
        ponto_d_ini_iso = dt_min_iso
        ponto_d_fim_iso = dt_max_iso
    periodo    = f"{data_ini} a {data_fim}"
    agora_str  = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── Logo em base64 ──────────────────────────────────────────────
    _logo_b64   = ""
    _logo_mime  = "image/png"
    _logo_candidates = [
        (os.path.join(_BASE_DIR, "logo_locvix.png"),  "image/png"),
        (os.path.join(_BASE_DIR, "logo_locvix.jpg"),  "image/jpeg"),
        (os.path.join(_BASE_DIR, "logo_locvix.jfif"), "image/jpeg"),
        (os.path.join(_BASE_DIR, "logo locvix.png"),  "image/png"),
        (os.path.join(_BASE_DIR, "logo locvix.jpg"),  "image/jpeg"),
        (os.path.join(_BASE_DIR, "logo locvix.jfif"), "image/jpeg"),
    ]
    for _lp, _lm in _logo_candidates:
        if os.path.exists(_lp):
            with open(_lp, "rb") as f:
                _logo_b64  = base64.b64encode(f.read()).decode()
            _logo_mime = _lm
            break

    _logo_alfa_b64  = ""
    _logo_alfa_path = os.path.join(_BASE_DIR, "logo_alfa.jpg")
    if not os.path.exists(_logo_alfa_path):
        _logo_alfa_path = r"Z:\codigos\Fabio\Logo Alfa.jpg"
    if os.path.exists(_logo_alfa_path):
        with open(_logo_alfa_path, "rb") as f:
            _logo_alfa_b64 = base64.b64encode(f.read()).decode()

    logo_tag = (f'<img src="data:{_logo_mime};base64,{_logo_b64}" '
                f'style="height:52px;width:auto;border-radius:6px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,.3);object-fit:contain;"/>'
               ) if _logo_b64 else '<span style="font-size:24px;font-weight:800">LOCVIX</span>'
    logo_alfa_tag = (f'<img src="data:image/jpeg;base64,{_logo_alfa_b64}" '
                     f'style="height:30px;width:auto;object-fit:contain;border-radius:4px;"/>'
                    ) if _logo_alfa_b64 else "<strong>Alfa Soluções</strong>"

    # ── Dados para o JS ─────────────────────────────────────────────
    def prep_vendas():
        rows = []
        for _, r in df_vendas.iterrows():
            d = r.get("Data")
            rows.append({
                "id":        str(r.get("ID","") or ""),
                "data":      d.strftime("%Y-%m-%d") if isinstance(d, datetime) else "",
                "cliente":   str(r.get("Cliente",""))[:50],
                "cod":       str(r.get("Cod. Produto","") or ""),
                "produto":   str(r.get("Produto","") or "")[:50],
                "categoria": str(r.get("Categoria","") or "SEM CATEGORIA"),
                "cc":        str(r.get("Centro Custo","") or "SEM CENTRO DE CUSTO"),
                "vendedor":  str(r.get("Vendedor","") or "Sem Vendedor"),
                "status":    str(r.get("Status","") or ""),
                "qtd":       round(float(r.get("Qtd",0) or 0), 4),
                "bruto":     round(float(r.get("Vlr Bruto",0) or 0), 2),
                "desc":      round(float(r.get("Desconto",0) or 0), 2),
                "liq":       round(float(r.get("Vlr Líquido",0) or 0), 2),
                "loja":      str(r.get("Loja","") or ""),
            })
        return rows

    def prep_financeiro(lista: list, tipo: str):
        rows = []
        for r in lista:
            def _dt(k):
                v = r.get(k)
                if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
                if isinstance(v, str) and v: return v[:10]
                return ""
            rows.append({
                "id":       r.get("ID",""),
                "desc":     r.get("Descrição","")[:60],
                "pessoa":   r.get("Pessoa","")[:50],
                "valor":    round(float(r.get("Valor",0) or 0), 2),
                "pago":     round(float(r.get("Valor Pago",0) or 0), 2),
                "saldo":    round(float(r.get("Saldo",0) or 0), 2),
                "venc":     _dt("Vencimento"),
                "pgto":     _dt("Pagamento"),
                "status":   r.get("Status",""),
                "cat":      r.get("Categoria",""),
                "cc":       r.get("Centro Custo",""),
                "tipo":     tipo,
                "loja":     r.get("Loja",""),
                "parc_n":   0,
                "parc_tot": 0,
            })
        # Detecta parcelas: agrupa por (cc, desc) e numera
        from collections import Counter
        key_count = Counter((row["cc"], row["desc"].strip()) for row in rows)
        key_idx = {}
        for row in sorted(rows, key=lambda x: (x["cc"], x["desc"].strip(), x["venc"])):
            k = (row["cc"], row["desc"].strip())
            if key_count[k] > 1:
                key_idx[k] = key_idx.get(k, 0) + 1
                row["parc_n"]   = key_idx[k]
                row["parc_tot"] = key_count[k]
        return rows

    def prep_os():
        rows = []
        for r in os_list:
            def _dt(k):
                v = r.get(k)
                if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
                return ""
            rows.append({
                "id":   r.get("Número",""),
                "data": _dt("Data"),
                "cli":  r.get("Cliente","")[:40],
                "tec":  r.get("Tecnico","")[:30],
                "desc": r.get("Descricao","")[:60],
                "st":   r.get("Status",""),
                "prio": r.get("Prioridade",""),
                "val":  round(float(r.get("Valor",0) or 0), 2),
                "fech": _dt("Fechamento"),
            })
        return rows

    raw_vendas = prep_vendas()
    raw_rec    = prep_financeiro(receber, "receber")
    raw_pag    = prep_financeiro(pagar, "pagar")
    raw_pag_all = prep_financeiro(pagar_all, "pagar")
    raw_os     = prep_os()
    raw_contr  = [{
        "id":     c.get("Número",""),
        "cli":    c.get("Cliente","")[:40],
        "val":    round(float(c.get("Valor",0) or 0), 2),
        "period": c.get("Periodicidade",""),
        "st":     c.get("Status",""),
    } for c in contratos]
    raw_orc = orcamentos or []  # já preparado em buscar_orcamentos()
    raw_med = medicoes or []    # já preparado em buscar_medicoes()
    raw_horas_app = horas_app or []  # registros do LocvixApp via Supabase

    # ── Status de manutenção preventiva por Centro de Custo ─────────
    EXCLUIR_CC = {'ADM/FINANCEIRO','SUBLOCAÇÕES - TERCEIROS','SEM CENTRO DE CUSTO','MANUTENÇÃO'}
    _cc_set = sorted({
        r["cc"].strip() for r in raw_pag_all
        if r["cc"] and r["cc"].strip().upper() not in EXCLUIR_CC
    })
    _manut_dict = {rec["equipamento"]: rec for rec in (manutencoes or [])}
    _hoje_d = date.today()
    raw_manutencoes = []
    for _cc in _cc_set:
        _rec = _manut_dict.get(_cc, {})
        _ultima   = (_rec.get("ultima_manutencao") or "")[:10]
        _intervalo = int(_rec.get("intervalo_meses") or 2)
        _email    = _rec.get("responsavel_email") or ""
        _tipo_srv = _rec.get("tipo_servico") or ""
        if _ultima:
            try:
                _dt_ultima  = date.fromisoformat(_ultima)
                _dt_proxima = _dt_ultima + timedelta(days=_intervalo * 30)
                _dias       = (_dt_proxima - _hoje_d).days
                _status     = "vencida" if _dias < 0 else ("proxima" if _dias <= 5 else "ok")
                _proxima_s  = _dt_proxima.isoformat()
            except Exception:
                _status, _dias, _proxima_s = "vencida", -9999, ""
        else:
            _status, _dias, _proxima_s = "vencida", -9999, ""
        raw_manutencoes.append({
            "cc":       _cc,
            "ultima":   _ultima,
            "proxima":  _proxima_s,
            "status":   _status,
            "dias":     _dias,
            "email":    _email,
            "intervalo": _intervalo,
        "tipo_servico": _tipo_srv,
        })

    def _clean_surrogates(o):
        """Remove lone surrogate characters that break UTF-8/JSON serialization."""
        if isinstance(o, str):
            return ''.join(c for c in o if not ('\ud800' <= c <= '\udfff'))
        if isinstance(o, list):  return [_clean_surrogates(x) for x in o]
        if isinstance(o, dict):  return {k: _clean_surrogates(v) for k, v in o.items()}
        return o
    jv = lambda v: _json.dumps(_clean_surrogates(v), ensure_ascii=True).replace('</', r'<\/')

    # Supabase — para uso no JS do formulário de manutenção
    supabase_url  = (SUPABASE_URL  or "").replace("'", "")
    supabase_anon = (SUPABASE_ANON or "").replace("'", "")

    # Ponto data
    ponto_func = (ponto_data or {}).get("funcionarios", [])
    ponto_marc = (ponto_data or {}).get("marcacoes", [])

    # ── Categorias únicas ───────────────────────────────────────────
    categorias   = sorted(df_vendas["Categoria"].dropna().unique().tolist())
    vendedores   = sorted(df_vendas["Vendedor"].dropna().unique().tolist())
    opt_cat  = "\n".join(f'<option value="{c}">{c}</option>' for c in categorias)
    opt_vend = "\n".join(f'<option value="{v}">{v}</option>' for v in vendedores)

    # ── HTML completo ───────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>Dashboard — Locvix</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4f8;color:#1e293b;font-size:14px;}}

/* ── TOPBAR ── */
.topbar{{background:linear-gradient(135deg,#0f2027 0%,#1a3a4a 60%,#203a43 100%);
  color:#fff;padding:18px 32px;display:flex;align-items:center;justify-content:space-between;
  box-shadow:0 2px 12px rgba(0,0,0,.35);}}
.topbar-title{{font-size:21px;font-weight:800;letter-spacing:.8px;}}
.topbar .sub{{font-size:12px;color:#90cdf4;margin-top:3px;}}
.topbar .periodo{{font-size:12px;color:#7dd3fc;text-align:right;}}

/* ── FILTER BAR ── */
.filter-bar{{background:#fff;border-bottom:2px solid #e2e8f0;padding:12px 32px;
  display:flex;align-items:flex-end;gap:14px;flex-wrap:wrap;
  box-shadow:0 1px 4px rgba(0,0,0,.07);position:sticky;top:0;z-index:100;}}
.filter-group{{display:flex;flex-direction:column;gap:3px;}}
.filter-group label{{font-size:11px;font-weight:700;color:#718096;text-transform:uppercase;letter-spacing:.4px;}}
.filter-group input[type=date],
.filter-group select{{border:1px solid #e2e8f0;border-radius:7px;padding:7px 10px;
  font-size:13px;color:#1e293b;background:#f8fafc;outline:none;min-width:140px;
  cursor:pointer;transition:border-color .15s;}}
.filter-group input:focus,.filter-group select:focus{{border-color:#1a3a4a;background:#fff;}}
.filter-sep{{width:1px;height:36px;background:#e2e8f0;margin:0 4px;align-self:center;}}
.btn{{padding:8px 18px;border:none;border-radius:7px;font-size:13px;font-weight:700;cursor:pointer;transition:all .15s;}}
.btn-apply{{background:#1a3a4a;color:#fff;}}
.btn-apply:hover{{background:#0f2027;}}
.btn-clear{{background:#e2e8f0;color:#4a5568;}}
.btn-clear:hover{{background:#cbd5e0;}}
.filter-info{{font-size:12px;color:#059669;font-weight:600;margin-left:auto;align-self:center;white-space:nowrap;}}
.loja-pills{{display:flex;gap:5px;}}
.loja-pill{{background:#e2e8f0;color:#4a5568;border:1px solid #cbd5e0;border-radius:6px;
  padding:5px 13px;font-size:12px;font-weight:700;cursor:pointer;transition:all .15s;}}
.loja-pill:hover{{background:#1a3a4a;color:#fff;border-color:#1a3a4a;}}
.loja-pill.active{{background:#1a3a4a;color:#fff;border-color:#1a3a4a;}}

/* ── FILTROS INTERNOS DE MÓDULO ── */
.fin-filter-bar{{background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;
  padding:12px 18px;margin-bottom:18px;display:flex;align-items:flex-end;gap:14px;flex-wrap:wrap;}}
body[data-theme="dark"] .fin-filter-bar{{background:#1e293b;border-color:#334155;}}

/* ── LAYOUT ── */
.container{{max-width:1520px;margin:0 auto;padding:22px 20px;}}
.section-title{{font-size:14px;font-weight:800;color:#1a3a4a;margin:26px 0 11px;
  padding-left:10px;border-left:4px solid #1a3a4a;letter-spacing:.5px;text-transform:uppercase;}}

/* ── KPI CARDS ── */
.kpi-grid{{display:grid;gap:14px;margin-bottom:24px;}}
.kpi-grid.col3{{grid-template-columns:repeat(3,1fr);}}
.kpi-grid.col4{{grid-template-columns:repeat(4,1fr);}}
.kpi-grid.col5{{grid-template-columns:repeat(5,1fr);}}
.kpi-grid.col6{{grid-template-columns:repeat(6,1fr);}}
@media(max-width:1100px){{.kpi-grid.col6,.kpi-grid.col5{{grid-template-columns:repeat(3,1fr);}}}}
@media(max-width:700px){{.kpi-grid{{grid-template-columns:repeat(2,1fr)!important;}}}}
.kpi-card{{background:#fff;border-radius:10px;padding:16px 14px;
  box-shadow:0 1px 4px rgba(0,0,0,.1);border-top:4px solid #1a3a4a;text-align:center;
  transition:transform .15s;}}
.kpi-card:hover{{transform:translateY(-3px);box-shadow:0 4px 12px rgba(0,0,0,.15);}}
.kpi-card.green{{border-top-color:#059669;}}
.kpi-card.teal{{border-top-color:#0891b2;}}
.kpi-card.orange{{border-top-color:#d97706;}}
.kpi-card.red{{border-top-color:#dc2626;}}
.kpi-card.purple{{border-top-color:#7c3aed;}}
.kpi-card.blue{{border-top-color:#2563eb;}}
/* ── BOTÕES FILTRO ORÇAMENTO ── */
.btn-filter-orc{{background:#1e3a5f;color:#94a3b8;border:1px solid #334155;border-radius:6px;
  padding:5px 14px;font-size:12px;cursor:pointer;transition:all .2s;}}
.btn-filter-orc:hover{{background:#2563eb;color:#fff;border-color:#2563eb;}}
.btn-filter-orc.active{{background:#2563eb;color:#fff;border-color:#2563eb;}}
/* ── BADGE STATUS ── */
.badge{{display:inline-block;padding:2px 10px;border-radius:12px;font-size:11px;font-weight:700;}}
/* ── MODAL DETALHE CC (Manutenção) ── */
.cc-overlay{{position:fixed;inset:0;background:rgba(0,0,0,.65);display:none;
  justify-content:center;align-items:center;z-index:25000;padding:16px;}}
.cc-overlay.open{{display:flex;}}
.cc-modal{{background:#1e293b;border-radius:14px;width:100%;max-width:760px;
  max-height:88vh;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,.6);
  animation:hmFade .2s ease;}}
.cc-modal-header{{display:flex;align-items:center;justify-content:space-between;
  padding:16px 20px 12px;border-bottom:1px solid #334155;}}
.cc-modal-header h2{{color:#f59e0b;font-size:15px;margin:0;}}
.cc-modal-body{{overflow-y:auto;padding:14px 20px;flex:1;}}
.cc-modal table{{width:100%;border-collapse:collapse;font-size:12.5px;}}
.cc-modal th{{background:#0f172a;color:#94a3b8;padding:7px 8px;text-align:left;
  position:sticky;top:0;font-size:11px;text-transform:uppercase;letter-spacing:.04em;}}
.cc-modal th.num{{text-align:right;}}
.cc-modal td{{padding:6px 8px;border-bottom:1px solid #1e293b;color:#e2e8f0;vertical-align:top;}}
.cc-modal td.num{{text-align:right;white-space:nowrap;color:#fcd34d;font-weight:600;}}
.cc-modal tr:hover td{{background:#1e3a5f33;}}
.cc-modal tfoot td{{background:#0f172a;color:#f59e0b;font-weight:700;padding:8px;}}
.cc-modal tfoot td.num{{color:#fcd34d;font-size:13px;}}
.tip-cell{{position:relative;cursor:help;}}
.tip-cell::after{{content:attr(data-tip);position:absolute;bottom:calc(100% + 6px);left:0;
  background:#0f172a;color:#f1f5f9;padding:5px 12px;border-radius:8px;font-size:11px;
  white-space:nowrap;border:1px solid #334155;opacity:0;pointer-events:none;
  transition:opacity .18s;z-index:200;box-shadow:0 4px 12px rgba(0,0,0,.4);}}
.tip-cell:hover::after{{opacity:1;}}
/* ── MODAL AJUDA ── */
.fab-help{{position:fixed;bottom:24px;right:136px;width:48px;height:48px;border-radius:50%;
  background:#1a3a4a;color:#fff;border:2px solid rgba(255,255,255,.18);font-size:20px;font-weight:800;
  cursor:pointer;box-shadow:0 4px 18px rgba(0,0,0,.35);z-index:10000;transition:all .2s;
  display:flex;align-items:center;justify-content:center;line-height:1;}}
.fab-help:hover{{transform:scale(1.12);}}
.help-overlay{{position:fixed;inset:0;background:rgba(0,0,0,.6);display:none;
  justify-content:center;align-items:center;z-index:20000;padding:16px;}}
.help-overlay.open{{display:flex;}}
.help-modal{{background:#1e293b;border-radius:14px;width:100%;max-width:680px;
  max-height:88vh;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,.6);
  animation:hmFade .2s ease;}}
@keyframes hmFade{{from{{opacity:0;transform:translateY(10px)}}to{{opacity:1;transform:none}}}}
.hm-header{{display:flex;align-items:center;justify-content:space-between;
  padding:18px 22px 14px;border-bottom:1px solid #334155;}}
.hm-header h2{{color:#f1f5f9;font-size:16px;margin:0;}}
.hm-header small{{font-size:12px;color:#94a3b8;margin-left:8px;}}
.hm-close{{background:none;border:none;color:#94a3b8;font-size:20px;cursor:pointer;
  padding:4px 8px;border-radius:6px;line-height:1;}}
.hm-close:hover{{background:#334155;color:#f1f5f9;}}
.hm-tabs{{display:flex;gap:3px;padding:10px 22px 0;flex-wrap:wrap;border-bottom:1px solid #334155;}}
.hm-tab{{padding:6px 12px;border-radius:7px 7px 0 0;font-size:12px;font-weight:600;
  cursor:pointer;color:#94a3b8;background:none;border:none;
  border-bottom:2px solid transparent;transition:all .15s;}}
.hm-tab:hover{{color:#f1f5f9;background:#334155;}}
.hm-tab.active{{color:#38bdf8;border-bottom-color:#38bdf8;}}
.hm-body{{overflow-y:auto;padding:20px 22px;flex:1;}}
.hm-panel{{display:none;}}.hm-panel.active{{display:block;}}
.hm-body h3{{color:#38bdf8;font-size:14px;margin:0 0 10px;display:flex;align-items:center;gap:7px;}}
.hm-body p{{color:#cbd5e1;font-size:13px;line-height:1.7;margin-bottom:10px;}}
.hm-body ul{{padding-left:17px;margin-bottom:12px;}}
.hm-body li{{color:#cbd5e1;font-size:13px;line-height:1.85;}}
.hm-body li strong{{color:#f1f5f9;}}
.hm-tip{{background:#0f2d3d;border-left:3px solid #0891b2;border-radius:0 7px 7px 0;
  padding:9px 13px;margin:12px 0;font-size:12.5px;color:#7dd3fc;}}
.hm-tip strong{{color:#38bdf8;}}
.hm-step{{display:flex;gap:10px;align-items:flex-start;margin-bottom:12px;}}
.hm-snum{{background:#1a3a4a;color:#38bdf8;border-radius:50%;width:26px;height:26px;
  min-width:26px;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:12px;}}
.hm-stxt{{color:#cbd5e1;font-size:13px;line-height:1.7;padding-top:3px;}}
.hm-stxt strong{{color:#f1f5f9;}}
.hm-hr{{border:none;border-top:1px solid #334155;margin:14px 0;}}
.hm-footer{{padding:12px 22px;border-top:1px solid #334155;
  display:flex;justify-content:space-between;align-items:center;}}
.hm-footer span{{font-size:11px;color:#64748b;}}
.hm-btn{{background:#1a3a4a;color:#fff;border:none;border-radius:7px;
  padding:7px 18px;font-size:13px;font-weight:700;cursor:pointer;transition:background .15s;}}
.hm-btn:hover{{background:#2563eb;}}
.badge.green{{background:#dcfce7;color:#166534;}}
.badge.red{{background:#fee2e2;color:#991b1b;}}
.badge.yellow{{background:#fef9c3;color:#854d0e;}}
.badge.blue{{background:#dbeafe;color:#1e40af;}}
.badge.gray{{background:#f1f5f9;color:#475569;}}
.kpi-label{{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.4px;margin-bottom:6px;}}
.kpi-value{{font-size:20px;font-weight:800;color:#1e293b;}}
.kpi-value.small{{font-size:15px;}}

/* ── CHARTS ── */
.chart-row{{display:grid;gap:16px;margin-bottom:16px;}}
.chart-row.col2{{grid-template-columns:1fr 1fr;}}
.chart-row.col3{{grid-template-columns:1fr 1fr 1fr;}}
@media(max-width:900px){{.chart-row.col2,.chart-row.col3{{grid-template-columns:1fr;}}}}
.chart-card{{background:#fff;border-radius:10px;padding:18px 20px;box-shadow:0 1px 4px rgba(0,0,0,.1);}}
.chart-card h3{{font-size:13px;font-weight:700;color:#4a5568;margin-bottom:14px;
  text-transform:uppercase;letter-spacing:.4px;}}

/* ── TABLE CARD ── */
.table-card{{background:#fff;border-radius:10px;padding:18px 20px;
  box-shadow:0 1px 4px rgba(0,0,0,.1);margin-bottom:16px;overflow-x:auto;}}
.table-card h3{{font-size:13px;font-weight:700;color:#4a5568;margin-bottom:14px;
  text-transform:uppercase;letter-spacing:.4px;}}
table.data-tbl{{width:100%;border-collapse:collapse;font-size:13px;}}
table.data-tbl thead th{{background:#1a3a4a;color:#fff;padding:9px 14px;text-align:left;
  font-size:11px;font-weight:700;letter-spacing:.3px;white-space:nowrap;}}
table.data-tbl thead th.num{{text-align:right;}}
table.data-tbl tbody tr:nth-child(even){{background:#f8fafc;}}
table.data-tbl tbody tr:hover{{background:#edf2f7;}}
table.data-tbl tbody td{{padding:8px 14px;border-bottom:1px solid #e2e8f0;color:#1e293b;}}
table.data-tbl tbody td.num{{text-align:right;font-variant-numeric:tabular-nums;font-weight:600;}}
table.data-tbl tfoot td{{background:#e2e8f0;font-weight:800;padding:9px 14px;}}
table.data-tbl tfoot td.num{{text-align:right;}}

/* ── STATUS BADGES ── */
.badge{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;}}
.badge.verde{{background:#d1fae5;color:#065f46;}}
.badge.vermelho{{background:#fef2f2;color:#991b1b;}}
.badge.amarelo{{background:#fef3c7;color:#92400e;}}
.badge.cinza{{background:#f1f5f9;color:#64748b;}}

/* ── FOOTER ── */
.footer{{margin-top:32px;padding:20px 32px;border-top:1px solid #e2e8f0;background:#f8fafc;
  display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap;}}
.footer-dev{{display:flex;align-items:center;gap:10px;font-size:12px;color:#94a3b8;}}
.footer-dev strong{{color:#64748b;}}
.footer-sep{{width:1px;height:32px;background:#e2e8f0;}}
.footer-gen{{font-size:11px;color:#b0bec5;text-align:right;}}

/* ── DARK MODE BUTTON ── */
#btn-theme{{position:fixed;bottom:24px;right:24px;z-index:10000;width:48px;height:48px;
  border-radius:50%;border:2px solid rgba(255,255,255,.18);cursor:pointer;font-size:20px;
  box-shadow:0 4px 18px rgba(0,0,0,.35);background:#1a3a4a;color:#f1f5f9;
  transition:all .2s;display:flex;align-items:center;justify-content:center;line-height:1;}}
#btn-theme:hover{{transform:scale(1.12);}}
#btn-fullscreen{{position:fixed;bottom:24px;right:80px;z-index:10000;width:48px;height:48px;
  border-radius:50%;border:2px solid rgba(255,255,255,.18);cursor:pointer;font-size:18px;
  box-shadow:0 4px 18px rgba(0,0,0,.35);background:#1a3a4a;color:#f1f5f9;
  transition:all .2s;display:flex;align-items:center;justify-content:center;line-height:1;}}
#btn-fullscreen:hover{{transform:scale(1.12);}}
body[data-theme="dark"] #btn-fullscreen{{background:#e2e8f0;color:#1e293b;}}

/* ── DARK MODE ── */
body[data-theme="dark"]{{background:#0f172a;color:#e2e8f0;}}
body[data-theme="dark"] .filter-bar{{background:#1e293b;border-color:#334155;}}
body[data-theme="dark"] .filter-group label{{color:#94a3b8;}}
body[data-theme="dark"] .filter-group input,
body[data-theme="dark"] .filter-group select{{background:#0f172a;color:#e2e8f0;border-color:#475569;}}
body[data-theme="dark"] .filter-sep{{background:#334155;}}
body[data-theme="dark"] .btn-clear{{background:#334155;color:#cbd5e0;}}
body[data-theme="dark"] .btn-apply{{background:#3b82f6;}}
body[data-theme="dark"] .section-title{{color:#93c5fd;border-color:#3b82f6;}}
body[data-theme="dark"] .kpi-card{{background:#1e293b;box-shadow:0 2px 8px rgba(0,0,0,.5);}}
body[data-theme="dark"] .kpi-label{{color:#94a3b8;}}
body[data-theme="dark"] .kpi-value{{color:#f1f5f9;}}
body[data-theme="dark"] .chart-card{{background:#1e293b;}}
body[data-theme="dark"] .chart-card h3{{color:#94a3b8;}}
body[data-theme="dark"] .table-card{{background:#1e293b;}}
body[data-theme="dark"] .table-card h3{{color:#94a3b8;}}
body[data-theme="dark"] table.data-tbl thead th{{background:#0f2027;}}
body[data-theme="dark"] table.data-tbl tbody td{{color:#e2e8f0;border-color:#334155;}}
body[data-theme="dark"] table.data-tbl tbody tr:nth-child(even){{background:#1e293b;}}
body[data-theme="dark"] table.data-tbl tbody tr:hover{{background:#334155;}}
body[data-theme="dark"] table.data-tbl tfoot td{{background:#334155;}}
body[data-theme="dark"] .footer{{background:#1e293b;border-color:#334155;}}
body[data-theme="dark"] .footer-dev{{color:#64748b;}}
body[data-theme="dark"] .footer-sep{{background:#334155;}}
body[data-theme="dark"] #btn-theme{{background:#e2e8f0;color:#1e293b;}}

/* ── MODULE NAV ── */
.mod-nav{{background:#fff;border-radius:10px;padding:10px 16px;margin-bottom:24px;
  display:flex;align-items:center;gap:6px;flex-wrap:wrap;
  box-shadow:0 1px 4px rgba(0,0,0,.1);}}
.nav-tab{{padding:10px 18px;border:none;background:none;font-size:13px;font-weight:700;
  color:#718096;cursor:pointer;border-radius:7px;transition:all .15s;white-space:nowrap;}}
.nav-tab:hover{{color:#1a3a4a;background:#f0f4f8;}}
.nav-tab.active{{background:#1a3a4a;color:#fff;box-shadow:0 2px 8px rgba(15,32,39,.3);}}
.mod-section{{display:block;}}
body[data-theme="dark"] .mod-nav{{background:#1e293b;box-shadow:0 2px 8px rgba(0,0,0,.5);}}
body[data-theme="dark"] .nav-tab{{color:#94a3b8;}}
body[data-theme="dark"] .nav-tab:hover{{color:#e2e8f0;background:#334155;}}
body[data-theme="dark"] .nav-tab.active{{background:#3b82f6;color:#fff;}}

/* ── MOBILE ── */
html,body{{overflow-x:hidden;max-width:100%;box-sizing:border-box;}}
@media(max-width:640px){{
  .topbar{{padding:10px 12px;gap:8px;flex-wrap:wrap;}}
  .topbar-title{{font-size:15px;}}
  .topbar .periodo{{text-align:left;width:100%;}}
  .container{{padding:10px 8px;}}
  .filter-bar{{padding:8px 10px;gap:8px;}}
  .filter-group input[type=date],
  .filter-group select{{min-width:100px;font-size:12px;padding:6px 8px;}}
  .section-title{{font-size:12px;margin:14px 0 7px;}}
  .mod-nav{{padding:6px 8px;gap:3px;}}
  .nav-tab{{padding:7px 9px;font-size:11px;}}
  .kpi-card{{padding:10px 8px;}}
  .kpi-value{{font-size:16px;}}
  .kpi-label{{font-size:10px;}}
  .chart-card{{padding:12px 10px;}}
  .table-card{{padding:10px 6px;}}
  table.data-tbl{{font-size:11px;}}
  table.data-tbl thead th,
  table.data-tbl tbody td,
  table.data-tbl tfoot td{{padding:5px 7px;}}
  #btn-theme{{width:40px;height:40px;font-size:16px;bottom:12px;right:12px;}}
  #btn-fullscreen{{width:40px;height:40px;font-size:14px;bottom:12px;right:60px;}}
  .filter-info{{font-size:11px;}}
}}

</style>
</head>
<body data-theme="dark">

<button id="btn-theme" onclick="toggleTheme()" title="Alternar modo claro/escuro">🌙</button>
<button id="btn-fullscreen" onclick="abrirTelaCheia()" title="Abrir em tela cheia">&#x26F6;</button>


<!-- TOPBAR -->
<div class="topbar">
  <div style="display:flex;align-items:center;gap:18px;">
    {logo_tag}
    <div>
      <div class="topbar-title">📊 DASHBOARD — LOCVIX</div>
      <div class="sub">Análise de Vendas · Financeiro · Clientes · OS · Contratos · GestãoClick ERP</div>
    </div>
  </div>
  <div class="periodo">
    Período: <strong>{periodo}</strong><br/>
    <span style="color:#94a3b8;font-size:11px;">Gerado em {agora_str}</span><br/>
    <span style="color:#64748b;font-size:10px;">Desenvolvido por <strong style="color:#94a3b8">Fabrício Zamprogno</strong></span>
  </div>
</div>

<!-- FILTROS -->
<div class="filter-bar">
  <div class="filter-group">
    <label>📅 Data Início</label>
    <input type="date" id="fDateIni" value="{dt_min_iso}"/>
  </div>
  <div class="filter-group">
    <label>📅 Data Fim</label>
    <input type="date" id="fDateFim" value="{dt_max_iso}"/>
  </div>
  <div class="filter-sep"></div>
  <button class="btn btn-apply" onclick="aplicarFiltros()">▶ Aplicar</button>
  <button class="btn btn-clear" onclick="limparFiltros()">✕ Limpar</button>
  <div class="filter-sep"></div>
  <div class="filter-group">
    <label>🏪 Empresa</label>
    <div class="loja-pills">
      <button class="loja-pill active" data-loja="ambas" onclick="setLoja('ambas')">Ambas</button>
      <button class="loja-pill" data-loja="GJ" onclick="setLoja('GJ')">G&amp;J</button>
      <button class="loja-pill" data-loja="WA" onclick="setLoja('WA')">W&amp;A</button>
    </div>
  </div>
  <div class="filter-info" id="filtroInfo"></div>
</div>

<!-- Botão flutuante de Ajuda -->
<button class="fab-help" onclick="abrirAjuda()" title="Ajuda">?</button>

<!-- Modal Detalhe CC Manutenção -->
<div class="cc-overlay" id="ccOverlay" onclick="if(event.target===this)fecharCCModal()">
  <div class="cc-modal">
    <div class="cc-modal-header">
      <h2 id="ccModalTitle">🏷️ Discriminação dos Lançamentos</h2>
      <button class="hm-close" onclick="fecharCCModal()">&#10005;</button>
    </div>
    <div class="cc-modal-body" id="ccModalBody"></div>
  </div>
</div>

<!-- Modal de Ajuda -->
<div class="help-overlay" id="helpOverlay" onclick="if(event.target===this)fecharAjuda()">
  <div class="help-modal">
    <div class="hm-header">
      <div><h2>&#128218; Guia de Uso — Dashboard LOCVIX</h2><small>Aprenda a usar cada módulo</small></div>
      <button class="hm-close" onclick="fecharAjuda()">&#10005;</button>
    </div>
    <div class="hm-tabs">
      <button class="hm-tab active" onclick="setHmTab('inicio')">&#127968; Início</button>
      <button class="hm-tab" onclick="setHmTab('filtros')">&#128269; Filtros</button>
      <button class="hm-tab" onclick="setHmTab('vendas')">&#128176; Vendas</button>
      <button class="hm-tab" onclick="setHmTab('financeiro')">&#128179; Financeiro</button>
      <button class="hm-tab" onclick="setHmTab('operacoes')">&#128295; Operações</button>
      <button class="hm-tab" onclick="setHmTab('manutencao')">&#128736; Manutenção</button>
      <button class="hm-tab" onclick="setHmTab('ponto')">&#128336; Ponto</button>
      <button class="hm-tab" onclick="setHmTab('orcamento')">&#128203; Orçamento</button>
    </div>
    <div class="hm-body">

      <div class="hm-panel active" id="hmp-inicio">
        <h3>&#127968; Bem-vindo ao Dashboard LOCVIX</h3>
        <p>Este painel centraliza todos os dados operacionais e financeiros da Locvix, integrado em tempo real com o ERP <strong>GestãoClick</strong>.</p>
        <div class="hm-step"><div class="hm-snum">1</div><div class="hm-stxt"><strong>Defina o período</strong> — Na barra lateral esquerda do Streamlit, escolha a Data Início e Data Fim para analisar.</div></div>
        <div class="hm-step"><div class="hm-snum">2</div><div class="hm-stxt"><strong>Clique em “Atualizar Dados”</strong> — O sistema busca os dados mais recentes do GestãoClick. Aguarde o carregamento.</div></div>
        <div class="hm-step"><div class="hm-snum">3</div><div class="hm-stxt"><strong>Escolha o módulo</strong> — Navegue pelas abas no topo do dashboard: Vendas, Financeiro, Operações, Ponto, Orçamento.</div></div>
        <div class="hm-step"><div class="hm-snum">4</div><div class="hm-stxt"><strong>Use os filtros</strong> — Ajuste datas e empresa na barra superior para refinar os dados exibidos.</div></div>
        <div class="hm-tip"><strong>&#128161; Dica:</strong> Use os botões <strong>G&amp;J</strong> / <strong>W&amp;A</strong> / <strong>Ambas</strong> para ver os dados de cada empresa separadamente ou consolidados.</div>
      </div>

      <div class="hm-panel" id="hmp-filtros">
        <h3>&#128269; Como usar os Filtros</h3>
        <p>A <strong>barra de filtros</strong> fica fixada no topo e afeta todos os módulos ao mesmo tempo.</p>
        <ul>
          <li><strong>&#128197; Data Início / Data Fim</strong> — Define o período de análise. Após alterar, clique em <em>“&#9654; Aplicar”</em>.</li>
          <li><strong>&#9654; Aplicar</strong> — Confirma os filtros e recalcula gráficos e tabelas.</li>
          <li><strong>&#10005; Limpar</strong> — Restaura o período original carregado.</li>
          <li><strong>&#127978; Empresa</strong> — Filtra por empresa: <em>Ambas</em> consolida G&amp;J e W&amp;A.</li>
        </ul>
        <hr class="hm-hr"/>
        <h3>Como atualizar os dados</h3>
        <p>Na tela principal, clique em <strong>“&#128260; Atualizar Dados”</strong>. Isso força nova busca na API, ignorando o cache local.</p>
        <div class="hm-tip"><strong>&#9203; Tempo:</strong> O carregamento pode levar de 30 segundos a 2 minutos dependendo do volume de dados.</div>
      </div>

      <div class="hm-panel" id="hmp-vendas">
        <h3>&#128176; Módulo Vendas</h3>
        <p>Exibe todas as vendas (pedidos/NFs) registradas no GestãoClick no período selecionado.</p>
        <ul>
          <li><strong>Fat. Líquido</strong> — Valor após descontos.</li>
          <li><strong>Fat. Bruto</strong> — Valor antes dos descontos.</li>
          <li><strong>Desconto Total</strong> — Soma de todos os descontos aplicados.</li>
          <li><strong>Qtd Vendida</strong> — Total de itens/serviços vendidos.</li>
          <li><strong>Clientes Ativos</strong> — Quantidade de clientes distintos com venda.</li>
          <li><strong>Pedidos / NFs</strong> — Número de documentos emitidos.</li>
        </ul>
        <hr class="hm-hr"/>
        <p>Gráficos disponíveis: <strong>Faturamento Mensal</strong>, <strong>por Categoria</strong>, <strong>Top 10 Produtos</strong> e <strong>Top 10 Clientes</strong>.</p>
        <div class="hm-tip"><strong>&#128161; Dica:</strong> Clique em uma barra do gráfico de Categorias para filtrar a tabela por aquela categoria.</div>
      </div>

      <div class="hm-panel" id="hmp-financeiro">
        <h3>&#128179; Módulo Financeiro</h3>
        <p>Exibe as <strong>Contas a Pagar</strong> com análise por categoria e centro de custo.</p>
        <ul>
          <li><strong>A Pagar (Total)</strong> — Soma de todos os valores a pagar no período.</li>
          <li><strong>A Pagar (Pago)</strong> — Parcela já liquidada.</li>
        </ul>
        <p>Filtros adicionais: <strong>Plano de Contas</strong> e <strong>Centro de Custo</strong> (departamento/projeto).</p>
        <p>Gráficos: <strong>Distribuição por Categoria</strong> (donut), <strong>Despesas Mensais</strong> e <strong>Resultado Mensal</strong> (receitas vs. despesas).</p>
        <div class="hm-tip"><strong>&#128161; Dica:</strong> Use o filtro de Centro de Custo para identificar quais setores estão gerando mais despesas.</div>
      </div>

      <div class="hm-panel" id="hmp-operacoes">
        <h3>&#128295; Módulo Operações</h3>
        <p>Concentra as informações operacionais: <strong>Ordens de Serviço</strong>, <strong>Contratos</strong>, <strong>Boletins de Medição</strong> e <strong>Horas do App</strong>.</p>
        <ul>
          <li><strong>Ordens de Serviço (OS)</strong> — Lista OS abertas e concluídas, com técnico, status e valor.</li>
          <li><strong>Contratos</strong> — Contratos ativos: cliente, valor e vigência.</li>
          <li><strong>Boletins de Medição</strong> — Registro das medições de serviços executados.</li>
          <li><strong>Horas do App</strong> — Horas registradas via app mobile pelos operadores.</li>
        </ul>
        <div class="hm-tip"><strong>&#128161; Dica:</strong> O status das OS usa cores: <span style="color:#059669">&#9632;</span> Concluída, <span style="color:#d97706">&#9632;</span> Em andamento, <span style="color:#dc2626">&#9632;</span> Cancelada.</div>
      </div>

      <div class="hm-panel" id="hmp-manutencao">
        <h3>&#128736; Módulo Manutenção</h3>
        <p>Este módulo está em desenvolvimento e em breve exibirá informações sobre manutenções preventivas e corretivas dos equipamentos da Locvix.</p>
        <div class="hm-tip"><strong>&#128161; Em breve:</strong> Histórico de manutenções, custos, equipamentos e técnicos responsáveis.</div>
      </div>

      <div class="hm-panel" id="hmp-ponto">
        <h3>&#128336; Módulo Ponto Colaborador</h3>
        <p>Exibe os registros de ponto integrados ao sistema <strong>DixiPonto</strong>.</p>
        <ul>
          <li><strong>Listagem por Colaborador</strong> — Entradas, saídas e total de horas por dia.</li>
          <li><strong>Filtro por Nome</strong> — Digite o nome do colaborador para filtrar.</li>
          <li><strong>Total de Horas</strong> — Somatório de horas trabalhadas no período.</li>
        </ul>
        <div class="hm-step"><div class="hm-snum">1</div><div class="hm-stxt">Ajuste o <strong>período nas datas</strong> da barra superior.</div></div>
        <div class="hm-step"><div class="hm-snum">2</div><div class="hm-stxt">Use o campo de busca para <strong>filtrar por colaborador</strong>.</div></div>
        <div class="hm-step"><div class="hm-snum">3</div><div class="hm-stxt">A tabela mostra <strong>data, entrada, saída e total de horas</strong>.</div></div>
        <div class="hm-tip"><strong>&#9888; Atenção:</strong> Se não aparecer nada, verifique se o período selecionado possui registros no DixiPonto.</div>
      </div>

      <div class="hm-panel" id="hmp-orcamento">
        <h3>&#128203; Módulo Orçamento</h3>
        <p>Exibe todos os <strong>Orçamentos</strong> cadastrados no GestãoClick com análise por situação e valores.</p>
        <ul>
          <li><strong>Total de Propostas</strong> — Quantidade de orçamentos no período.</li>
          <li><strong>Em Aberto</strong> — Orçamentos aguardando resposta do cliente.</li>
          <li><strong>Em Andamento</strong> — Orçamentos em negociação ou execução.</li>
          <li><strong>Concretizado</strong> — Orçamentos aprovados e fechados.</li>
          <li><strong>Cancelado</strong> — Orçamentos cancelados pelo cliente ou empresa.</li>
          <li><strong>Valor Total</strong> — Soma de todos os orçamentos do período.</li>
        </ul>
        <p>Filtros rápidos por situação: <strong>Todas / Em Aberto / Em Andamento / Concretizado / Cancelado</strong>.</p>
        <div class="hm-tip"><strong>&#128161; Dica:</strong> Acompanhe a evolução dos orçamentos por vendedor no gráfico de barras e a distribuição geral no gráfico de rosca.</div>
      </div>

    </div>
    <div class="hm-footer">
      <span>Dashboard LOCVIX — Sistema integrado GestãoClick</span>
      <button class="hm-btn" onclick="fecharAjuda()">Fechar</button>
    </div>
  </div>
</div>

<div class="container">

  <!-- MODULE NAVIGATION -->
  <div class="mod-nav" id="modNav">
    <button class="nav-tab active" data-mod="geral" onclick="setModulo('geral')">📊 Visão Geral</button>
    <button class="nav-tab" data-mod="vendas" onclick="setModulo('vendas')">💰 Vendas</button>
    <button class="nav-tab" data-mod="financeiro" onclick="setModulo('financeiro')">💳 Financeiro</button>
    <button class="nav-tab" data-mod="operacoes" onclick="setModulo('operacoes')">🔧 Operações</button>
    <button class="nav-tab" data-mod="manutencao" onclick="setModulo('manutencao')">🛠 Manutenção</button>
    <button class="nav-tab" data-mod="ponto" onclick="setModulo('ponto')">🕐 Ponto Colaborador</button>
    <button class="nav-tab" data-mod="orcamento" onclick="setModulo('orcamento')">📋 Orçamento</button>
  </div>

  <div class="mod-section" data-mod="vendas">
  <!-- ── KPIs DE VENDAS ── -->
  <div class="section-title">💰 Vendas — Resumo do Período</div>
  <div class="kpi-grid col6">
    <div class="kpi-card green">
      <div class="kpi-label">Fat. Líquido</div>
      <div class="kpi-value small" id="kLiq">—</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Fat. Bruto</div>
      <div class="kpi-value small" id="kBruto">—</div>
    </div>
    <div class="kpi-card orange">
      <div class="kpi-label">Desconto Total</div>
      <div class="kpi-value small" id="kDesc">—</div>
    </div>
    <div class="kpi-card blue">
      <div class="kpi-label">Qtd Vendida</div>
      <div class="kpi-value" id="kQtd">—</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Clientes Ativos</div>
      <div class="kpi-value" id="kCli">—</div>
    </div>
    <div class="kpi-card teal">
      <div class="kpi-label">Pedidos / NFs</div>
      <div class="kpi-value" id="kNFs">—</div>
    </div>
  </div>

  </div><!-- /mod vendas-kpi -->

  <div class="mod-section" data-mod="financeiro">
  <!-- ── KPIs FINANCEIRO ── -->
  <div class="section-title">🏦 Financeiro — Resumo</div>
  <div class="kpi-grid col4">
    <div class="kpi-card red">
      <div class="kpi-label">A Pagar (Total)</div>
      <div class="kpi-value small" id="kPagTotal">—</div>
    </div>
    <div class="kpi-card orange">
      <div class="kpi-label">A Pagar (Pago)</div>
      <div class="kpi-value small" id="kPagPago">—</div>
    </div>
  </div>

  </div><!-- /mod financeiro-kpi -->

  <div class="mod-section" data-mod="operacoes">

  <!-- ── BOLETINS DE MEDIÇÃO ── -->
  <div class="section-title" style="margin-top:28px;">📋 Boletins de Medição</div>
  <!-- KPIs Medição -->
  <div class="kpi-grid col5" id="secMedicaoKpis">
    <div class="kpi-card blue">
      <div class="kpi-label">Boletins</div>
      <div class="kpi-value" id="kMedTotal">—</div>
    </div>
    <div class="kpi-card teal">
      <div class="kpi-label">Horas Diurno</div>
      <div class="kpi-value small" id="kMedHrDia">—</div>
    </div>
    <div class="kpi-card purple">
      <div class="kpi-label">Horas Noturno</div>
      <div class="kpi-value small" id="kMedHrNot">—</div>
    </div>
    <div class="kpi-card orange">
      <div class="kpi-label">Horas Extra</div>
      <div class="kpi-value small" id="kMedHrExt">—</div>
    </div>
    <div class="kpi-card green">
      <div class="kpi-label">Total Faturado</div>
      <div class="kpi-value small" id="kMedTotal$">—</div>
    </div>
  </div>
  <!-- Gráficos Medição -->
  <div class="chart-row col2" style="align-items:start;margin-top:16px;" id="secMedicaoCharts">
    <div class="chart-card">
      <h3>🏗️ Horas por Equipamento</h3>
      <div style="position:relative;height:240px;">
        <canvas id="chartMedHoras"></canvas>
      </div>
    </div>
    <div class="chart-card">
      <h3>💰 Faturamento por Boletim</h3>
      <div style="position:relative;height:240px;">
        <canvas id="chartMedFat"></canvas>
      </div>
    </div>
  </div>
  <!-- Tabela Medição -->
  <div class="table-card" style="margin-top:16px;" id="secMedicaoTabela">
    <h3>📄 Detalhamento dos Boletins</h3>
    <div style="overflow-x:auto;">
      <table class="data-tbl" id="tblMedicoes">
        <thead>
          <tr>
            <th>Nº Boletim</th><th>Equipamento</th><th>Cliente</th><th>Período</th>
            <th class="num">H.Diurno</th><th class="num">H.Noturno</th><th class="num">H.Extra</th>
            <th class="num">Vl.Hora Dia</th><th class="num">Vl.Hora Not</th>
            <th class="num">Turno Diurno</th><th class="num">Turno Noturno</th>
            <th class="num">Horas Extra</th><th class="num">Desmobilização</th>
            <th class="num">Total Medição</th><th class="num">Total a Pagar</th>
          </tr>
        </thead>
        <tbody id="tblMedicoesBdy"></tbody>
        <tfoot id="tblMedicoesFoot"></tfoot>
      </table>
    </div>
  </div>

  <div id="secMedicaoVazio" style="display:none;padding:32px;text-align:center;color:#94a3b8;font-size:15px;">
    📂 Nenhum boletim de medição encontrado na pasta <code>MEDIÇÃO/</code>
  </div>

  <!-- ── HORAS REGISTRADAS PELO APP ── -->
  <div class="section-title" style="margin-top:36px;">📱 Horas Registradas pelo App</div>

  <!-- KPIs Horas App -->
  <div class="kpi-grid col5" id="secHorasAppKpis">
    <div class="kpi-card blue">
      <div class="kpi-label">Registros</div>
      <div class="kpi-value" id="kAppTotal">—</div>
    </div>
    <div class="kpi-card teal">
      <div class="kpi-label">Total de Horas</div>
      <div class="kpi-value small" id="kAppHoras">—</div>
    </div>
    <div class="kpi-card orange">
      <div class="kpi-label">Turno Diurno</div>
      <div class="kpi-value small" id="kAppDiurno">—</div>
    </div>
    <div class="kpi-card purple">
      <div class="kpi-label">Turno Noturno</div>
      <div class="kpi-value small" id="kAppNoturno">—</div>
    </div>
    <div class="kpi-card green">
      <div class="kpi-label">Operadores</div>
      <div class="kpi-value" id="kAppOperadores">—</div>
    </div>
  </div>

  <!-- Gráficos Horas App -->
  <div class="chart-row col2" style="align-items:start;margin-top:16px;" id="secHorasAppCharts">
    <div class="chart-card">
      <h3>🛠️ Horas por Equipamento</h3>
      <div style="position:relative;height:240px;">
        <canvas id="chartAppEquip"></canvas>
      </div>
    </div>
    <div class="chart-card">
      <h3>👷 Horas por Operador</h3>
      <div style="position:relative;height:240px;">
        <canvas id="chartAppOper"></canvas>
      </div>
    </div>
  </div>

  <!-- Tabela Horas App -->
  <div class="table-card" style="margin-top:16px;" id="secHorasAppTabela">
    <h3>🗓️ Detalhe dos Registros</h3>
    <div style="overflow-x:auto;">
      <table class="data-tbl" id="tblHorasApp">
        <thead>
          <tr>
            <th>Data</th><th>Equipamento</th><th>Placa</th><th>Cliente</th>
            <th>Operador</th><th>Turno</th>
            <th class="num">Entrada</th><th class="num">Saída</th>
            <th class="num">Horas</th><th>Status</th><th>Observações</th>
          </tr>
        </thead>
        <tbody id="tblHorasAppBdy"></tbody>
        <tfoot id="tblHorasAppFoot"></tfoot>
      </table>
    </div>
  </div>

  <div id="secHorasAppVazio" style="display:none;padding:32px;text-align:center;color:#94a3b8;font-size:15px;">
    📱 Nenhum registro de horas encontrado para o período selecionado.
  </div>

  </div><!-- /mod operacoes -->

  <div class="mod-section" data-mod="manutencao">
    <div class="section-title">🛠 Manutenção Preventiva — Equipamentos</div>

    <!-- Formulário de registro -->
    <div class="chart-card" style="margin-bottom:20px;">
      <h3 style="margin-bottom:12px;">🔧 Registrar / Atualizar Manutenção</h3>
      <p style="font-size:13px;color:#64748b;margin:0 0 14px">Preencha abaixo para registrar a data da última manutenção. O próximo ciclo será calculado automaticamente (2 meses).</p>
      <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end;">
        <div style="flex:2;min-width:180px;">
          <label style="font-size:12px;font-weight:600;color:#64748b;display:block;margin-bottom:4px">Equipamento / Centro de Custo</label>
          <select id="mFormEquip"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;box-sizing:border-box;background:#fff;cursor:pointer;">
            <option value="">— Selecione o equipamento —</option>
          </select>
        </div>
        <div style="flex:1;min-width:140px;">
          <label style="font-size:12px;font-weight:600;color:#64748b;display:block;margin-bottom:4px">Data da última manutenção</label>
          <input id="mFormData" type="date"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;box-sizing:border-box;">
        </div>
        <div style="flex:2;min-width:180px;">
          <label style="font-size:12px;font-weight:600;color:#64748b;display:block;margin-bottom:4px">Tipo de serviço</label>
          <input id="mFormServico" type="text" placeholder="Ex: Troca de óleo"
            style="width:100%;padding:8px 10px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;box-sizing:border-box;">
        </div>
        <div style="flex:none;display:flex;gap:8px;">
          <button onclick="salvarManutencao()"
            style="background:#1e3a5f;color:#fff;border:none;border-radius:6px;padding:9px 22px;font-size:13px;font-weight:700;cursor:pointer;white-space:nowrap;">
            💾 Salvar
          </button>
          <button onclick="deletarManutencao()"
            style="background:#dc2626;color:#fff;border:none;border-radius:6px;padding:9px 16px;font-size:13px;font-weight:700;cursor:pointer;white-space:nowrap;">
            🗑 Excluir
          </button>
        </div>
      </div>
      <div id="mFormMsg" style="margin-top:10px;font-size:13px;display:none;"></div>
    </div>

    <!-- KPIs Manutenção -->
    <div class="kpi-grid col3">
      <div class="kpi-card red">
        <div class="kpi-label">🔴 Vencidas</div>
        <div class="kpi-value" id="kManutVencidas">—</div>
      </div>
      <div class="kpi-card orange">
        <div class="kpi-label">⚠️ Próximas (≤5 dias)</div>
        <div class="kpi-value" id="kManutProximas">—</div>
      </div>
      <div class="kpi-card green">
        <div class="kpi-label">✅ Em Dia</div>
        <div class="kpi-value" id="kManutOk">—</div>
      </div>
    </div>

    <!-- Tabela de status por equipamento -->
    <div class="table-card" style="margin-top:16px;">
      <h3>📋 Status de Manutenção por Equipamento (ciclo a cada 2 meses)</h3>
      <div style="overflow-x:auto;">
        <table class="data-tbl" id="tblManutencao">
          <thead>
            <tr>
              <th>Equipamento / Centro de Custo</th>
              <th>Status</th>
              <th>Última Manutenção</th>
              <th>Próxima Manutenção</th>
              <th class="num">Dias Restantes</th>
            </tr>
          </thead>
          <tbody id="tblManutencaoBdy"></tbody>
        </table>
      </div>
      <div id="manutVazioMsg" style="display:none;padding:24px;text-align:center;color:#94a3b8;font-size:14px;">
        Nenhum equipamento encontrado nas despesas do período.
      </div>
    </div>

    <!-- Gráfico de despesas por CC -->
    <div class="chart-row" style="align-items:start;margin-top:24px;">
      <div class="chart-card" style="flex:1;">
        <h3>🏷️ Despesas por Centro de Custo (Equipamentos)</h3>
        <div id="wrapManutCC" style="position:relative;height:320px;"><canvas id="chartManutCC"></canvas></div>
      </div>
    </div>
  </div><!-- /mod manutencao -->

  <div class="mod-section" data-mod="vendas">
  <!-- ── GRÁFICOS DE VENDAS ── -->
  <div class="section-title">📈 Análise de Vendas</div>
  <div class="chart-row col2" style="align-items:start;">
    <div class="chart-card">
      <h3>📅 Faturamento Líquido por Mês</h3>
      <div style="position:relative;height:260px;">
        <canvas id="chartMensal"></canvas>
      </div>
    </div>
    <div class="chart-card">
      <h3>🍩 Faturamento por Categoria</h3>
      <div style="position:relative;height:260px;">
        <canvas id="chartCategoria"></canvas>
      </div>
    </div>
  </div>
  <div class="chart-row col2" style="align-items:start;">
    <div class="chart-card">
      <h3>🏆 Top 10 Produtos — Fat. Líquido</h3>
      <div style="position:relative;height:260px;">
        <canvas id="chartProdutos"></canvas>
      </div>
    </div>
    <div class="chart-card">
      <h3>👥 Top 10 Clientes — Fat. Líquido</h3>
      <div style="position:relative;height:260px;">
        <canvas id="chartClientes"></canvas>
      </div>
    </div>
  </div>
  <div class="chart-row" style="align-items:start;">
    <div class="chart-card" style="flex:1;">
      <h3>🏷️ Faturamento Líquido por Centro de Custo</h3>
      <div id="wrapVendasCC" style="position:relative;height:320px;"><canvas id="chartVendasCC"></canvas></div>
    </div>
  </div>

  </div><!-- /mod vendas-charts -->

  <div class="mod-section" data-mod="financeiro">
  <!-- ── FILTROS FINANCEIRO ── -->
  <div class="section-title">💳 Financeiro</div>
  <div class="fin-filter-bar" id="finFilterBar">
    <div class="filter-group">
      <label>🏷 Plano de Contas</label>
      <select id="fFinCat" onchange="aplicarFiltrosFinanceiro()">
        <option value="">— Todas —</option>
      </select>
    </div>
    <div class="filter-group">
      <label>🏢 Centro de Custo</label>
      <select id="fFinCC" onchange="aplicarFiltrosFinanceiro()">
        <option value="">— Todos —</option>
      </select>
    </div>
    <button class="btn btn-clear" style="align-self:flex-end" onclick="limparFiltrosFinanceiro()">✕ Limpar</button>
  </div>
  <div class="chart-row col2" style="align-items:start;">
    <div class="chart-card">
      <h3> Contas a Pagar por Plano de Contas</h3>
      <div style="position:relative;height:260px;">
        <canvas id="chartPagar"></canvas>
      </div>
    </div>
    <div class="chart-card">
      <h3>📅 Contas a Pagar — Vencimentos por Mês (±12 meses)</h3>
      <div style="position:relative;height:260px;">
        <canvas id="chartPagarMensal"></canvas>
      </div>
    </div>
  </div>
  <div class="chart-row" style="align-items:start;">
    <div class="chart-card" style="flex:1;">
      <h3>🏷️ Despesas por Centro de Custo</h3>
      <div id="wrapPagarCC" style="position:relative;height:260px;"><canvas id="chartPagarCC"></canvas></div>
    </div>
  </div>
  <div class="chart-row" style="align-items:start;">
    <div class="chart-card" style="flex:1;">
      <h3>📈 Receitas vs Despesas por Mês — Resultado</h3>
      <div style="position:relative;height:320px;"><canvas id="chartResultado"></canvas></div>
    </div>
  </div>

  </div><!-- /mod financeiro-charts -->

  <div class="mod-section" data-mod="vendas">
  <!-- ── TABELA TOP PRODUTOS ── -->
  <div class="section-title">📋 Top 20 Produtos</div>
  <div class="table-card">
    <h3>Ranking de produtos por faturamento líquido</h3>
    <table class="data-tbl">
      <thead>
        <tr>
          <th>#</th>
          <th>Código</th>
          <th>Produto</th>
          <th>Categoria</th>
          <th class="num">Qtd</th>
          <th class="num">Fat. Bruto</th>
          <th class="num">Desconto</th>
          <th class="num">Fat. Líquido</th>
          <th class="num">Part. %</th>
        </tr>
      </thead>
      <tbody id="tblProdCorpo"></tbody>
      <tfoot><tr id="tblProdTotal"></tr></tfoot>
    </table>
  </div>

  <!-- ── TABELA TOP CLIENTES ── -->
  <div class="section-title">👥 Top 20 Clientes</div>
  <div class="table-card">
    <h3>Ranking de clientes por faturamento líquido</h3>
    <table class="data-tbl">
      <thead>
        <tr>
          <th>#</th>
          <th>Cliente</th>
          <th>Vendedor</th>
          <th class="num">Pedidos</th>
          <th class="num">Fat. Bruto</th>
          <th class="num">Desconto</th>
          <th class="num">Fat. Líquido</th>
          <th class="num">Part. %</th>
        </tr>
      </thead>
      <tbody id="tblCliCorpo"></tbody>
      <tfoot><tr id="tblCliTotal"></tr></tfoot>
    </table>
  </div>
  </div><!-- /mod vendas-tables -->

  <!-- ── PONTO COLABORADOR ── -->
  <div class="mod-section" data-mod="ponto">
  <div class="section-title" id="tituloPonto">🕐 Ponto Colaborador — {periodo}</div>
  <div class="kpi-grid col6">
    <div class="kpi-card blue"><div class="kpi-label">Funcionários</div><div class="kpi-value" id="kPontoFunc">—</div></div>
    <div class="kpi-card green"><div class="kpi-label">Marcações no Período</div><div class="kpi-value" id="kPontoTotal">—</div></div>
    <div class="kpi-card teal"><div class="kpi-label">Com Registro Hoje</div><div class="kpi-value" id="kPontoHoje">—</div></div>
    <div class="kpi-card orange"><div class="kpi-label">Sem Registro Hoje</div><div class="kpi-value" id="kPontoAus">—</div></div>
    <div class="kpi-card purple"><div class="kpi-label">✏️ Ajustes Manuais</div><div class="kpi-value" id="kPontoManual">—</div></div>
    <div class="kpi-card red"><div class="kpi-label">⏰ HE Hoje</div><div class="kpi-value" id="kPontoHE">—</div></div>
  </div>
  <div class="chart-row col2" style="align-items:start;">
    <div class="chart-card"><h3>⚠️ Ausências por Funcionário — Faltas e Saídas Antecipadas (horas)</h3><div id="wrapPontoAus" style="position:relative;height:280px;"><canvas id="chartPontoAus"></canvas></div></div>
    <div class="chart-card"><h3>👤 Marcações por Funcionário</h3><div style="position:relative;height:280px;"><canvas id="chartPontoFunc"></canvas></div></div>
  </div>
  <div class="chart-row col2" style="align-items:start;">
    <div class="chart-card"><h3>⏰ Horas Regulares vs Hora Extra (período)</h3><div id="wrapPontoHoras" style="position:relative;height:260px;"><canvas id="chartPontoHoras"></canvas></div></div>
    <div class="chart-card"><h3>⏰ Ranking Hora Extra no Período</h3><div id="wrapPontoHoraExtra" style="position:relative;height:260px;"><canvas id="chartPontoHoraExtra"></canvas></div></div>
  </div>
  <div class="section-title">📋 Resumo do Dia de Hoje</div>
  <div class="table-card">
    <h3>Marcações de hoje por funcionário</h3>
    <table class="data-tbl">
      <thead><tr>
        <th>#</th><th>Funcionário</th>
        <th class="num">1ª Entrada</th><th class="num">Saída Almoço</th><th class="num">Retorno</th><th class="num">Última Saída</th>
        <th class="num">H. Trabalhadas</th><th>Status</th><th>Origem</th>
      </tr></thead>
      <tbody id="tblPontoHoje"></tbody>
    </table>
  </div>
  <div class="section-title">✏️ Motivos dos Ajustes Manuais</div>
  <div class="table-card">
    <h3>Registros manuais agrupados por justificativa no período</h3>
    <table class="data-tbl">
      <thead><tr>
        <th>#</th><th>Motivo</th><th class="num">Qtd</th><th>Funcionários</th>
      </tr></thead>
      <tbody id="tblPontoJustif"></tbody>
    </table>
  </div>
  </div><!-- /mod ponto -->

  <!-- ═══════════════════════════════════════════ ORÇAMENTO ═══ -->
  <div class="mod-section" data-mod="orcamento" style="display:none">
  <div class="section-title">📋 Orçamento — Propostas Comerciais</div>

  <!-- KPIs -->
  <div class="kpi-grid col6">
    <div class="kpi-card blue">
      <div class="kpi-label">Total Propostas</div>
      <div class="kpi-value" id="kOrcTotal">—</div>
    </div>
    <div class="kpi-card yellow">
      <div class="kpi-label">📂 Em Aberto</div>
      <div class="kpi-value" id="kOrcAberto">—</div>
    </div>
    <div class="kpi-card teal">
      <div class="kpi-label">⏳ Em Andamento</div>
      <div class="kpi-value" id="kOrcAndamento">—</div>
    </div>
    <div class="kpi-card green">
      <div class="kpi-label">✅ Concretizado</div>
      <div class="kpi-value" id="kOrcConc">—</div>
    </div>
    <div class="kpi-card red">
      <div class="kpi-label">❌ Cancelado</div>
      <div class="kpi-value" id="kOrcCanc">—</div>
    </div>
    <div class="kpi-card purple">
      <div class="kpi-label">Valor Total</div>
      <div class="kpi-value small" id="kOrcValTotal">—</div>
    </div>
  </div>

  <!-- Gráficos -->
  <div class="chart-row" style="align-items:start;">
    <div class="chart-card" style="flex:1;">
      <h3>📊 Propostas por Vendedor</h3>
      <div style="position:relative;height:280px;"><canvas id="chartOrcVendedor"></canvas></div>
    </div>
    <div class="chart-card" style="flex:0 0 280px;">
      <h3>🎯 Situação dos Orçamentos</h3>
      <div style="position:relative;height:280px;"><canvas id="chartOrcTaxa"></canvas></div>
    </div>
  </div>

  <!-- Filtro de status -->
  <div class="section-title">🔍 Propostas Detalhadas</div>
  <div style="margin-bottom:10px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;">
    <span style="color:#94a3b8;font-size:13px;">Filtrar:</span>
    <button class="btn-filter-orc active" data-f="todos"       onclick="filtrarOrc(this)">Todas</button>
    <button class="btn-filter-orc"        data-f="Em aberto"   onclick="filtrarOrc(this)">📂 Em Aberto</button>
    <button class="btn-filter-orc"        data-f="Em andamento" onclick="filtrarOrc(this)">⏳ Em Andamento</button>
    <button class="btn-filter-orc"        data-f="Concretizado" onclick="filtrarOrc(this)">✅ Concretizado</button>
    <button class="btn-filter-orc"        data-f="Cancelado"   onclick="filtrarOrc(this)">❌ Cancelado</button>
  </div>

  <div class="table-card">
    <table class="data-tbl" id="tblOrcamentos">
      <thead><tr>
        <th>#</th>
        <th>Nº</th>
        <th>Data</th>
        <th>Cliente</th>
        <th>Vendedor</th>
        <th>Centro de Custo</th>
        <th>Loja</th>
        <th class="num">Valor Total</th>
        <th>Situação</th>
        <th style="text-align:center">PDF</th>
      </tr></thead>
      <tbody id="tblOrcBody"></tbody>
    </table>
  </div>
  </div><!-- /mod orcamento -->

</div>

<div class="footer">
  <div class="footer-dev">
    <div>
      <div>Desenvolvido por <strong>Fabrício Zamprogno</strong></div>
    </div>
  </div>
  <div class="footer-sep"></div>
  <div class="footer-gen">
    Dashboard Locvix — GestãoClick ERP<br/>
    Gerado em {agora_str}
  </div>
</div>

<!-- dados embutidos como JSON puro — nunca executados como JavaScript -->
<script type="application/json" id="_dVENDAS">{jv(raw_vendas)}</script>
<script type="application/json" id="_dRECEBER">{jv(raw_rec)}</script>
<script type="application/json" id="_dPAGAR">{jv(raw_pag)}</script>
<script type="application/json" id="_dPAGAR_ALL">{jv(raw_pag_all)}</script>
<script type="application/json" id="_dOS_LIST">{jv(raw_os)}</script>
<script type="application/json" id="_dCONTRATOS">{jv(raw_contr)}</script>
<script type="application/json" id="_dPONTO_FUNC">{jv(ponto_func)}</script>
<script type="application/json" id="_dPONTO_MARC">{jv(ponto_marc)}</script>
<script type="application/json" id="_dORCAMENTOS">{jv(raw_orc)}</script>
<script type="application/json" id="_dMEDICOES">{jv(raw_med)}</script>
<script type="application/json" id="_dHORAS_APP">{jv(raw_horas_app)}</script>
<script type="application/json" id="_dMANUTENCAO">{jv(raw_manutencoes)}</script>

<script>
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  DADOS BRUTOS (lidos via JSON.parse - nunca como JS literal)
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
function _pd(id) {{
  const el = document.getElementById(id);
  if (!el) return [];
  try {{ return JSON.parse(el.textContent); }} catch(e) {{ console.warn('_pd', id, e); return []; }}
}}
const VENDAS    = _pd('_dVENDAS');
const RECEBER   = _pd('_dRECEBER');
const PAGAR     = _pd('_dPAGAR');
const PAGAR_ALL = _pd('_dPAGAR_ALL');
const OS_LIST   = _pd('_dOS_LIST');
const CONTRATOS = _pd('_dCONTRATOS');
const PONTO_FUNC = _pd('_dPONTO_FUNC');
const PONTO_MARC = _pd('_dPONTO_MARC');
const ORCAMENTOS = _pd('_dORCAMENTOS');
const MEDICOES   = _pd('_dMEDICOES');
const HORAS_APP  = _pd('_dHORAS_APP');
const MANUTENCAO = _pd('_dMANUTENCAO');
const PERIODO_INI = '{ponto_d_ini_iso}';  // yyyy-mm-dd do per\u00EDodo selecionado
const PERIODO_FIM = '{ponto_d_fim_iso}';
const _SB_URL  = '{supabase_url}';
const _SB_ANON = '{supabase_anon}';

const BRL = v => 'R$\u00a0' + v.toLocaleString('pt-BR',{{minimumFractionDigits:2,maximumFractionDigits:2}});
const NUM = v => v.toLocaleString('pt-BR');
const PCT = (a,b) => b > 0 ? (a/b*100).toFixed(1)+'%' : '\u2014';

// \u2500\u2500 Cores \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
const CORES = ['#1a3a4a','#0891b2','#059669','#d97706','#7c3aed',
               '#2563eb','#dc2626','#0d9488','#65a30d','#ea580c',
               '#db2777','#6366f1','#78716c','#0369a1','#15803d'];

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  ESTADO DE FILTROS
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
let dadosFilt    = VENDAS;
let pontoMarcFilt = PONTO_MARC;
let pagarFilt    = PAGAR;
let pagarFiltFin = PAGAR;  // filtro interno do m\u00F3dulo Financeiro (cat + cc)
let lojaFiltAtivo = 'ambas';  // 'ambas' | 'GJ' | 'WA'

function setLoja(loja) {{
  lojaFiltAtivo = loja;
  document.querySelectorAll('.loja-pill').forEach(b => {{
    b.classList.toggle('active', b.dataset.loja === loja);
  }});
  filtrar();
  atualizar();
}}

// \u2500\u2500 Popula os selects de filtro do m\u00F3dulo Financeiro
function _populaFiltrosFinanceiro() {{
  const cats = [...new Set(PAGAR.map(r => r.cat || '').filter(v => v))].sort();
  const ccs  = [...new Set(PAGAR.map(r => r.cc  || '').filter(v => v))].sort();
  const elCat = document.getElementById('fFinCat');
  const elCC  = document.getElementById('fFinCC');
  if (elCat && elCat.options.length <= 1) {{
    cats.forEach(c => {{ const o = new Option(c, c); elCat.add(o); }});
  }}
  if (elCC && elCC.options.length <= 1) {{
    ccs.forEach(c => {{ const o = new Option(c, c); elCC.add(o); }});
  }}
}}

function aplicarFiltrosFinanceiro() {{
  const cat = (document.getElementById('fFinCat')?.value  || '').trim();
  const cc  = (document.getElementById('fFinCC')?.value   || '').trim();
  pagarFiltFin = pagarFilt.filter(r => {{
    if (cat && (r.cat || '') !== cat) return false;
    if (cc  && (r.cc  || '') !== cc)  return false;
    return true;
  }});
  mkFinanceiro();
}}

function limparFiltrosFinanceiro() {{
  const elCat = document.getElementById('fFinCat');
  const elCC  = document.getElementById('fFinCC');
  if (elCat) elCat.value = '';
  if (elCC)  elCC.value  = '';
  pagarFiltFin = pagarFilt;
  mkFinanceiro();
}}

function filtrar() {{
  const ini  = document.getElementById('fDateIni').value;
  const fim  = document.getElementById('fDateFim').value;
  dadosFilt = VENDAS.filter(r => {{
    if (ini && r.data < ini) return false;
    if (fim && r.data > fim) return false;
    if (lojaFiltAtivo !== 'ambas' && r.loja && r.loja !== lojaFiltAtivo) return false;
    return true;
  }});
  pagarFilt = PAGAR.filter(r => {{
    if (!r.venc) return true;
    if (ini && r.venc < ini) return false;
    if (fim && r.venc > fim) return false;
    if (lojaFiltAtivo !== 'ambas' && r.loja && r.loja !== lojaFiltAtivo) return false;
    return true;
  }});
  pagarFiltFin = pagarFilt;  // reseta filtro interno ao alterar per\u00EDodo global
  pontoMarcFilt = PONTO_MARC.filter(r => {{
    if (ini && r.data < ini) return false;
    if (fim && r.data > fim) return false;
    return true;
  }});
  const fmtD = s => s ? s.split('-').reverse().join('/') : '?';
  const tP = document.getElementById('tituloPonto');
  if (tP) tP.textContent = '\uD83D\uDD50 Ponto Colaborador \u2014 ' + fmtD(ini) + ' a ' + fmtD(fim);
  const info = document.getElementById('filtroInfo');
  info.textContent = dadosFilt.length === VENDAS.length ? '' :
    `\u2714 ${{NUM(dadosFilt.length)}} de ${{NUM(VENDAS.length)}} itens filtrados`;
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  CHARTS
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
const charts = {{}};
function destroyChart(id) {{
  // 1) destroy via our registry
  if (charts[id]) {{ try {{ charts[id].destroy(); }} catch(e){{}} delete charts[id]; }}
  // 2) destroy via Chart.js native registry (catches orphaned charts after iframe re-render)
  const el = document.getElementById(id);
  if (el) {{ try {{ const c = Chart.getChart(el); if (c) c.destroy(); }} catch(e){{}} }}
}}

function mkMensal(rows) {{
  const m = {{}};
  rows.forEach(r => {{
    const ym = (r.data||'').substring(0,7);
    if (!ym) return;
    m[ym] = (m[ym]||0) + r.liq;
  }});
  const entries = Object.entries(m).sort((a,b)=>a[0]>b[0]?1:-1);
  const meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
  const labels = entries.map(([ym]) => {{
    const [y,mo] = ym.split('-');
    return meses[parseInt(mo)-1]+'/'+y.substring(2);
  }});
  destroyChart('chartMensal');
  charts['chartMensal'] = new Chart(document.getElementById('chartMensal'), {{
    type: 'bar',
    data: {{
      labels,
      datasets: [{{
        label: 'Fat. L\u00EDquido',
        data: entries.map(e=>Math.round(e[1])),
        backgroundColor: '#0891b2', borderRadius: 5, borderSkipped: false,
      }}]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{
        legend:{{display:false}},
        subtitle:{{
          display:true,
          text:'Total: '+BRL(entries.reduce((s,e)=>s+e[1],0)),
          color:'#38bdf8',
          font:{{size:13,weight:'bold'}},
          padding:{{bottom:8}}
        }},
        tooltip:{{callbacks:{{label:c=>BRL(c.raw)}}}}}},
      scales: {{
        y: {{ticks:{{callback:v=>'R$'+(v>=1000?(v/1000).toFixed(0)+'k':v),color:'#cbd5e1'}},grid:{{color:'#334155'}}}},
        x: {{ticks:{{color:'#cbd5e1'}},grid:{{display:false}}}}
      }}
    }}
  }});
}}

function mkCategoria(rows) {{
  const m = {{}};
  rows.forEach(r => {{ m[r.categoria||'SEM CATEGORIA'] = (m[r.categoria||'SEM CATEGORIA']||0)+r.liq; }});
  const entries = Object.entries(m).sort((a,b)=>b[1]-a[1]).slice(0,10);
  const total = entries.reduce((s,e)=>s+e[1],0);
  destroyChart('chartCategoria');
  const canvasCat = document.getElementById('chartCategoria');
  const isDarkCat = document.body.getAttribute('data-theme') !== 'light';
  const centerPluginCat = [{{
    id: 'centerTextCat',
    beforeDraw(chart) {{
      const {{ctx}} = chart;
      const cx = chart.chartArea ? (chart.chartArea.left+chart.chartArea.right)/2 : chart.width/2;
      const cy = chart.chartArea ? (chart.chartArea.top+chart.chartArea.bottom)/2 : chart.height/2;
      const dark = document.body.getAttribute('data-theme') !== 'light';
      ctx.save();
      ctx.textAlign='center'; ctx.textBaseline='middle';
      ctx.fillStyle = dark ? '#cbd5e1' : '#1e293b';
      ctx.font='bold 11px Inter,sans-serif';
      ctx.fillText('TOTAL', cx, cy-10);
      ctx.font='bold 13px Inter,sans-serif';
      ctx.fillStyle = dark ? '#38bdf8' : '#0284c7';
      ctx.fillText(BRL(total), cx, cy+8);
      ctx.restore();
    }}
  }}];
  charts['chartCategoria'] = new Chart(canvasCat, {{
    type: 'doughnut',
    data: {{
      labels: entries.map(e=>e[0]),
      datasets: [{{ data: entries.map(e=>Math.round(e[1])),
        backgroundColor: CORES, borderWidth: 2, borderColor: '#fff', hoverOffset: 10 }}]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      cutout: '55%',
      plugins: {{
        legend: {{position:'right',labels:{{font:{{size:11}},boxWidth:14,color:'#cbd5e1'}}}},
        tooltip: {{callbacks:{{label:c=>c.label+': '+BRL(c.raw)}}}}
      }}
    }},
    plugins: centerPluginCat
  }});
}}

function mkVendasCC(rows) {{
  destroyChart('chartVendasCC');
  const canvas = document.getElementById('chartVendasCC');
  if (!canvas) return;
  const mCC = {{}};
  rows.forEach(r => {{
    const cc = (r.cc || '').trim() || 'SEM CENTRO DE CUSTO';
    mCC[cc] = (mCC[cc] || 0) + r.liq;
  }});
  const entries = Object.entries(mCC).sort((a, b) => b[1] - a[1]);
  const wrap = document.getElementById('wrapVendasCC');
  if (!entries.length) {{
    if (wrap) wrap.innerHTML = '<div style="display:flex;align-items:center;justify-content:center;height:100%;color:#94a3b8;font-size:14px;">Nenhuma venda no período</div>';
    return;
  }}
  const total = entries.reduce((s, e) => s + e[1], 0);
  const altura = Math.max(320, entries.length * 44);
  if (wrap) wrap.style.height = altura + 'px';
  const isDark = document.body.getAttribute('data-theme') !== 'light';
  const gridClr = isDark ? '#334155' : '#e2e8f0';
  const txtClr  = isDark ? '#cbd5e1' : '#1e293b';
  const CORES_V = ['#0891b2','#0ea5e9','#38bdf8','#7dd3fc','#22d3ee','#06b6d4','#0284c7','#0369a1','#14b8a6','#2dd4bf','#34d399','#4ade80'];
  charts['chartVendasCC'] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels: entries.map(e => e[0]),
      datasets: [{{
        label: 'Faturamento Líquido',
        data: entries.map(e => Math.round(e[1] * 100) / 100),
        backgroundColor: entries.map((_, i) => CORES_V[i % CORES_V.length]),
        borderRadius: 4, borderSkipped: false
      }}]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false, indexAxis: 'y',
      plugins: {{
        legend: {{ display: false }},
        subtitle: {{ display: true, text: 'Total: ' + BRL(total),
          color: '#38bdf8', font: {{ size: 12, weight: 'bold' }}, padding: {{ bottom: 6 }} }},
        tooltip: {{ callbacks: {{ label: c => BRL(c.raw) + ' (' + ((c.raw / total) * 100).toFixed(1) + '%)' }} }}
      }},
      scales: {{
        x: {{ grid: {{ color: gridClr }}, ticks: {{ color: txtClr, callback: v => v >= 1000 ? 'R$' + (v/1000).toFixed(0) + 'k' : 'R$' + v }} }},
        y: {{ grid: {{ display: false }}, ticks: {{ color: txtClr, font: {{ size: 10 }} }} }}
      }}
    }}
  }});
}}

function mkHorizBar(id, entries) {{
  destroyChart(id);
  charts[id] = new Chart(document.getElementById(id), {{
    type: 'bar',
    data: {{
      labels: entries.map(e=>e[0].length>28?e[0].substring(0,25)+'...':e[0]),
      datasets: [{{ data: entries.map(e=>Math.round(e[1])),
        backgroundColor: CORES, borderRadius: 4, borderSkipped: false }}]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      indexAxis: 'y',
      plugins: {{legend:{{display:false}},
        tooltip:{{callbacks:{{label:c=>BRL(c.raw)}}}}}},
      scales: {{
        x: {{ticks:{{callback:v=>'R$'+(v>=1000?(v/1000).toFixed(0)+'k':v),color:'#cbd5e1'}},grid:{{color:'#334155'}}}},
        y: {{ticks:{{color:'#cbd5e1'}},grid:{{display:false}}}}
      }}
    }}
  }});
}}

function mkDonut(id, entries, centerLabel) {{
  destroyChart(id);
  const canvas = document.getElementById(id);
  if (!canvas) return;
  // Remove mensagem de "sem dados" anterior, se houver
  const prev = document.getElementById(id + '-nodata');
  if (prev) prev.remove();
  if (!entries || entries.length === 0) {{
    canvas.style.display = 'none';
    const msg = document.createElement('div');
    msg.id = id + '-nodata';
    msg.style.cssText = 'display:flex;align-items:center;justify-content:center;height:120px;color:#94a3b8;font-size:13px;text-align:center;';
    msg.innerHTML = '\u26A0\uFE0F M\u00F3dulo financeiro n\u00E3o dispon\u00EDvel<br/>no plano atual da API';
    canvas.parentNode.insertBefore(msg, canvas.nextSibling);
    return;
  }}
  canvas.style.display = '';
  const centerPlugin = centerLabel ? [{{
    id: 'centerText_' + id,
    beforeDraw(chart) {{
      const {{width, height, ctx}} = chart;
      const cx = chart.chartArea ? (chart.chartArea.left + chart.chartArea.right) / 2 : width / 2;
      const cy = chart.chartArea ? (chart.chartArea.top + chart.chartArea.bottom) / 2 : height / 2;
      ctx.save();
      const isDark = document.body.getAttribute('data-theme') !== 'light';
      ctx.fillStyle = isDark ? '#cbd5e1' : '#1e293b';
      ctx.font = 'bold 11px Inter,sans-serif';
      ctx.textAlign = 'center';
      ctx.textBaseline = 'middle';
      ctx.fillText('TOTAL', cx, cy - 10);
      ctx.font = 'bold 13px Inter,sans-serif';
      ctx.fillStyle = isDark ? '#38bdf8' : '#0284c7';
      ctx.fillText(centerLabel, cx, cy + 8);
      ctx.restore();
    }}
  }}] : [];
  charts[id] = new Chart(canvas, {{
    type: 'doughnut',
    data: {{
      labels: entries.map(e=>e[0]),
      datasets: [{{ data: entries.map(e=>Math.round(e[1])),
        backgroundColor: CORES, borderWidth: 2, borderColor: '#fff' }}]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      cutout: '55%',
      plugins: {{
        legend: {{position:'right',labels:{{font:{{size:11}},boxWidth:14,color:'#cbd5e1'}}}},
        tooltip: {{callbacks:{{label:c=>c.label+': '+BRL(c.raw)}}}}
      }}
    }},
    plugins: centerPlugin
  }});
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  ATUALIZA\u00C7\u00C3O DOS KPIs
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
function atualizarKPIVendas(rows) {{
  const liq   = rows.reduce((s,r)=>s+r.liq,0);
  const bruto = rows.reduce((s,r)=>s+r.bruto,0);
  const desc  = rows.reduce((s,r)=>s+r.desc,0);
  const qtd   = rows.reduce((s,r)=>s+r.qtd,0);
  const clis  = new Set(rows.map(r=>r.cliente)).size;
  const nfs   = new Set(rows.map(r=>r.id)).size;
  document.getElementById('kLiq').textContent   = BRL(liq);
  document.getElementById('kBruto').textContent = BRL(bruto);
  document.getElementById('kDesc').textContent  = BRL(desc);
  document.getElementById('kQtd').textContent   = NUM(Math.round(qtd));
  document.getElementById('kCli').textContent   = NUM(clis);
  document.getElementById('kNFs').textContent   = NUM(nfs);
}}

function atualizarKPIFinanceiro() {{
  const pagTot  = pagarFilt.reduce((s,r)=>s+r.valor,0);
  const pagPago = pagarFilt.reduce((s,r)=>s+r.pago,0);
  document.getElementById('kPagTotal').textContent = BRL(pagTot);
  document.getElementById('kPagPago').textContent  = BRL(pagPago);
}}

function atualizarKPIOS() {{
  const total   = OS_LIST.length;
  const abertas = OS_LIST.filter(o=>!['CONCLUIDO','FECHADA','CANCELADA'].includes(o.st)).length;
  const conc    = OS_LIST.filter(o=>['CONCLUIDO','CONCLU\u00CDDA','FECHADA'].includes(o.st)).length;
  const contativos = CONTRATOS.filter(c=>c.st && !['CANCELADO','INATIVO'].includes(c.st.toUpperCase())).length;
  const mrr = CONTRATOS
    .filter(c=>c.st && !['CANCELADO','INATIVO'].includes(c.st.toUpperCase()))
    .filter(c=>c.period && (c.period.includes('MENSAL') || c.period.includes('M\u00CAS')))
    .reduce((s,c)=>s+c.val,0);
  document.getElementById('kOsAbert').textContent  = NUM(abertas);
  document.getElementById('kOsConc').textContent   = NUM(conc);
  document.getElementById('kContrAt').textContent  = NUM(contativos);
  document.getElementById('kMRR').textContent      = BRL(mrr);
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  TABELAS
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
function renderTblProdutos(rows) {{
  const m = {{}};
  rows.forEach(r => {{
    const k = r.cod || r.produto;
    if (!m[k]) m[k] = {{cod:r.cod,prod:r.produto,cat:r.categoria,qtd:0,bruto:0,desc:0,liq:0}};
    m[k].qtd+=r.qtd; m[k].bruto+=r.bruto; m[k].desc+=r.desc; m[k].liq+=r.liq;
  }});
  const arr = Object.values(m).sort((a,b)=>b.liq-a.liq).slice(0,20);
  const totLiq = arr.reduce((s,r)=>s+r.liq,0);
  let html = '';
  arr.forEach((r,i) => {{
    html += `<tr>
      <td>${{i+1}}</td><td><code style="font-size:11px;color:#64748b">${{r.cod||'\u2014'}}</code></td>
      <td>${{r.prod||'\u2014'}}</td><td>${{r.cat||'\u2014'}}</td>
      <td class="num">${{r.qtd.toLocaleString('pt-BR',{{maximumFractionDigits:0}})}}</td>
      <td class="num">${{BRL(r.bruto)}}</td><td class="num">${{BRL(r.desc)}}</td>
      <td class="num">${{BRL(r.liq)}}</td>
      <td class="num">${{PCT(r.liq,totLiq)}}</td>
    </tr>`;
  }});
  document.getElementById('tblProdCorpo').innerHTML = html;
  const tTot = arr.reduce((s,r)=>s+r.bruto,0);
  const tDesc = arr.reduce((s,r)=>s+r.desc,0);
  const tQtd = arr.reduce((s,r)=>s+r.qtd,0);
  document.getElementById('tblProdTotal').innerHTML =
    `<td colspan="4">TOTAL (TOP 20)</td>
     <td class="num">${{tQtd.toLocaleString('pt-BR',{{maximumFractionDigits:0}})}}</td>
     <td class="num">${{BRL(tTot)}}</td><td class="num">${{BRL(tDesc)}}</td>
     <td class="num">${{BRL(totLiq)}}</td><td class="num">100%</td>`;
}}

function renderTblClientes(rows) {{
  const m = {{}};
  rows.forEach(r => {{
    const k = r.cliente;
    if (!m[k]) m[k] = {{cli:r.cliente,vend:r.vendedor,ids:new Set(),bruto:0,desc:0,liq:0}};
    m[k].ids.add(r.id); m[k].bruto+=r.bruto; m[k].desc+=r.desc; m[k].liq+=r.liq;
  }});
  const arr = Object.values(m).sort((a,b)=>b.liq-a.liq).slice(0,20);
  const totLiq = arr.reduce((s,r)=>s+r.liq,0);
  let html = '';
  arr.forEach((r,i) => {{
    html += `<tr>
      <td>${{i+1}}</td><td>${{r.cli||'\u2014'}}</td><td>${{r.vend||'\u2014'}}</td>
      <td class="num">${{NUM(r.ids.size)}}</td>
      <td class="num">${{BRL(r.bruto)}}</td><td class="num">${{BRL(r.desc)}}</td>
      <td class="num">${{BRL(r.liq)}}</td>
      <td class="num">${{PCT(r.liq,totLiq)}}</td>
    </tr>`;
  }});
  document.getElementById('tblCliCorpo').innerHTML = html;
  const tTot = arr.reduce((s,r)=>s+r.bruto,0);
  const tDesc = arr.reduce((s,r)=>s+r.desc,0);
  document.getElementById('tblCliTotal').innerHTML =
    `<td colspan="3">TOTAL (TOP 20)</td><td class="num">\u2014</td>
     <td class="num">${{BRL(tTot)}}</td><td class="num">${{BRL(tDesc)}}</td>
     <td class="num">${{BRL(totLiq)}}</td><td class="num">100%</td>`;
}}

function statusBadge(st) {{
  const s = (st||'').toUpperCase();
  if (['PAGO','RECEBIDO','QUITADO','CONCLUIDO','CONCLU\u00CDDA','ATIVO'].some(x=>s.includes(x)))
    return `<span class="badge verde">${{st}}</span>`;
  if (['VENCIDO','INADIMPLENTE','ATRASADO','CANCELADO'].some(x=>s.includes(x)))
    return `<span class="badge vermelho">${{st}}</span>`;
  if (['ABERTO','PENDENTE','AGUARDANDO'].some(x=>s.includes(x)))
    return `<span class="badge amarelo">${{st}}</span>`;
  return `<span class="badge cinza">${{st||'\u2014'}}</span>`;
}}

function renderTblReceber() {{
  const sorted = [...RECEBER].sort((a,b)=>b.valor-a.valor).slice(0,30);
  let html = '';
  sorted.forEach(r => {{
    html += `<tr>
      <td>${{r.desc||'\u2014'}}</td><td>${{r.pessoa||'\u2014'}}</td>
      <td class="num">${{BRL(r.valor)}}</td>
      <td class="num">${{BRL(r.pago)}}</td>
      <td class="num" style="color:${{r.saldo>0?'#dc2626':'#059669'}}">${{BRL(r.saldo)}}</td>
      <td>${{r.venc?r.venc.split('-').reverse().join('/'):'\u2014'}}</td>
      <td>${{statusBadge(r.status)}}</td>
    </tr>`;
  }});
  document.getElementById('tblRecCorpo').innerHTML = html;
}}

function renderTblOS() {{
  const sorted = [...OS_LIST].sort((a,b)=>(b.data||'')>(a.data||'')?1:-1).slice(0,30);
  let html = '';
  sorted.forEach(r => {{
    html += `<tr>
      <td>${{r.id||'\u2014'}}</td>
      <td>${{r.data?r.data.split('-').reverse().join('/'):'\u2014'}}</td>
      <td>${{r.cli||'\u2014'}}</td><td>${{r.tec||'\u2014'}}</td>
      <td title="${{r.desc||''}}">${{(r.desc||'').substring(0,45)}}...</td>
      <td>${{statusBadge(r.st)}}</td>
      <td class="num">${{BRL(r.val)}}</td>
    </tr>`;
  }});
  document.getElementById('tblOsCorpo').innerHTML = html;
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  GR\u00C1FICOS FINANCEIROS
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
function mkFinanceiro() {{
  // Contas a pagar por categoria
  const mPag = {{}};
  pagarFiltFin.forEach(r => {{ mPag[r.cat||'OUTROS'] = (mPag[r.cat||'OUTROS']||0)+r.valor; }});
  const pagEntries = Object.entries(mPag).sort((a,b)=>b[1]-a[1]).slice(0,8);
  const totalPagar = pagarFiltFin.reduce((s,r)=>s+r.valor, 0);
  mkDonut('chartPagar', pagEntries, BRL(totalPagar));
  mkPagarMensal();
  mkPagarCentroCusto();
  mkResultadoMensal();
}}

function mkResultadoMensal() {{
  destroyChart('chartResultado');
  const canvas = document.getElementById('chartResultado');
  if (!canvas) return;

  // Coleta meses presentes nos dados FILTRADOS pelo per\u00EDodo
  const mesesSet = new Set();
  dadosFilt.forEach(r  => {{ if (r.data) mesesSet.add(r.data.slice(0,7)); }});
  pagarFiltFin.forEach(r => {{ if (r.venc) mesesSet.add(r.venc.slice(0,7)); }});
  if (!mesesSet.size) return;
  // Filtra apenas meses dentro do per\u00EDodo selecionado pelo usu\u00E1rio
  const _mkIni = PERIODO_INI.slice(0,7);
  const _mkFim = PERIODO_FIM.slice(0,7);
  const meses = [...mesesSet].sort().filter(m => m >= _mkIni && m <= _mkFim);
  if (!meses.length) return;

  // Agrupa receitas (vendas liq filtradas) por m\u00EAs
  const recMes = {{}};
  meses.forEach(m => recMes[m] = 0);
  dadosFilt.forEach(r => {{
    const mk = (r.data || '').slice(0,7);
    if (recMes[mk] !== undefined) recMes[mk] += (r.liq || 0);
  }});

  // Agrupa despesas (pagar do per\u00EDodo) por m\u00EAs de vencimento
  const despMes = {{}};
  meses.forEach(m => despMes[m] = 0);
  pagarFiltFin.forEach(r => {{
    const mk = (r.venc || '').slice(0,7);
    if (despMes[mk] !== undefined) despMes[mk] += (r.valor || 0);
  }});

  const labels   = meses.map(m => {{
    const [y,mo] = m.split('-');
    return new Date(+y, +mo-1, 1).toLocaleDateString('pt-BR',{{month:'short',year:'2-digit'}});
  }});
  const recData  = meses.map(m => Math.round(recMes[m]  * 100) / 100);
  const despData = meses.map(m => Math.round(despMes[m] * 100) / 100);
  const resData  = meses.map(m => Math.round((recMes[m] - despMes[m]) * 100) / 100);

  const totalRec  = recData.reduce((s,v)=>s+v,0);
  const totalDesp = despData.reduce((s,v)=>s+v,0);
  const totalRes  = totalRec - totalDesp;
  const corRes    = totalRes >= 0 ? '#22c55e' : '#ef4444';

  const isDark  = document.body.getAttribute('data-theme') !== 'light';
  const gridClr = isDark ? '#334155' : '#e2e8f0';
  const txtClr  = isDark ? '#cbd5e1' : '#1e293b';

  charts['chartResultado'] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels,
      datasets: [
        {{ type: 'bar',  label: 'Receitas',  data: recData,  backgroundColor: '#059669', stack: 's', borderRadius: 0, order: 2 }},
        {{ type: 'bar',  label: 'Despesas',  data: despData, backgroundColor: '#ef4444', stack: 's', borderRadius: 4, order: 2 }},
        {{ type: 'line', label: 'Resultado', data: resData,
           borderColor: '#f59e0b', backgroundColor: 'transparent',
           pointBackgroundColor: resData.map(v => v >= 0 ? '#22c55e' : '#ef4444'),
           pointBorderColor: '#fff', pointBorderWidth: 1,
           pointRadius: 7, pointHoverRadius: 10,
           borderWidth: 2, tension: 0.3, yAxisID: 'yRes', order: 0 }}
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ labels: {{ color: txtClr, font: {{size:11}}, boxWidth:12 }} }},
        subtitle: {{
          display: true,
          text: 'Receitas: ' + BRL(totalRec) + '   |   Despesas: ' + BRL(totalDesp) + '   |   Resultado: ' + BRL(totalRes),
          color: corRes, font: {{ size: 12, weight: 'bold' }}, padding: {{ bottom: 6 }}
        }},
        tooltip: {{ callbacks: {{ label: c => c.dataset.label + ': ' + BRL(c.raw) }} }}
      }},
      scales: {{
        x: {{ grid: {{ color: gridClr }}, ticks: {{ color: txtClr, maxRotation: 45, font: {{size:9}} }} }},
        y: {{ stacked: true, grid: {{ color: gridClr }}, ticks: {{ color: txtClr, callback: v => v>=1000?'R$'+(v/1000).toFixed(0)+'k':'R$'+v }} }},
        yRes: {{
          position: 'right', grid: {{ display: false }},
          ticks: {{ color: '#f59e0b', callback: v => v>=1000?'R$'+(v/1000).toFixed(0)+'k':'R$'+v }}
        }}
      }}
    }}
  }});
}}

function abrirCCModal(ccNome) {{
  const EXCLUIR = ['ADM/FINANCEIRO','SUBLOCAÇÕES - TERCEIROS','SEM CENTRO DE CUSTO','MANUTENÇÃO'];
  const rows = pagarFiltFin.filter(r => {{
    const cc = (r.cc || '').trim() || 'SEM CENTRO DE CUSTO';
    if (EXCLUIR.includes(cc.toUpperCase())) return false;
    return cc === ccNome;
  }}).sort((a, b) => {{
    const d = (a.desc || '').localeCompare(b.desc || '', 'pt-BR');
    if (d !== 0) return d;
    return (a.venc || '').localeCompare(b.venc || '');
  }});
  const total = rows.reduce((s, r) => s + r.valor, 0);
  document.getElementById('ccModalTitle').textContent = '\uD83C\uDFF7\uFE0F ' + ccNome + ' \u2014 Discrimina\u00E7\u00E3o dos Lan\u00E7amentos';
  let html = '<table><thead><tr><th>#</th><th>Descri\u00E7\u00E3o</th><th>Fornecedor</th><th>Vencimento</th><th class="num">Valor</th></tr></thead><tbody>';
  rows.forEach((r, i) => {{
    const venc = r.venc ? r.venc.split('-').reverse().join('/') : '\u2014';
    const desc = r.desc || '\u2014';
    let descHtml = desc;
    let extraClass = '';
    let dataTip = '';
    if (r.parc_n > 0) {{
      descHtml += ' <span style="background:#f59e0b;color:#1e293b;font-size:10px;font-weight:700;padding:1px 8px;border-radius:10px;white-space:nowrap;vertical-align:middle">Parcela ' + r.parc_n + ' de ' + r.parc_tot + '</span>';
      extraClass = ' class="tip-cell"';
      dataTip = ' data-tip="Parcela ' + r.parc_n + ' de ' + r.parc_tot + ' \u2014 mesma descri\u00E7\u00E3o/fornecedor"';
    }}
    html += '<tr><td style="color:#64748b;font-size:11px">' + (i+1) + '</td><td' + extraClass + dataTip + '>' + descHtml + '</td><td style="color:#94a3b8">' + (r.pessoa||'\u2014') + '</td><td style="color:#94a3b8">' + venc + '</td><td class="num">' + BRL(r.valor) + '</td></tr>';
  }});
  html += `</tbody><tfoot><tr><td colspan="4">TOTAL (${{rows.length}} lançamentos)</td><td class="num">${{BRL(total)}}</td></tr></tfoot></table>`;
  document.getElementById('ccModalBody').innerHTML = html;
  document.getElementById('ccOverlay').classList.add('open');
}}
function fecharCCModal() {{ document.getElementById('ccOverlay').classList.remove('open'); }}

function mkManutencaoCC() {{
  destroyChart('chartManutCC');
  const canvas = document.getElementById('chartManutCC');
  if (!canvas) return;
  const EXCLUIR = ['ADM/FINANCEIRO','SUBLOCAÇÕES - TERCEIROS','SEM CENTRO DE CUSTO','MANUTENÇÃO'];
  const mCC = {{}};
  pagarFiltFin.forEach(r => {{
    const cc = (r.cc || '').trim() || 'SEM CENTRO DE CUSTO';
    if (EXCLUIR.includes(cc.toUpperCase())) return;
    mCC[cc] = (mCC[cc] || 0) + r.valor;
  }});
  const entries = Object.entries(mCC).sort((a, b) => b[1] - a[1]);
  const wrap = document.getElementById('wrapManutCC');
  if (!entries.length) {{
    if (wrap) wrap.innerHTML = '<div style="display:flex;align-items:center;justify-content:center;height:100%;color:#94a3b8;font-size:14px;">Nenhuma despesa de equipamento no período</div>';
    return;
  }}
  const total = entries.reduce((s, e) => s + e[1], 0);
  const altura = Math.max(320, entries.length * 44);
  if (wrap) wrap.style.height = altura + 'px';
  const isDark = document.body.getAttribute('data-theme') !== 'light';
  const gridClr = isDark ? '#334155' : '#e2e8f0';
  const txtClr  = isDark ? '#cbd5e1' : '#1e293b';
  const CORES = ['#f59e0b','#f97316','#ef4444','#eab308','#84cc16','#22c55e','#14b8a6','#06b6d4','#3b82f6','#8b5cf6','#ec4899','#d946ef'];
  charts['chartManutCC'] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels: entries.map(e => e[0]),
      datasets: [{{
        label: 'Despesas',
        data: entries.map(e => Math.round(e[1] * 100) / 100),
        backgroundColor: entries.map((_, i) => CORES[i % CORES.length]),
        borderRadius: 4, borderSkipped: false,
        cursor: 'pointer'
      }}]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false, indexAxis: 'y',
      onClick: (evt, els) => {{
        if (els.length) abrirCCModal(entries[els[0].index][0]);
      }},
      plugins: {{
        legend: {{ display: false }},
        subtitle: {{ display: true, text: 'Total: ' + BRL(total) + '  —  clique numa barra para ver os lançamentos',
          color: '#f59e0b', font: {{ size: 11, weight: 'bold' }}, padding: {{ bottom: 6 }} }},
        tooltip: {{ callbacks: {{ label: c => BRL(c.raw) + ' (' + ((c.raw / total) * 100).toFixed(1) + '%)' }} }}
      }},
      scales: {{
        x: {{ grid: {{ color: gridClr }}, ticks: {{ color: txtClr, callback: v => v >= 1000 ? 'R$' + (v/1000).toFixed(0) + 'k' : 'R$' + v }} }},
        y: {{ grid: {{ display: false }}, ticks: {{ color: txtClr, font: {{ size: 10 }} }} }}
      }}
    }}
  }});
  canvas.style.cursor = 'pointer';
}}

// ── Tabela de Status de Manutenção Preventiva ──────────────────────
function mkManutencao() {{
  const tbody = document.getElementById('tblManutencaoBdy');
  const vazioMsg = document.getElementById('manutVazioMsg');
  if (!tbody) return;
  tbody.innerHTML = '';
  const data = MANUTENCAO;
  if (!data || !data.length) {{
    if (vazioMsg) vazioMsg.style.display = '';
    document.getElementById('kManutVencidas').textContent = '0';
    document.getElementById('kManutProximas').textContent = '0';
    document.getElementById('kManutOk').textContent = '0';
    return;
  }}
  if (vazioMsg) vazioMsg.style.display = 'none';
  let vencidas = 0, proximas = 0, ok = 0;
  data.forEach(r => {{
    const st = r.status || 'vencida';
    if (st === 'vencida') vencidas++;
    else if (st === 'proxima') proximas++;
    else ok++;
    let badge, rowBg;
    if (st === 'vencida') {{
      badge  = '<span style="background:#dc2626;color:#fff;padding:2px 10px;border-radius:12px;font-size:11px;font-weight:700">&#128308; VENCIDA</span>';
      rowBg  = 'background:rgba(220,38,38,.07)';
    }} else if (st === 'proxima') {{
      badge  = '<span style="background:#d97706;color:#fff;padding:2px 10px;border-radius:12px;font-size:11px;font-weight:700">&#9888; PR&#211;XIMA</span>';
      rowBg  = 'background:rgba(217,119,6,.07)';
    }} else {{
      badge  = '<span style="background:#059669;color:#fff;padding:2px 10px;border-radius:12px;font-size:11px;font-weight:700">&#9989; EM DIA</span>';
      rowBg  = '';
    }}
    const fmtDt = s => s ? s.split('-').reverse().join('/') : '<em style="color:#94a3b8">Não registrada</em>';
    let diasTxt = '—', diasStyle = '';
    if (r.dias !== undefined && r.ultima) {{
      if (r.dias < 0) {{
        diasTxt  = Math.abs(r.dias) + ' dias atrás';
        diasStyle = 'color:#dc2626;font-weight:600';
      }} else if (r.dias <= 5) {{
        diasTxt  = r.dias + ' dias';
        diasStyle = 'color:#d97706;font-weight:600';
      }} else {{
        diasTxt  = r.dias + ' dias';
        diasStyle = 'color:#059669';
      }}
    }} else if (!r.ultima) {{
      diasTxt  = 'Nunca realizada';
      diasStyle = 'color:#dc2626;font-weight:600';
    }}
    const tr = document.createElement('tr');
    if (rowBg) tr.setAttribute('style', rowBg);
    tr.innerHTML =
      '<td><strong>' + r.cc + '</strong></td>' +
      '<td>' + badge + '</td>' +
      '<td>' + fmtDt(r.ultima) + '</td>' +
      '<td>' + fmtDt(r.proxima) + '</td>' +
      '<td class="num" style="' + diasStyle + '">' + diasTxt + '</td>';
    tbody.appendChild(tr);
  }});
  document.getElementById('kManutVencidas').textContent = vencidas;
  document.getElementById('kManutProximas').textContent = proximas;
  document.getElementById('kManutOk').textContent = ok;
}}

// ── Salvar manutenção via Supabase REST (fetch direto do browser) ──
async function salvarManutencao() {{
  const equip = (document.getElementById('mFormEquip').value || '').trim();
  const data  = (document.getElementById('mFormData').value  || '').trim();
  const serv  = (document.getElementById('mFormServico').value || '').trim();
  const msg   = document.getElementById('mFormMsg');

  if (!equip) {{ _mMsg(msg, '\u274c Selecione o equipamento.', '#dc2626'); return; }}
  if (!data)  {{ _mMsg(msg, '\u274c Selecione a data da \u00faltima manuten\u00e7\u00e3o.', '#dc2626'); return; }}

  const sbUrl  = _SB_URL;
  const sbAnon = _SB_ANON;
  if (!sbUrl || sbUrl === '') {{
    _mMsg(msg, '\u26a0\ufe0f Supabase n\u00e3o configurado.', '#d97706');
    return;
  }}

  const hdrs = {{
    'apikey': sbAnon, 'Authorization': 'Bearer ' + sbAnon,
    'Content-Type': 'application/json', 'Prefer': 'resolution=merge-duplicates'
  }};
  const payload = JSON.stringify({{
    equipamento: equip, ultima_manutencao: data,
    tipo_servico: serv || null,
    intervalo_meses: 2, updated_at: new Date().toISOString()
  }});
  _mMsg(msg, '\u23f3 Salvando...', '#0891b2');
  try {{
    const r = await fetch(sbUrl + '/rest/v1/manutencoes_equipamentos', {{
      method: 'POST', headers: hdrs, body: payload
    }});
    if (!r.ok) throw new Error('HTTP ' + r.status);
    _mMsg(msg, '\u2705 ' + equip + ' \u2014 manuten\u00e7\u00e3o registrada em ' + data.split('-').reverse().join('/'), '#059669');
    document.getElementById('mFormEquip').value = '';
    document.getElementById('mFormServico').value = '';
    // Atualiza tabela de status localmente
    const idx = MANUTENCAO.findIndex(r => r.cc === equip);
    const hoje = new Date(); hoje.setHours(0,0,0,0);
    const dt = new Date(data + 'T00:00:00');
    const prox = new Date(dt); prox.setDate(prox.getDate() + 60);
    const dias = Math.round((prox - hoje) / 86400000);
    const st = dias < 0 ? 'vencida' : (dias <= 5 ? 'proxima' : 'ok');
    const rec = {{ cc: equip, ultima: data, proxima: prox.toISOString().slice(0,10), status: st, dias: dias, tipo_servico: serv }};
    if (idx >= 0) MANUTENCAO[idx] = rec; else MANUTENCAO.push(rec);
    mkManutencao();
  }} catch(e) {{
    _mMsg(msg, '\u274c Erro ao salvar: ' + e.message, '#dc2626');
  }}
}}
function _mMsg(el, txt, cor) {{
  el.textContent = txt; el.style.color = cor; el.style.display = '';
}}

// ── Excluir manutenção via Supabase REST ───────────────────────────
async function deletarManutencao() {{
  const sel   = document.getElementById('mFormEquip');
  const equip = (sel ? sel.value : '').trim();
  const msg   = document.getElementById('mFormMsg');
  if (!equip) {{ _mMsg(msg, '\u274c Selecione o equipamento a excluir.', '#dc2626'); return; }}
  if (!confirm('Tem certeza que deseja excluir o registro de manuten\u00e7\u00e3o de "' + equip + '"?')) return;
  const sbUrl  = _SB_URL;
  const sbAnon = _SB_ANON;
  if (!sbUrl) {{ _mMsg(msg, '\u26a0\ufe0f Supabase n\u00e3o configurado.', '#d97706'); return; }}
  _mMsg(msg, '\u23f3 Excluindo...', '#0891b2');
  try {{
    const r = await fetch(
      sbUrl + '/rest/v1/manutencoes_equipamentos?equipamento=eq.' + encodeURIComponent(equip),
      {{ method: 'DELETE', headers: {{ 'apikey': sbAnon, 'Authorization': 'Bearer ' + sbAnon }} }}
    );
    if (!r.ok) throw new Error('HTTP ' + r.status);
    _mMsg(msg, '\u2705 Registro de ' + equip + ' exclu\u00eddo.', '#059669');
    // Reseta localmente
    const idx = MANUTENCAO.findIndex(x => x.cc === equip);
    if (idx >= 0) {{
      MANUTENCAO[idx] = {{ cc: equip, ultima: null, proxima: null, status: 'vencida', dias: -9999, email: null }};
    }}
    mkManutencao();
    sel.value = '';
    document.getElementById('mFormData').value = new Date().toISOString().slice(0,10);
    document.getElementById('mFormServico').value = '';
  }} catch(e) {{
    _mMsg(msg, '\u274c Erro ao excluir: ' + e.message, '#dc2626');
  }}
}}

function mkPagarCentroCusto() {{
  destroyChart('chartPagarCC');
  const canvas = document.getElementById('chartPagarCC');
  if (!canvas) return;
  const mCC = {{}};
  pagarFiltFin.forEach(r => {{
    const cc = (r.cc || '').trim() || 'SEM CENTRO DE CUSTO';
    mCC[cc] = (mCC[cc] || 0) + r.valor;
  }});
  const entries = Object.entries(mCC).sort((a, b) => b[1] - a[1]);
  if (!entries.length) return;
  const total = entries.reduce((s, e) => s + e[1], 0);
  const alturaCC = Math.max(260, entries.length * 42);
  const wrapCC = document.getElementById('wrapPagarCC');
  if (wrapCC) wrapCC.style.height = alturaCC + 'px';
  const isDark  = document.body.getAttribute('data-theme') !== 'light';
  const gridClr = isDark ? '#334155' : '#e2e8f0';
  const txtClr  = isDark ? '#cbd5e1' : '#1e293b';
  const CORES_CC = ['#6366f1','#8b5cf6','#a78bfa','#c4b5fd','#ddd6fe','#818cf8','#4f46e5','#4338ca'];
  charts['chartPagarCC'] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels: entries.map(e => e[0]),
      datasets: [{{
        label: 'Despesas',
        data: entries.map(e => Math.round(e[1] * 100) / 100),
        backgroundColor: entries.map((_, i) => CORES_CC[i % CORES_CC.length]),
        borderRadius: 4, borderSkipped: false
      }}]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false, indexAxis: 'y',
      plugins: {{
        legend: {{ display: false }},
        subtitle: {{ display: true, text: 'Total: ' + BRL(total),
          color: '#6366f1', font: {{ size: 12, weight: 'bold' }}, padding: {{ bottom: 6 }} }},
        tooltip: {{ callbacks: {{ label: c => BRL(c.raw) + ' (' + ((c.raw / total) * 100).toFixed(1) + '%)' }} }}
      }},
      scales: {{
        x: {{ grid: {{ color: gridClr }}, ticks: {{ color: txtClr, callback: v => v >= 1000 ? 'R$' + (v/1000).toFixed(0) + 'k' : 'R$' + v }} }},
        y: {{ grid: {{ display: false }}, ticks: {{ color: txtClr, font: {{ size: 10 }} }} }}
      }}
    }}
  }});
}}

function mkPagarMensal() {{
  destroyChart('chartPagarMensal');
  const canvas = document.getElementById('chartPagarMensal');
  if (!canvas) return;
  const hoje = new Date();
  const meses = [];
  for (let i = -12; i <= 12; i++) {{
    const d = new Date(hoje.getFullYear(), hoje.getMonth() + i, 1);
    meses.push({{ key: d.getFullYear() + '-' + String(d.getMonth()+1).padStart(2,'0'),
                  label: d.toLocaleDateString('pt-BR',{{month:'short',year:'2-digit'}}) }});
  }}
  const pago    = {{}};
  const aVencer = {{}};
  const vencido = {{}};
  const hojeStr = hoje.getFullYear() + '-' + String(hoje.getMonth()+1).padStart(2,'0');
  meses.forEach(m => {{ pago[m.key]=0; aVencer[m.key]=0; vencido[m.key]=0; }});
  pagarFiltFin.forEach(r => {{
    if (!r.venc) return;
    const mk = r.venc.slice(0,7);
    if (!Object.prototype.hasOwnProperty.call(pago, mk)) return;
    if (r.status === 'PAGO') {{
      pago[mk] += r.valor;
    }} else if (mk < hojeStr) {{
      vencido[mk] += r.valor;
    }} else {{
      aVencer[mk] += r.valor;
    }}
  }});
  const labels  = meses.map(m=>m.label);
  const isDark  = document.body.getAttribute('data-theme') !== 'light';
  const gridClr = isDark ? '#334155' : '#e2e8f0';
  const txtClr  = isDark ? '#cbd5e1' : '#1e293b';
  charts['chartPagarMensal'] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels,
      datasets: [
        {{ label: 'Pago',     data: meses.map(m=>pago[m.key]),    backgroundColor: '#22c55e', stack:'s' }},
        {{ label: 'Vencido',  data: meses.map(m=>vencido[m.key]), backgroundColor: '#ef4444', stack:'s' }},
        {{ label: 'A Vencer', data: meses.map(m=>aVencer[m.key]), backgroundColor: '#38bdf8', stack:'s' }}
      ]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{
        legend: {{ labels: {{ color: txtClr, font: {{size:11}}, boxWidth:12 }} }},
        tooltip: {{ callbacks: {{ label: c => c.dataset.label+': '+BRL(c.raw) }} }}
      }},
      scales: {{
        x: {{ stacked:true, grid:{{color:gridClr}}, ticks:{{color:txtClr, maxRotation:45, font:{{size:9}}}} }},
        y: {{ stacked:true, grid:{{color:gridClr}}, ticks:{{color:txtClr, callback:v=>v>=1000?'R$'+(v/1000).toFixed(0)+'k':'R$'+v}} }}
      }}
    }}
  }});
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  M\u00D3DULO PONTO COLABORADOR
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550

// Calcula horas trabalhadas por funcion\u00E1rio/dia pareando as batidas
function calcJornadas(marc) {{
  const LIMITE = 9 * 60; // 9h em minutos (07:00-17:00 ou 06:30-16:30, 1h almo\u00E7o)
  const byFD = {{}};
  marc.forEach(r => {{
    const k = r.funcionario + '|' + r.data;
    if (!byFD[k]) byFD[k] = {{func: r.funcionario, func_id: r.funcionario_id, data: r.data, mins: []}};
    const [hh, mm] = r.hora.split(':').map(Number);
    byFD[k].mins.push(hh * 60 + mm);
  }});
  return Object.values(byFD).map(e => {{
    const pts = e.mins.slice().sort((a, b) => a - b);
    // Detec\u00E7\u00E3o de turno noturno: span > 10h E primeira batida antes das 07:00
    // Situa\u00E7\u00E3o: sa\u00EDda da noite anterior (ex: 05:30) misturada com entrada da noite seguinte (ex: 19:30)
    const SPAN_MAX    = 10 * 60; // 600 min \u2014 acima disso suspeito de mix noturno
    const HORA_LIMITE =  7 * 60; // 07:00 \u2014 batida antes disso = fim de turno noturno
    const HORA_NOITE  = 20 * 60; // 20:00 \u2014 turno noturno s\u00F3 se \u00FAltima batida for ap\u00F3s 20h
    const span0 = pts.length >= 2 ? pts[pts.length - 1] - pts[0] : 0;
    let turnoFora = false;
    let ptsCalc   = pts;
    // S\u00F3 remove batidas madrugada se: span > 10h E primeiro < 07:00 E \u00FAltimo > 20:00
    // (trabalhadores diurnos que chegam cedo N\u00C3O devem ser afetados)
    if (span0 > SPAN_MAX && pts[0] < HORA_LIMITE && pts[pts.length - 1] > HORA_NOITE) {{
      turnoFora = true;
      // Remove batidas da madrugada (fim do turno anterior) para n\u00E3o inflar HE
      ptsCalc = pts.filter(m => m >= HORA_LIMITE);
    }}
    // Par a par: (sa\u00EDda_almo\u00E7o - entrada) + (volta - sa\u00EDda_final) \u2014 exato com 4 batidas
    let minTrab = 0;
    if (ptsCalc.length >= 4) {{
      // 4+ batidas: soma dos pares (entrada\u2192sa\u00EDda, volta\u2192sa\u00EDda)
      for (let i = 0; i + 1 < ptsCalc.length; i += 2) {{
        minTrab += ptsCalc[i + 1] - ptsCalc[i];
      }}
    }} else if (ptsCalc.length >= 2) {{
      // 2 batidas: span - estimativa almo\u00E7o
      const span   = ptsCalc[ptsCalc.length - 1] - ptsCalc[0];
      const almoco = span > 240 ? 60 : 0;
      minTrab = Math.max(0, span - almoco);
    }}
    minTrab = Math.max(0, minTrab);
    const hTrab   = minTrab / 60;
    const hReg    = Math.min(minTrab, LIMITE) / 60;
    const hExtra  = Math.max(0, minTrab - LIMITE) / 60;
    return {{func: e.func, func_id: e.func_id, data: e.data,
             nMarc: pts.length, minTrab, hTrab, hReg, hExtra, turnoFora}};
  }});
}}

function mkPontoAusencias(marc) {{
  destroyChart('chartPontoAus');
  const canvas = document.getElementById('chartPontoAus');
  if (!canvas) return;

  // Apenas dias \u00FAteis (seg-sex) dentro do per\u00EDodo
  const diasUteis = [...new Set(marc.map(r => r.data))]
    .filter(d => {{ const dow = new Date(d + 'T00:00:00').getDay(); return dow >= 1 && dow <= 5; }})
    .sort();
  if (!diasUteis.length || !PONTO_FUNC.length) return;

  const JORNADA = 9; // horas esperadas por dia \u00FAtil (07:00-17:00, 1h almo\u00E7o)
  const byFD = {{}};
  marc.forEach(r => {{
    const k = r.funcionario_id + '|' + r.data;
    if (!byFD[k]) byFD[k] = [];
    const [hh, mm] = r.hora.split(':').map(Number);
    byFD[k].push(hh * 60 + mm);
  }});

  const result = PONTO_FUNC.map(f => {{
    let dFalta = 0, dAntecip = 0;
    diasUteis.forEach(dia => {{
      const k = f.id + '|' + dia;
      if (!byFD[k]) {{
        // Nenhuma batida no dia \u00FAtil = falta completa
        dFalta += 1;
      }} else {{
        const pts = byFD[k].slice().sort((a, b) => a - b);
        let minTrabAus = 0;
        if (pts.length >= 4) {{
          for (let i = 0; i + 1 < pts.length; i += 2) minTrabAus += pts[i + 1] - pts[i];
        }} else if (pts.length >= 2) {{
          const span = pts[pts.length - 1] - pts[0];
          minTrabAus = Math.max(0, span - (span > 240 ? 60 : 0));
        }}
        const trab = minTrabAus / 60;
        const faltou = Math.max(0, JORNADA - trab);
        if (faltou >= JORNADA) {{
          // S\u00F3 1 batida = dia incompleto \u2192 conta como sa\u00EDda antecipada
          dAntecip += 1;
        }} else if (faltou > 0.5) {{
          // Trabalhou menos de 7.5h = sa\u00EDda antecipada
          dAntecip += 1;
        }}
      }}
    }});
    const partes = f.nome.split(' ');
    const label  = partes[0] + (partes.length > 1 ? ' ' + partes[partes.length - 1] : '');
    return {{ label, dFalta, dAntecip }};
  }}).filter(r => r.dFalta > 0 || r.dAntecip > 0)
    .sort((a, b) => (b.dFalta + b.dAntecip) - (a.dFalta + a.dAntecip));

  const totalDias = result.reduce((s,r)=>s+r.dFalta+r.dAntecip, 0);
  const alturaAus = Math.max(280, result.length * 40);
  const wrapAus = document.getElementById('wrapPontoAus');
  if (wrapAus) wrapAus.style.height = alturaAus + 'px';
  charts['chartPontoAus'] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels: result.map(r=>r.label),
      datasets: [
        {{ label: 'Falta completa', data: result.map(r=>r.dFalta),
           backgroundColor: '#ef4444', stack: 's', borderRadius: 0 }},
        {{ label: 'Sa\u00EDda antecipada', data: result.map(r=>r.dAntecip),
           backgroundColor: '#f59e0b', stack: 's', borderRadius: 4 }},
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false, indexAxis: 'y',
      plugins: {{
        legend: {{ labels: {{ color:'#cbd5e1', font:{{size:11}}, boxWidth:12 }} }},
        subtitle: {{ display:true, text:'Total: '+totalDias+' ocorr\u00EAncia(s) no per\u00EDodo',
          color:'#fca5a5', font:{{size:12,weight:'bold'}}, padding:{{bottom:6}} }},
        tooltip: {{ callbacks: {{ label: c => c.dataset.label+': '+c.raw+' dia(s)' }} }}
      }},
      scales: {{
        x: {{ stacked:true, ticks:{{color:'#cbd5e1',stepSize:1,callback:v=>Number.isInteger(v)?v:''}},
             grid:{{color:'#334155'}},
             title:{{display:true,text:'Dias',color:'#94a3b8'}} }},
        y: {{ stacked:true, ticks:{{color:'#cbd5e1',font:{{size:10}}}}, grid:{{display:false}} }}
      }}
    }}
  }});
}}

function mkPontoFuncChart(marc) {{
  destroyChart('chartPontoFunc');
  const m = {{}};
  marc.forEach(r => {{ m[r.funcionario] = (m[r.funcionario]||0)+1; }});
  const entries = Object.entries(m).sort((a,b)=>b[1]-a[1]);
  if (!entries.length) return;
  charts['chartPontoFunc'] = new Chart(document.getElementById('chartPontoFunc'), {{
    type: 'bar',
    data: {{
      labels: entries.map(e=>e[0].split(' ')[0]),
      datasets: [{{ data: entries.map(e=>e[1]),
        backgroundColor: CORES, borderRadius: 4, borderSkipped: false }}]
    }},
    options: {{
      responsive:true, maintainAspectRatio:false, indexAxis:'y',
      plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>c.raw+' batidas'}}}}}},
      scales:{{
        x:{{ticks:{{color:'#cbd5e1',stepSize:1}},grid:{{color:'#334155'}}}},
        y:{{ticks:{{color:'#cbd5e1',font:{{size:10}}}},grid:{{display:false}}}}
      }}
    }}
  }});
}}

function mkPontoHoras(marc) {{
  destroyChart('chartPontoHoras');
  const canvas = document.getElementById('chartPontoHoras');
  if (!canvas) return;
  const jornadas = calcJornadas(marc);
  if (!jornadas.length) return;
  // Agrega por funcion\u00E1rio: soma hReg e hExtra de todos os dias
  const byFunc = {{}};
  jornadas.forEach(j => {{
    if (!byFunc[j.func]) byFunc[j.func] = {{hReg: 0, hExtra: 0}};
    byFunc[j.func].hReg   += j.hReg;
    byFunc[j.func].hExtra += j.hExtra;
  }});
  const entries = Object.entries(byFunc).sort((a, b) => (b[1].hReg + b[1].hExtra) - (a[1].hReg + a[1].hExtra));
  // Altura din\u00E2mica: m\u00EDn 260px, 40px por funcion\u00E1rio
  const alturaHoras = Math.max(260, entries.length * 40);
  const wrapH = document.getElementById('wrapPontoHoras');
  if (wrapH) wrapH.style.height = alturaHoras + 'px';
  charts['chartPontoHoras'] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels: entries.map(e => e[0].split(' ')[0]),
      datasets: [
        {{ label: 'Regular (\u22649h/dia)', data: entries.map(e => Math.round(e[1].hReg * 10) / 10),
           backgroundColor: '#059669', stack: 's', borderRadius: 0 }},
        {{ label: 'Hora Extra (>9h/dia)', data: entries.map(e => Math.round(e[1].hExtra * 10) / 10),
           backgroundColor: '#f97316', stack: 's', borderRadius: 4 }},
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false, indexAxis: 'y',
      plugins: {{
        legend: {{ labels: {{ color: '#cbd5e1', font: {{size: 11}}, boxWidth: 12 }} }},
        tooltip: {{ callbacks: {{ label: c => c.dataset.label + ': ' + c.raw + 'h' }} }}
      }},
      scales: {{
        x: {{ stacked: true, ticks: {{ color: '#cbd5e1', callback: v => v + 'h' }}, grid: {{ color: '#334155' }},
              title: {{ display: true, text: 'Horas no Per\u00EDodo', color: '#94a3b8' }} }},
        y: {{ stacked: true, ticks: {{ color: '#cbd5e1', font: {{ size: 10 }} }}, grid: {{ display: false }} }}
      }}
    }}
  }});
}}

function mkPontoHoraExtra(marc) {{
  destroyChart('chartPontoHoraExtra');
  const canvas = document.getElementById('chartPontoHoraExtra');
  if (!canvas) return;
  // Agrega hExtra por funcionario em todo o periodo
  const byFunc = {{}};
  calcJornadas(marc).forEach(j => {{
    if (j.hExtra > 0) {{
      byFunc[j.func] = (byFunc[j.func] || 0) + j.hExtra;
    }}
  }});
  const entries = Object.entries(byFunc).sort((a, b) => b[1] - a[1]);
  if (!entries.length) {{
    destroyChart('chartPontoHoraExtra');
    canvas.style.display = 'none';
    const prev = document.getElementById('chartPontoHoraExtra-nodata');
    if (!prev) {{
      const msg = document.createElement('div');
      msg.id = 'chartPontoHoraExtra-nodata';
      msg.style.cssText = 'display:flex;align-items:center;justify-content:center;height:260px;color:#94a3b8;font-size:13px;text-align:center;flex-direction:column;gap:6px';
      msg.innerHTML = '<span style="font-size:28px">\u2705</span>Nenhuma hora extra no per\u00edodo';
      canvas.parentNode.insertBefore(msg, canvas.nextSibling);
    }}
    return;
  }}
  // Remove mensagem de sem dados se existir
  const nodata = document.getElementById('chartPontoHoraExtra-nodata');
  if (nodata) nodata.remove();
  canvas.style.display = '';
  const totalHE = entries.reduce((s, e) => s + e[1], 0);
  const alturaHE = Math.max(260, entries.length * 40);
  const wrapHE = document.getElementById('wrapPontoHoraExtra');
  if (wrapHE) wrapHE.style.height = alturaHE + 'px';
  charts['chartPontoHoraExtra'] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels: entries.map(e => e[0].split(' ')[0]),
      datasets: [{{
        label: 'Hora Extra',
        data: entries.map(e => Math.round(e[1] * 10) / 10),
        backgroundColor: entries.map((_, i) => i === 0 ? '#f97316' : i === 1 ? '#fb923c' : '#fed7aa'),
        borderRadius: 4, borderSkipped: false
      }}]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false, indexAxis: 'y',
      plugins: {{
        legend: {{ display: false }},
        subtitle: {{ display: true,
          text: 'Total no per\u00EDodo: ' + totalHE.toFixed(1) + 'h',
          color: '#f97316', font: {{ size: 12, weight: 'bold' }}, padding: {{ bottom: 6 }} }},
        tooltip: {{ callbacks: {{ label: c => c.raw + 'h de HE' }} }}
      }},
      scales: {{
        x: {{ ticks: {{ color: '#cbd5e1', callback: v => v + 'h' }}, grid: {{ color: '#334155' }},
             title: {{ display: true, text: 'Horas Extras no Per\u00EDodo', color: '#94a3b8' }} }},
        y: {{ ticks: {{ color: '#cbd5e1', font: {{ size: 10 }} }}, grid: {{ display: false }} }}
      }}
    }}
  }});
}}

function renderTblPontoHoje(marc) {{
  // Dia de refer\u00EAncia = \u00FAltimo dia do per\u00EDodo filtrado (ou hoje se in range)
  const hoje    = new Date().toISOString().substring(0,10);
  const diasMarc = [...new Set(marc.map(r=>r.data))].sort();
  const diaRef  = diasMarc.includes(hoje) ? hoje : (diasMarc[diasMarc.length-1] || hoje);
  const labelDia = diaRef.split('-').reverse().join('/');
  // Atualiza t\u00EDtulo da tabela
  const th3 = document.querySelector('#tblPontoHoje')?.closest('.table-card')?.querySelector('h3');
  if (th3) th3.textContent = 'Marca\u00E7\u00F5es de ' + labelDia + ' por funcion\u00E1rio';
  // Pre-calcula horas por funcion\u00E1rio/dia via calcJornadas
  const jDia = {{}};
  calcJornadas(marc.filter(r => r.data === diaRef)).forEach(j => {{
    jDia[j.func] = j;
  }});
  const byFunc  = {{}};
  marc.filter(r=>r.data===diaRef).forEach(r => {{
    if (!byFunc[r.funcionario]) byFunc[r.funcionario] = {{func:r.funcionario,horas:[],manuais:0}};
    const [hh,mm] = r.hora.split(':').map(Number);
    byFunc[r.funcionario].horas.push(hh*60+mm);
    if (r.origem_id === 2) byFunc[r.funcionario].manuais = (byFunc[r.funcionario].manuais||0)+1;
  }});
  PONTO_FUNC.forEach(f => {{
    if (!byFunc[f.nome]) byFunc[f.nome] = {{func:f.nome,horas:[],ausente:true,manuais:0}};
  }});
  const arr = Object.values(byFunc).sort((a,b)=>a.func>b.func?1:-1);
  let html = '';
  arr.forEach((r,i) => {{
    const hrs     = r.horas.sort((a,b)=>a-b);
    const fmt     = m => Math.floor(m/60)+':'+String(m%60).padStart(2,'0');
    const entrada     = hrs.length > 0 ? fmt(hrs[0]) : '\u2014';
    const saidaAlmoco = hrs.length >= 4 ? fmt(hrs[1]) : '\u2014';
    const retorno     = hrs.length >= 4 ? fmt(hrs[2]) : (hrs.length === 3 ? fmt(hrs[1]) : '\u2014');
    const saida       = hrs.length > 1 ? fmt(hrs[hrs.length-1]) : '\u2014';
    const jd      = jDia[r.func];
    const hTrabNum = jd ? jd.hTrab : 0;
    const hExtra  = jd ? jd.hExtra : 0;
    let hTrabCell = '\u2014';
    if (hrs.length > 1) {{
      hTrabCell = `<span style="font-weight:700">${{hTrabNum.toFixed(1)}}h</span>`;
      if (hExtra > 0) hTrabCell += ` <span style="color:#f97316;font-size:11px;font-weight:600">+${{hExtra.toFixed(1)}}h\u23F0</span>`;
    }}
    const foraJornada = jd && jd.turnoFora;
    const heStyle = (!r.ausente && hExtra > 0) ? ' style="background:rgba(249,115,22,0.07)"'
      : foraJornada ? ' style="background:rgba(99,102,241,0.07)"' : '';
    const heStatus = (!r.ausente && hExtra > 0)
      ? ' <span class="badge" style="background:#f97316;color:#fff;font-size:10px">HE</span>' : '';
    const foraStatus = foraJornada
      ? ' <span class="badge" style="background:#6366f1;color:#fff;font-size:10px" title="Batidas fora do hor\u00E1rio da jornada cadastrada">\uD83C\uDF19 Fora Jornada</span>' : '';
    const status  = r.ausente
      ? '<span class="badge vermelho">SEM REGISTRO</span>'
      : hrs.length >= 4 ? `<span class="badge verde">COMPLETO</span>${{heStatus}}${{foraStatus}}`
      : hrs.length >= 2 ? `<span class="badge amarelo">PARCIAL</span>${{heStatus}}${{foraStatus}}`
      : `<span class="badge amarelo">ENTRADA</span>${{foraStatus}}`;
    const origemCell = r.ausente ? '\u2014'
      : r.manuais > 0 ? `<span style="color:#f59e0b;font-weight:600">\u270F\uFE0F Manual (${{r.manuais}})</span>`
      : '<span style="color:#94a3b8;font-size:11px">Autom\u00E1tico</span>';
    html += `<tr${{heStyle}}><td>${{i+1}}</td><td>${{r.func}}</td>
      <td class="num">${{entrada}}</td><td class="num" style="color:#94a3b8">${{saidaAlmoco}}</td><td class="num" style="color:#94a3b8">${{retorno}}</td><td class="num">${{saida}}</td>
      <td class="num">${{hTrabCell}}</td><td>${{status}}</td><td>${{origemCell}}</td></tr>`;
  }});
  const el = document.getElementById('tblPontoHoje');
  if (el) el.innerHTML = html || '<tr><td colspan="9" style="text-align:center;color:#94a3b8">Sem dados de ponto dispon\u00EDveis para hoje</td></tr>';
}}

function renderTblPontoUlt(marc) {{
  const arr = [...marc].sort((a,b)=>{{
    if (b.data > a.data) return 1; if (a.data > b.data) return -1;
    return b.hora > a.hora ? 1 : -1;
  }}).slice(0,50);
  let html = '';
  arr.forEach(r => {{
    const isManual = r.origem_id === 2;
    const origemCell = isManual
      ? `<span style="color:#f59e0b;font-weight:600">\u270F\uFE0F Manual<br><span style="font-size:10px;font-weight:400;color:#fcd34d">${{r.descricao||''}}</span></span>`
      : `<span style="color:#94a3b8;font-size:11px">${{r.origem||'Autom\u00E1tico'}}</span>`;
    html += `<tr${{isManual?' style="background:rgba(245,158,11,0.07)"':''}}>\n      <td>${{r.data?r.data.split('-').reverse().join('/'):'\u2014'}}</td>\n      <td>${{r.funcionario||'\u2014'}}</td>\n      <td class="num" style="font-weight:700;color:#38bdf8">${{r.hora||'\u2014'}}</td>\n      <td>${{origemCell}}</td>\n    </tr>`;
  }});
  const el = document.getElementById('tblPontoUlt');
  if (el) el.innerHTML = html || '<tr><td colspan="4" style="text-align:center;color:#94a3b8">Sem marca\u00E7\u00F5es no per\u00EDodo selecionado</td></tr>';
}}

function renderTblPontoJustif(marc) {{
  const manuais = marc.filter(r => r.origem_id === 2 && r.descricao && r.descricao.trim());
  const el = document.getElementById('tblPontoJustif');
  if (!el) return;
  if (!manuais.length) {{
    el.innerHTML = '<tr><td colspan="4" style="text-align:center;color:#94a3b8">Nenhum ajuste manual com justificativa no per\u00EDodo</td></tr>';
    return;
  }}
  const grouped = {{}};
  manuais.forEach(r => {{
    const motivo = r.descricao.trim();
    if (!grouped[motivo]) grouped[motivo] = {{count: 0, funcs: new Set()}};
    grouped[motivo].count++;
    grouped[motivo].funcs.add(r.funcionario.split(' ')[0]);
  }});
  const arr = Object.entries(grouped).sort((a, b) => b[1].count - a[1].count);
  let html = '';
  arr.forEach(([motivo, info], i) => {{
    const funcs = [...info.funcs].join(', ');
    html += `<tr><td>${{i+1}}</td><td style="color:#fcd34d">${{motivo}}</td>
      <td class="num" style="font-weight:700;color:#f59e0b">${{info.count}}</td>
      <td style="font-size:12px;color:#94a3b8">${{funcs}}</td></tr>`;
  }});
  el.innerHTML = html;
}}

function initPonto(marc) {{
  marc = marc || PONTO_MARC;
  // Dia de refer\u00EAncia = \u00FAltimo dia do per\u00EDodo filtrado (ou hoje se in range)
  const hoje    = new Date().toISOString().substring(0,10);
  const diasMarc = [...new Set(marc.map(r=>r.data))].sort();
  const diaRef   = diasMarc.includes(hoje) ? hoje : (diasMarc[diasMarc.length-1] || hoje);
  const el = id => document.getElementById(id);
  if (el('kPontoFunc'))  el('kPontoFunc').textContent  = PONTO_FUNC.length || '0';
  if (el('kPontoTotal')) el('kPontoTotal').textContent = marc.length || '0';
  const presHoje   = new Set(marc.filter(r=>r.data===diaRef).map(r=>r.funcionario_id)).size;
  const nManuais   = marc.filter(r=>r.origem_id===2).length;
  const jHoje      = calcJornadas(marc.filter(r => r.data === diaRef));
  const nHoraExtra = jHoje.filter(j => j.hExtra > 0).length;
  if (el('kPontoHoje'))   el('kPontoHoje').textContent   = presHoje || '0';
  if (el('kPontoAus'))    el('kPontoAus').textContent    = Math.max(0, PONTO_FUNC.length - presHoje) || '0';
  if (el('kPontoManual')) el('kPontoManual').textContent = nManuais || '0';
  if (el('kPontoHE'))     el('kPontoHE').textContent     = nHoraExtra || '0';
  mkPontoAusencias(marc);
  mkPontoFuncChart(marc);
  mkPontoHoras(marc);
  mkPontoHoraExtra(marc);
  renderTblPontoHoje(marc);
  renderTblPontoUlt(marc);
  renderTblPontoJustif(marc);
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  M\u00D3DULO OR\u00C7AMENTO
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
let _orcFiltro = 'todos';

function filtrarOrc(btn) {{
  _orcFiltro = btn.dataset.f;
  document.querySelectorAll('.btn-filter-orc').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  renderTblOrc();
}}

function _downloadOrcPdf(idx) {{
  const r = ORCAMENTOS[idx];
  if (!r || !r.pdf_b64) {{ alert('PDF não disponível para este orçamento.'); return; }}
  try {{
    const bin = atob(r.pdf_b64);
    const arr = new Uint8Array(bin.length);
    for (let i = 0; i < bin.length; i++) arr[i] = bin.charCodeAt(i);
    const blob = new Blob([arr], {{type:'application/pdf'}});
    const url  = URL.createObjectURL(blob);
    const a    = document.createElement('a');
    a.href = url; a.download = 'Proposta_' + (r.codigo || r.id) + '.pdf';
    document.body.appendChild(a); a.click();
    document.body.removeChild(a); URL.revokeObjectURL(url);
  }} catch(e) {{ alert('Erro ao abrir PDF: ' + e); }}
}}

function renderTblOrc() {{
  const tbody = document.getElementById('tblOrcBody');
  if (!tbody) return;
  const lista = _orcFiltro === 'todos' ? ORCAMENTOS
    : ORCAMENTOS.filter(r => r.situacao === _orcFiltro);
  const sorted = [...lista].sort((a, b) => (b.data || '').localeCompare(a.data || ''));
  const badgeMap = {{'Em aberto':'yellow','Em andamento':'blue','Concretizado':'green','Cancelado':'red'}};
  const iconMap  = {{'Em aberto':'\uD83D\uDCC2','Em andamento':'\u23F3','Concretizado':'\u2705','Cancelado':'\u274C'}};
  tbody.innerHTML = sorted.map((r, i) => {{
    const cor   = badgeMap[r.situacao] || 'gray';
    const icon  = iconMap[r.situacao]  || '';
    const badge = `<span class="badge ${{cor}}">${{icon}} ${{r.situacao}}</span>`;
    const val   = r.valor > 0 ? BRL(r.valor) : '\u2014';
    const origIdx = ORCAMENTOS.indexOf(r);
    const pdfBtn  = r.pdf_b64
      ? `<button onclick="_downloadOrcPdf(${{origIdx}})" style="background:#1e40af;color:#fff;border:none;border-radius:4px;padding:2px 8px;font-size:11px;cursor:pointer;white-space:nowrap">\uD83D\uDCC4 PDF</button>`
      : `<span style="color:#94a3b8;font-size:10px">—</span>`;
    return `<tr>
      <td>${{i+1}}</td>
      <td style="font-size:11px;color:#94a3b8">${{r.codigo}}</td>
      <td style="white-space:nowrap">${{r.data || '\u2014'}}</td>
      <td><strong>${{r.cliente || '\u2014'}}</strong></td>
      <td style="font-size:12px">${{r.vendedor || '\u2014'}}</td>
      <td style="font-size:12px">${{r.centro_custo || '\u2014'}}</td>
      <td style="font-size:12px">${{r.loja || '\u2014'}}</td>
      <td class="num">${{val}}</td>
      <td>${{badge}}</td>
      <td style="text-align:center">${{pdfBtn}}</td>
    </tr>`;
  }}).join('') || '<tr><td colspan="10" style="text-align:center;color:#94a3b8">Nenhum orçamento</td></tr>';
}}

function mkOrcamento() {{
  if (!ORCAMENTOS.length) return;
  const abertos   = ORCAMENTOS.filter(r => r.situacao === 'Em aberto');
  const andamento = ORCAMENTOS.filter(r => r.situacao === 'Em andamento');
  const concret   = ORCAMENTOS.filter(r => r.situacao === 'Concretizado');
  const cancelados= ORCAMENTOS.filter(r => r.situacao === 'Cancelado');
  const valTotal  = ORCAMENTOS.reduce((s, r) => s + (r.valor || 0), 0);

  const _set = (id, v) => {{ const el = document.getElementById(id); if (el) el.textContent = v; }};
  _set('kOrcTotal',     NUM(ORCAMENTOS.length));
  _set('kOrcAberto',    NUM(abertos.length));
  _set('kOrcAndamento', NUM(andamento.length));
  _set('kOrcConc',      NUM(concret.length));
  _set('kOrcCanc',      NUM(cancelados.length));
  _set('kOrcValTotal',  BRL(valTotal));

  renderTblOrc();

  // ── Gráfico: propostas por vendedor (stacked por situação)
  destroyChart('chartOrcVendedor');
  const cv = document.getElementById('chartOrcVendedor');
  if (cv) {{
    const vends = [...new Set(ORCAMENTOS.map(r => r.vendedor || 'Sem Vendedor'))].sort();
    const isDark = document.body.getAttribute('data-theme') !== 'light';
    const txtClr = isDark ? '#cbd5e1' : '#1e293b';
    const gridClr = isDark ? '#334155' : '#e2e8f0';
    const sits = ['Em aberto','Em andamento','Concretizado','Cancelado'];
    const cores = ['#eab308','#3b82f6','#059669','#ef4444'];
    charts['chartOrcVendedor'] = new Chart(cv, {{
      type: 'bar',
      data: {{
        labels: vends,
        datasets: sits.map((s,idx) => ({{
          label: s,
          data: vends.map(v => ORCAMENTOS.filter(r => (r.vendedor||'Sem Vendedor')===v && r.situacao===s).length),
          backgroundColor: cores[idx]
        }}))
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{ legend: {{ labels: {{ color: txtClr }} }}, tooltip: {{ callbacks: {{ label: c => c.dataset.label + ': ' + c.raw }} }} }},
        scales: {{
          x: {{ stacked: true, grid: {{ color: gridClr }}, ticks: {{ color: txtClr }} }},
          y: {{ stacked: true, grid: {{ color: gridClr }}, ticks: {{ color: txtClr, stepSize: 1 }}, beginAtZero: true }}
        }}
      }}
    }});
  }}

  // ── Gráfico: donut situação
  destroyChart('chartOrcTaxa');
  const ct = document.getElementById('chartOrcTaxa');
  if (ct) {{
    const isDark = document.body.getAttribute('data-theme') !== 'light';
    const txtClr = isDark ? '#cbd5e1' : '#1e293b';
    charts['chartOrcTaxa'] = new Chart(ct, {{
      type: 'doughnut',
      data: {{
        labels: ['Em aberto', 'Em andamento', 'Concretizado', 'Cancelado'],
        datasets: [{{ data: [abertos.length, andamento.length, concret.length, cancelados.length], backgroundColor: ['#eab308','#3b82f6','#059669','#ef4444'], borderWidth: 0 }}]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{
          legend: {{ position: 'bottom', labels: {{ color: txtClr }} }},
          tooltip: {{ callbacks: {{ label: c => c.label + ': ' + c.raw + ' (' + PCT(c.raw, ORCAMENTOS.length) + ')' }} }}
        }}
      }}
    }});
  }}
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  BOLETINS DE MEDI\u00C7\u00C3O
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
function mkMedicoes() {{
  const _set = (id, v) => {{ const el = document.getElementById(id); if (el) el.textContent = v; }};
  const vazio  = document.getElementById('secMedicaoVazio');
  const kpis   = document.getElementById('secMedicaoKpis');
  const secChartsMed = document.getElementById('secMedicaoCharts');
  const tabela = document.getElementById('secMedicaoTabela');

  if (!MEDICOES || !MEDICOES.length) {{
    if (vazio) vazio.style.display = '';
    if (kpis)         kpis.style.display         = 'none';
    if (secChartsMed) secChartsMed.style.display = 'none';
    if (tabela) tabela.style.display = 'none';
    return;
  }}
  if (vazio) vazio.style.display = 'none';
  if (kpis)         kpis.style.display         = '';
  if (secChartsMed) secChartsMed.style.display = '';
  if (tabela) tabela.style.display = '';

  // \u2500\u2500 KPIs
  const totBoletins  = MEDICOES.length;
  const totHrDia     = MEDICOES.reduce((s, r) => s + (r.horas_diurno  || 0), 0);
  const totHrNot     = MEDICOES.reduce((s, r) => s + (r.horas_noturno || 0), 0);
  const totHrExt     = MEDICOES.reduce((s, r) => s + (r.horas_extra   || 0), 0);
  const totFaturado  = MEDICOES.reduce((s, r) => s + (r.total_pagar   || 0), 0);

  _set('kMedTotal',   NUM(totBoletins));
  _set('kMedHrDia',   totHrDia.toLocaleString('pt-BR', {{maximumFractionDigits:1}}) + 'h');
  _set('kMedHrNot',   totHrNot.toLocaleString('pt-BR', {{maximumFractionDigits:1}}) + 'h');
  _set('kMedHrExt',   totHrExt.toLocaleString('pt-BR', {{maximumFractionDigits:1}}) + 'h');
  _set('kMedTotal$',  BRL(totFaturado));

  // \u2500\u2500 Tabela
  const tbody = document.getElementById('tblMedicoesBdy');
  const tfoot = document.getElementById('tblMedicoesFoot');
  if (tbody) {{
    tbody.innerHTML = MEDICOES.map(r => `
      <tr>
        <td><strong>${{r.numero || '\u2014'}}</strong></td>
        <td>${{r.equipamento || '\u2014'}}</td>
        <td>${{r.cliente || '\u2014'}}</td>
        <td style="white-space:nowrap">${{r.periodo || '\u2014'}}</td>
        <td class="num">${{(r.horas_diurno  || 0).toLocaleString('pt-BR',{{maximumFractionDigits:1}})}}h</td>
        <td class="num">${{(r.horas_noturno || 0).toLocaleString('pt-BR',{{maximumFractionDigits:1}})}}h</td>
        <td class="num">${{(r.horas_extra   || 0).toLocaleString('pt-BR',{{maximumFractionDigits:1}})}}h</td>
        <td class="num">${{BRL(r.valor_hora_dia || 0)}}</td>
        <td class="num">${{BRL(r.valor_hora_not || 0)}}</td>
        <td class="num">${{BRL(r.valor_diurno   || 0)}}</td>
        <td class="num">${{BRL(r.valor_noturno  || 0)}}</td>
        <td class="num">${{BRL(r.valor_extra    || 0)}}</td>
        <td class="num">${{BRL(r.valor_desmobi  || 0)}}</td>
        <td class="num"><strong>${{BRL(r.total_medicao || 0)}}</strong></td>
        <td class="num"><strong style="color:#059669">${{BRL(r.total_pagar   || 0)}}</strong></td>
      </tr>`).join('');
  }}
  if (tfoot) {{
    const totDesmobi = MEDICOES.reduce((s, r) => s + (r.valor_desmobi || 0), 0);
    const totMed     = MEDICOES.reduce((s, r) => s + (r.total_medicao || 0), 0);
    const totDia$    = MEDICOES.reduce((s, r) => s + (r.valor_diurno  || 0), 0);
    const totNot$    = MEDICOES.reduce((s, r) => s + (r.valor_noturno || 0), 0);
    const totExt$    = MEDICOES.reduce((s, r) => s + (r.valor_extra   || 0), 0);
    tfoot.innerHTML = `<tr style="font-weight:700;background:#f0fdf4">
      <td colspan="4">TOTAL (${{totBoletins}} boletim${{totBoletins!==1?'s':''}})</td>
      <td class="num">${{totHrDia.toLocaleString('pt-BR',{{maximumFractionDigits:1}})}}h</td>
      <td class="num">${{totHrNot.toLocaleString('pt-BR',{{maximumFractionDigits:1}})}}h</td>
      <td class="num">${{totHrExt.toLocaleString('pt-BR',{{maximumFractionDigits:1}})}}h</td>
      <td class="num">\u2014</td><td class="num">\u2014</td>
      <td class="num">${{BRL(totDia$)}}</td>
      <td class="num">${{BRL(totNot$)}}</td>
      <td class="num">${{BRL(totExt$)}}</td>
      <td class="num">${{BRL(totDesmobi)}}</td>
      <td class="num"><strong>${{BRL(totMed)}}</strong></td>
      <td class="num"><strong style="color:#059669">${{BRL(totFaturado)}}</strong></td>
    </tr>`;
  }}

  const isDark  = document.body.getAttribute('data-theme') !== 'light';
  const txtClr  = isDark ? '#cbd5e1' : '#1e293b';
  const gridClr = isDark ? '#334155' : '#e2e8f0';

  // \u2500\u2500 Gr\u00E1fico: Horas por Equipamento (diurno vs noturno)
  destroyChart('chartMedHoras');
  const cvH = document.getElementById('chartMedHoras');
  if (cvH) {{
    const labels = MEDICOES.map(r => (r.equipamento || r.numero || r.arquivo || '\u2014').replace(/\d+\/\d+T$/,'').trim());
    charts['chartMedHoras'] = new Chart(cvH, {{
      type: 'bar',
      data: {{
        labels,
        datasets: [
          {{ label: '\u2600\uFE0F Diurno',  data: MEDICOES.map(r => r.horas_diurno  || 0), backgroundColor: '#0891b2', borderRadius: 4 }},
          {{ label: '\uD83C\uDF19 Noturno', data: MEDICOES.map(r => r.horas_noturno || 0), backgroundColor: '#7c3aed', borderRadius: 4 }},
          {{ label: '\u26A1 Extra',   data: MEDICOES.map(r => r.horas_extra   || 0), backgroundColor: '#f59e0b', borderRadius: 4 }},
        ]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{
          legend: {{ labels: {{ color: txtClr }} }},
          tooltip: {{ callbacks: {{ label: c => c.dataset.label + ': ' + c.raw + 'h' }} }}
        }},
        scales: {{
          x: {{ stacked: true, ticks: {{ color: txtClr }}, grid: {{ color: gridClr }} }},
          y: {{ stacked: true, ticks: {{ color: txtClr, callback: v => v + 'h' }}, grid: {{ color: gridClr }}, beginAtZero: true }}
        }}
      }}
    }});
  }}

  // \u2500\u2500 Gr\u00E1fico: Faturamento por Boletim
  destroyChart('chartMedFat');
  const cvF = document.getElementById('chartMedFat');
  if (cvF) {{
    const labels = MEDICOES.map(r => 'N\u00BA ' + (r.numero || r.arquivo?.slice(0,12) || '\u2014'));
    const cores = MEDICOES.map((_, i) => CORES[i % CORES.length]);
    charts['chartMedFat'] = new Chart(cvF, {{
      type: 'bar',
      data: {{
        labels,
        datasets: [{{
          label: 'Total a Pagar',
          data: MEDICOES.map(r => r.total_pagar || 0),
          backgroundColor: cores, borderRadius: 6
        }}]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{
          legend: {{ display: false }},
          tooltip: {{ callbacks: {{ label: c => BRL(c.raw) }} }}
        }},
        scales: {{
          x: {{ ticks: {{ color: txtClr }}, grid: {{ color: gridClr }} }},
          y: {{ ticks: {{ color: txtClr, callback: v => 'R$' + (v/1000).toFixed(0) + 'k' }}, grid: {{ color: gridClr }}, beginAtZero: true }}
        }}
      }}
    }});
  }}
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  HORAS REGISTRADAS PELO APP (Supabase)
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
function mkHorasApp() {{
  try {{
  const _set = (id, v) => {{ const el = document.getElementById(id); if (el) el.textContent = v; }};
  const vazio        = document.getElementById('secHorasAppVazio');
  const kpis         = document.getElementById('secHorasAppKpis');
  const secChartsApp = document.getElementById('secHorasAppCharts');
  const tabela       = document.getElementById('secHorasAppTabela');

  if (!HORAS_APP || !HORAS_APP.length) {{
    if (vazio)        vazio.style.display        = '';
    if (kpis)         kpis.style.display         = 'none';
    if (secChartsApp) secChartsApp.style.display = 'none';
    if (tabela)       tabela.style.display       = 'none';
    return;
  }}
  if (vazio)        vazio.style.display        = 'none';
  if (kpis)         kpis.style.display         = '';
  if (secChartsApp) secChartsApp.style.display = '';
  if (tabela)       tabela.style.display       = '';

  // \u2500\u2500 KPIs
  const totReg    = HORAS_APP.length;
  const totHoras  = HORAS_APP.reduce((s, r) => s + (parseFloat(r.horas_trabalhadas) || 0), 0);
  const hrsDiurno = HORAS_APP.filter(r => (r.turno || '').toUpperCase() === 'DIURNO')
                             .reduce((s, r) => s + (parseFloat(r.horas_trabalhadas) || 0), 0);
  const hrsNot    = HORAS_APP.filter(r => (r.turno || '').toUpperCase() === 'NOTURNO')
                             .reduce((s, r) => s + (parseFloat(r.horas_trabalhadas) || 0), 0);
  const operadores = new Set(HORAS_APP.map(r => r.operador).filter(Boolean));

  _set('kAppTotal',     NUM(totReg));
  _set('kAppHoras',     totHoras.toLocaleString('pt-BR', {{maximumFractionDigits:1}}) + 'h');
  _set('kAppDiurno',    hrsDiurno.toLocaleString('pt-BR', {{maximumFractionDigits:1}}) + 'h');
  _set('kAppNoturno',   hrsNot.toLocaleString('pt-BR', {{maximumFractionDigits:1}}) + 'h');
  _set('kAppOperadores', NUM(operadores.size));

  // \u2500\u2500 Tabela
  const tbody = document.getElementById('tblHorasAppBdy');
  const tfoot = document.getElementById('tblHorasAppFoot');
  if (tbody) {{
    tbody.innerHTML = HORAS_APP.map(r => {{
      const st = (r.status || 'pendente');
      const stClr = st === 'aprovado' ? '#059669' : '#d97706';
      return `<tr>
        <td style="white-space:nowrap">${{r.data || '\u2014'}}</td>
        <td><strong>${{r.equipamento || '\u2014'}}</strong></td>
        <td>${{r.placa || '\u2014'}}</td>
        <td>${{r.cliente || '\u2014'}}</td>
        <td>${{r.operador || '\u2014'}}</td>
        <td><span style="color:${{(r.turno||'').toUpperCase()==='NOTURNO'?'#7c3aed':'#0891b2'}};font-weight:600">${{r.turno || '\u2014'}}</span></td>
        <td class="num">${{r.hora_inicio || '\u2014'}}</td>
        <td class="num">${{r.hora_fim || '\u2014'}}</td>
        <td class="num"><strong>${{(parseFloat(r.horas_trabalhadas)||0).toLocaleString('pt-BR',{{maximumFractionDigits:1}})}}h</strong></td>
        <td><span style="color:${{stClr}};font-weight:600">${{st}}</span></td>
        <td style="max-width:160px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis" title="${{r.observacoes||''}}">${{r.observacoes || ''}}</td>
      </tr>`;
    }}).join('');
  }}
  if (tfoot) {{
    tfoot.innerHTML = `<tr style="font-weight:700;background:#f0fdf4">
      <td colspan="8">TOTAL (${{totReg}} registro${{totReg!==1?'s':''}})</td>
      <td class="num">${{totHoras.toLocaleString('pt-BR',{{maximumFractionDigits:1}})}}h</td>
      <td colspan="2"></td>
    </tr>`;
  }}

  const isDark  = document.body.getAttribute('data-theme') !== 'light';
  const txtClr  = isDark ? '#cbd5e1' : '#1e293b';
  const gridClr = isDark ? '#334155' : '#e2e8f0';

  // \u2500\u2500 Gr\u00E1fico: Horas por Equipamento
  destroyChart('chartAppEquip');
  const cvE = document.getElementById('chartAppEquip');
  if (cvE) {{
    const mapE = {{}};
    HORAS_APP.forEach(r => {{
      const k = r.equipamento || 'Sem equipamento';
      mapE[k] = (mapE[k] || 0) + (parseFloat(r.horas_trabalhadas) || 0);
    }});
    const sorted = Object.entries(mapE).sort((a,b) => b[1]-a[1]).slice(0,10);
    charts['chartAppEquip'] = new Chart(cvE, {{
      type: 'bar',
      data: {{
        labels: sorted.map(e => e[0]),
        datasets: [{{ label: 'Horas', data: sorted.map(e => +e[1].toFixed(1)),
          backgroundColor: '#0891b2', borderRadius: 4 }}]
      }},
      options: {{
        indexAxis: 'y', responsive: true, maintainAspectRatio: false,
        plugins: {{ legend: {{ display: false }}, tooltip: {{ callbacks: {{ label: c => c.raw + 'h' }} }} }},
        scales: {{
          x: {{ ticks: {{ color: txtClr, callback: v => v + 'h' }}, grid: {{ color: gridClr }}, beginAtZero: true }},
          y: {{ ticks: {{ color: txtClr }}, grid: {{ color: gridClr }} }}
        }}
      }}
    }});
  }}

  // \u2500\u2500 Gr\u00E1fico: Horas por Operador
  destroyChart('chartAppOper');
  const cvO = document.getElementById('chartAppOper');
  if (cvO) {{
    const mapO = {{}};
    HORAS_APP.forEach(r => {{
      const k = r.operador || 'Sem operador';
      mapO[k] = (mapO[k] || 0) + (parseFloat(r.horas_trabalhadas) || 0);
    }});
    const sorted = Object.entries(mapO).sort((a,b) => b[1]-a[1]);
    const cores  = sorted.map((_,i) => CORES[i % CORES.length]);
    charts['chartAppOper'] = new Chart(cvO, {{
      type: 'doughnut',
      data: {{
        labels: sorted.map(e => e[0]),
        datasets: [{{ data: sorted.map(e => +e[1].toFixed(1)),
          backgroundColor: cores, borderWidth: 2 }}]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{
          legend: {{ position: 'right', labels: {{ color: txtClr, boxWidth: 10 }} }},
          tooltip: {{ callbacks: {{ label: c => c.label + ': ' + c.raw + 'h' }} }}
        }}
      }}
    }});
  }}
}} catch(e) {{ console.warn('mkHorasApp:', e); }}
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  NAVEGA\u00C7\u00C3O POR M\u00D3DULOS
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
function setModulo(m) {{
  document.querySelectorAll('.mod-section').forEach(el => {{
    el.style.display = (m === 'geral' || el.dataset.mod === m) ? '' : 'none';
  }});
  document.querySelectorAll('.nav-tab').forEach(btn => {{
    btn.classList.toggle('active', btn.dataset.mod === m);
  }});
  try {{ localStorage.setItem('locvix-modulo', m); }} catch(e) {{}}
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  ATUALIZA\u00C7\u00C3O GERAL
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
function atualizar() {{
  const rows = dadosFilt;
  try {{ atualizarKPIVendas(rows); }} catch(e) {{ console.warn('KPIVendas:', e); }}
  try {{ atualizarKPIFinanceiro(); }} catch(e) {{ console.warn('KPIFin:', e); }}
  // atualizarKPIOS() removido \u2014 elementos HTML foram removidos do m\u00F3dulo Opera\u00E7\u00F5es

  // Gr\u00E1ficos de vendas
  try {{ mkMensal(rows); }} catch(e) {{ console.warn('mkMensal:', e); }}
  try {{ mkCategoria(rows); }} catch(e) {{ console.warn('mkCategoria:', e); }}
  try {{ mkVendasCC(rows); }} catch(e) {{ console.warn('mkVendasCC:', e); }}

  // Top 10 produtos e clientes
  const mProd = {{}};
  rows.forEach(r=>{{ const k=r.cod||r.produto;
    if(!mProd[k]) mProd[k]=[r.produto,0]; mProd[k][1]+=r.liq; }});
  try {{ mkHorizBar('chartProdutos', Object.values(mProd).sort((a,b)=>b[1]-a[1]).slice(0,10)); }} catch(e) {{ console.warn('chartProdutos:', e); }}

  const mCli = {{}};
  rows.forEach(r=>{{ mCli[r.cliente]=(mCli[r.cliente]||0)+r.liq; }});
  try {{ mkHorizBar('chartClientes', Object.entries(mCli).sort((a,b)=>b[1]-a[1]).slice(0,10)); }} catch(e) {{ console.warn('chartClientes:', e); }}

  // Tabelas
  try {{ renderTblProdutos(rows); }} catch(e) {{ console.warn('tblProdutos:', e); }}
  try {{ renderTblClientes(rows); }} catch(e) {{ console.warn('tblClientes:', e); }}
  try {{ mkFinanceiro(); }} catch(e) {{ console.warn('mkFinanceiro:', e); }}
  try {{ initPonto(pontoMarcFilt); }} catch(e) {{ console.warn('initPonto:', e); }}
  try {{ mkOrcamento(); }} catch(e) {{ console.warn('mkOrcamento:', e); }}
  try {{ mkManutencaoCC(); }} catch(e) {{ console.warn('mkManutencaoCC:', e); }}
  try {{ mkManutencao(); }} catch(e) {{ console.warn('mkManutencao:', e); }}
  try {{ mkMedicoes(); }} catch(e) {{ console.warn('mkMedicoes:', e); }}
  try {{ mkHorasApp(); }} catch(e) {{ console.warn('mkHorasApp:', e); }}
}}

function aplicarFiltros() {{ filtrar(); atualizar(); }}
function limparFiltros() {{
  document.getElementById('fDateIni').value = '{dt_min_iso}';
  document.getElementById('fDateFim').value = '{dt_max_iso}';
  document.getElementById('filtroInfo').textContent = '';
  const tP = document.getElementById('tituloPonto');
  if (tP) tP.textContent = '\uD83D\uDD50 Ponto Colaborador \u2014 {periodo}';
  dadosFilt     = VENDAS;
  pagarFilt     = PAGAR;
  pagarFiltFin  = PAGAR;
  pontoMarcFilt = PONTO_MARC.filter(r => r.data >= '{ponto_d_ini_iso}' && r.data <= '{ponto_d_fim_iso}');
  lojaFiltAtivo = 'ambas';
  document.querySelectorAll('.loja-pill').forEach(b => {{
    b.classList.toggle('active', b.dataset.loja === 'ambas');
  }});
  atualizar();
}}

// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
//  DARK MODE
// \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
function toggleTheme() {{
  const body = document.body;
  const isDark = body.getAttribute('data-theme') === 'dark';
  const next = isDark ? 'light' : 'dark';
  body.setAttribute('data-theme', next);
  document.getElementById('btn-theme').textContent = isDark ? String.fromCodePoint(0x1F319) : String.fromCodePoint(0x2600);
  try {{ localStorage.setItem('locvix-theme', next); }} catch(e){{}}
  const gc = isDark ? '#e2e8f0' : '#e2e8f0';
  const lc = isDark ? '#94a3b8' : '#4a5568';
  const toDark = next === 'dark';
  Chart.defaults.color = toDark ? '#cbd5e1' : '#1e293b';
  Object.values(Chart.instances).forEach(ch => {{
    if (ch.options.scales) {{
      Object.values(ch.options.scales).forEach(sc => {{
        if (sc.grid) sc.grid.color = toDark ? '#334155' : '#e2e8f0';
        if (sc.ticks) sc.ticks.color = toDark ? '#cbd5e1' : '#1e293b';
      }});
    }}
    if (ch.options.plugins?.legend) ch.options.plugins.legend.labels.color = toDark ? '#cbd5e1' : '#1e293b';
    ch.update();
  }});
}}

function abrirTelaCheia() {{
  try {{
    const html = document.documentElement.outerHTML;
    const blob = new Blob([html], {{type: 'text/html;charset=utf-8'}});
    const url  = URL.createObjectURL(blob);
    window.open(url, '_blank');
  }} catch(e) {{
    alert('N\u00E3o foi poss\u00EDvel abrir em tela cheia: ' + e.message);
  }}
}}

document.addEventListener('DOMContentLoaded', () => {{
  const i = document.getElementById('pg-inp'); if(i) i.focus();
  // Define data padrão do formulário de manutenção como hoje
  const mfd = document.getElementById('mFormData');
  if (mfd) mfd.value = new Date().toISOString().slice(0,10);
  // Popula select de equipamentos
  const mSel = document.getElementById('mFormEquip');
  if (mSel && MANUTENCAO && MANUTENCAO.length) {{
    MANUTENCAO.slice().sort((a,b) => a.cc.localeCompare(b.cc)).forEach(r => {{
      const o = document.createElement('option'); o.value = r.cc; o.textContent = r.cc; mSel.appendChild(o);
    }});
    mSel.addEventListener('change', () => {{
      const rec = MANUTENCAO.find(r => r.cc === mSel.value);
      const dt = document.getElementById('mFormData');
      if (dt && rec && rec.ultima) dt.value = rec.ultima;
      const srv = document.getElementById('mFormServico');
      if (srv && rec && rec.tipo_servico) srv.value = rec.tipo_servico;
    }});
  }}
  const mSrv = document.getElementById('mFormServico');
  try {{
    const saved = localStorage.getItem('locvix-theme');
    if (saved === 'dark') {{
      document.body.setAttribute('data-theme','dark');
      const btn = document.getElementById('btn-theme');
      if (btn) btn.textContent = '\u2600\ufe0f';
    }}
  }} catch(e) {{}}
}});

// \u2500\u2500\u2500 Init \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
Chart.defaults.font.family = "'Segoe UI', Arial, sans-serif";
Chart.defaults.font.size   = 12;
Chart.defaults.color       = '#cbd5e1';
dadosFilt = VENDAS;
// Sincroniza pontoMarcFilt com o per\u00EDodo do ponto (independente do range de vendas)
(function() {{
  const ini = '{ponto_d_ini_iso}';
  const fim = '{ponto_d_fim_iso}';
  pontoMarcFilt = PONTO_MARC.filter(r => {{
    if (ini && r.data < ini) return false;
    if (fim && r.data > fim) return false;
    return true;
  }});
}})();
_populaFiltrosFinanceiro();
atualizar();
try {{ const sm = localStorage.getItem('locvix-modulo'); if(sm) setModulo(sm); }} catch(e){{}}
// Auto-resize iframe para mobile (Streamlit component resize protocol)
(function() {{
  function _sendH() {{
    var h = Math.max(document.body.scrollHeight, document.documentElement.scrollHeight, 600);
    try {{ window.parent.postMessage({{type:'streamlit:setFrameHeight', height:h}}, '*'); }} catch(e){{}}
  }}
  if (typeof ResizeObserver !== 'undefined') {{
    new ResizeObserver(_sendH).observe(document.body);
  }}
  window.addEventListener('load', _sendH);
  setTimeout(_sendH, 600);
  setTimeout(_sendH, 1500);
}})();

function abrirAjuda() {{
  document.getElementById('helpOverlay').classList.add('open');
  try {{
    var r = window.frameElement.getBoundingClientRect();
    window.parent.scrollTo({{top: window.parent.pageYOffset + r.top, behavior:'smooth'}});
  }} catch(e) {{}}
}}
function fecharAjuda() {{ document.getElementById('helpOverlay').classList.remove('open'); }}
function setHmTab(id) {{
  document.querySelectorAll('.hm-tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.hm-panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.hm-tab').forEach(t => {{
    if (t.getAttribute('onclick') === "setHmTab('" + id + "')") t.classList.add('active');
  }});
  document.getElementById('hmp-' + id).classList.add('active');
}}
document.addEventListener('keydown', e => {{ if(e.key==='Escape') fecharAjuda(); }});
</script>
</body>
</html>"""

    # Escape ALL non-ASCII chars inside executable <script> blocks to \uXXXX sequences.
    # This prevents any emoji / special char in JS string literals from corrupting the
    # JS parser when Streamlit embeds the HTML in a srcdoc iframe.
    # <script type="application/json"> data islands are intentionally left untouched.
    import re as _re

    def _escape_js_block(m):
        attrs, body = m.group(1), m.group(2)
        # leave data islands and external scripts untouched
        if "application/json" in attrs or "src=" in attrs:
            return m.group(0)
        out = []
        for ch in body:
            cp = ord(ch)
            if cp <= 127:
                out.append(ch)
            elif cp <= 0xFFFF:
                out.append(f"\\u{cp:04X}")
            else:
                # surrogate pair for chars outside BMP (e.g. emoji U+1F550)
                cp -= 0x10000
                out.append(f"\\u{0xD800 | (cp >> 10):04X}\\u{0xDC00 | (cp & 0x3FF):04X}")
        return f"<script{attrs}>" + "".join(out) + "</script>"

    html = _re.sub(r"<script([^>]*)>(.*?)</script>", _escape_js_block, html, flags=_re.DOTALL)

    # Force ALL output to pure ASCII — every non-ASCII char becomes &#NNNN;
    # This guarantees the srcdoc iframe never sees raw Unicode bytes.
    html_clean = html.encode("ascii", errors="xmlcharrefreplace").decode("ascii")
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(html_clean)
    print(f"  OK Dashboard HTML salvo: {caminho}")
    return html_clean


# ══════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════
def main(
    saida_html:   str | None = None,
    saida_excel:  str | None = None,
    data_ini:     str | None = None,
    data_fim:     str | None = None,
    fonte_vendas: str        = "api",   # "api" = GestãoClick  |  "excel" = planilha manual
    loja_filtro:  str | None = None,    # None | "ambas" | "521831" (G&J) | "65731" (W&A)
) -> str | None:
    """
    Executa a coleta completa de dados e gera Excel + Dashboard HTML.

    Parâmetros
    ----------
    saida_html   : caminho do HTML gerado (usa SAIDA_HTML global se None)
    saida_excel  : caminho do Excel (usa SAIDA_EXCEL global se None; "" = pula Excel)
    data_ini     : início do período "DD/MM/AAAA" (usa DATA_INI global se None)
    data_fim     : fim do período   "DD/MM/AAAA" (usa DATA_FIM global se None)
    fonte_vendas : "api" (GestãoClick) ou "excel" (planilhas manuais)

    Retorna o conteúdo HTML como string (útil para Streamlit).
    """
    global _client, LOJA_FILTRO

    LOJA_FILTRO = loja_filtro

    d_ini  = data_ini  or DATA_INI
    d_fim  = data_fim  or DATA_FIM
    h_path = saida_html  or SAIDA_HTML
    x_path = saida_excel if saida_excel is not None else SAIDA_EXCEL

    print(f"\n{'='*60}")
    print(f"  LOCVIX — ANÁLISE DE DADOS VIA GESTÃOCLICK API")
    print(f"  Período: {d_ini} → {d_fim}")
    print(f"{'='*60}")

    # Verifica credenciais
    if GCK_ACCESS_TOKEN == "SEU_ACCESS_TOKEN_AQUI":
        print("\n  ⚠ ATENÇÃO: Credenciais não configuradas!")
        print("  Configure as variáveis de ambiente:")
        print("    GCK_ACCESS_TOKEN  = seu Access Token do GestãoClick")
        print("    GCK_SECRET_TOKEN  = seu Secret Token do GestãoClick")
        print("  Ou edite as constantes GCK_ACCESS_TOKEN / GCK_SECRET_TOKEN")
        print("  no início deste arquivo.\n")
        print("  Gerando dashboard com dados de DEMONSTRAÇÃO...\n")
        return _main_demo(h_path, d_ini, d_fim)

    # Reinicia o cliente com as credenciais atuais
    _client = GCKClient(GCK_ACCESS_TOKEN, GCK_SECRET_TOKEN)

    # ── Coleta de dados ────────────────────────────────────────────
    if fonte_vendas == "excel":
        print("  📂 Fonte de vendas: Excel manual")
        vendas = buscar_vendas_excel(d_ini, d_fim)
    else:
        print("  🌐 Fonte de vendas: API GestãoClick")
        vendas = buscar_vendas(d_ini, d_fim)
    financ    = buscar_financeiro(d_ini, d_fim)

    # Busca pagamentos em janela fixa 2 anos atras / 2 anos a frente
    # para o grafico de vencimentos por mes (independente do filtro)
    _hoje_d   = (datetime.utcnow() - timedelta(hours=3)).date()
    _ini_all  = (_hoje_d - timedelta(days=730)).strftime("%d/%m/%Y")
    _fim_all  = (_hoje_d + timedelta(days=730)).strftime("%d/%m/%Y")
    financ_all = buscar_financeiro(_ini_all, _fim_all)
    pagar_all  = financ_all.get("pagar", [])
    clientes  = buscar_clientes()
    produtos  = buscar_produtos()
    contratos = buscar_contratos()
    orcamentos  = buscar_orcamentos()
    medicoes    = buscar_medicoes()
    horas_app   = buscar_horas_app(d_ini, d_fim)
    manutencoes = buscar_manutencoes()
    os_list     = buscar_ordens_servico(d_ini, d_fim)
    ponto_data  = buscar_ponto(d_ini, d_fim)

    receber = financ.get("receber", [])
    pagar   = financ.get("pagar", [])

    _prog(0.82, "Processando dados...")

    # ── DataFrame de vendas ─────────────────────────────────────────
    if vendas:
        df = pd.DataFrame(vendas)
        # Garante coluna Data como datetime
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    else:
        df = pd.DataFrame(columns=[
            "ID","NF","Data","Cliente","Status","Vendedor","Categoria",
            "Cod. Produto","Produto","Unidade","Qtd",
            "Vlr Unitário","Vlr Bruto","Desconto","Vlr Líquido"
        ])
        print("  ⚠ Sem dados de vendas no período.")

    # ── Gera arquivos ───────────────────────────────────────────────
    _prog(0.88, "Gerando Excel...")
    if x_path:
        gerar_excel(df, receber, pagar, os_list, contratos, x_path)

    _prog(0.93, "Gerando Dashboard HTML...")
    html_content = gerar_dashboard_html(
        df_vendas=df, receber=receber, pagar=pagar, pagar_all=pagar_all,
        os_list=os_list, contratos=contratos,
        caminho=h_path, data_ini=d_ini, data_fim=d_fim,
        ponto_data=ponto_data, orcamentos=orcamentos, medicoes=medicoes,
      horas_app=horas_app, manutencoes=manutencoes,
    )

    _prog(1.0, "✔ Concluído!")
    print(f"\n  ✅ Tudo pronto!")
    print(f"  📊 Dashboard: {h_path}")
    if x_path:
        print(f"  📋 Excel:     {x_path}")
    print(f"{'='*60}\n")
    return html_content


# ══════════════════════════════════════════════════════════════════
#  MODO DEMONSTRAÇÃO (sem credenciais)
# ══════════════════════════════════════════════════════════════════
def _main_demo(h_path: str, d_ini: str, d_fim: str) -> str:
    """
    Gera um dashboard de demonstração com dados fictícios
    para validar o layout sem precisar das credenciais.
    """
    import random
    rng = random.Random(42)

    clientes_demo = ["Empresa Alpha Ltda", "Beta Soluções S/A", "Gamma Tech",
                     "Delta Serviços", "Epsilon Comercial", "Zeta Distribuidora",
                     "Eta Logística", "Theta Consultoria", "Iota Indústria"]
    cats_demo     = ["SERVIÇOS", "PRODUTOS", "SUPORTE", "INSTALAÇÃO", "MANUTENÇÃO"]
    vendedores_demo = ["João Silva", "Maria Souza", "Carlos Oliveira", "Ana Santos"]
    produtos_demo = [
        {"cod":"PROD001","desc":"Sistema ERP Módulo"},
        {"cod":"PROD002","desc":"Licença Software Anual"},
        {"cod":"SERV001","desc":"Hora de Suporte Técnico"},
        {"cod":"SERV002","desc":"Implantação"},
        {"cod":"SERV003","desc":"Treinamento Usuários"},
        {"cod":"PROD003","desc":"Firewall Hardware"},
        {"cod":"PROD004","desc":"Switch 24 portas"},
        {"cod":"SERV004","desc":"Monitoramento Remoto"},
    ]

    # Gera vendas fictícias
    vendas_demo = []
    try:
        d1 = datetime.strptime(d_ini, "%d/%m/%Y")
        d2 = datetime.strptime(d_fim, "%d/%m/%Y")
    except Exception:
        d1 = datetime(datetime.now().year, 1, 1)
        d2 = datetime.now()

    delta = (d2 - d1).days
    for i in range(min(delta * 3, 500)):
        data = d1 + timedelta(days=rng.randint(0, max(delta, 1)))
        prod = rng.choice(produtos_demo)
        qtd  = rng.choice([1, 1, 1, 2, 5, 10, 20])
        prec = rng.uniform(50, 2000)
        desc = rng.uniform(0, 0.1) * prec * qtd
        vbruto = round(prec * qtd, 2)
        vliq   = round(vbruto - desc, 2)
        vendas_demo.append({
            "ID":           f"PED{i+1:04d}",
            "NF":           f"{1000+i}",
            "Data":         data,
            "Cliente":      rng.choice(clientes_demo),
            "Status":       rng.choice(["EMITIDO","EMITIDO","PAGO","CANCELADO"]),
            "Vendedor":     rng.choice(vendedores_demo),
            "Categoria":    rng.choice(cats_demo),
            "Cod. Produto": prod["cod"],
            "Produto":      prod["desc"],
            "Unidade":      "UN",
            "Qtd":          float(qtd),
            "Vlr Unitário": round(prec, 2),
            "Vlr Bruto":    vbruto,
            "Desconto":     round(desc, 2),
            "Vlr Líquido":  vliq,
        })

    df = pd.DataFrame(vendas_demo)

    # Financeiro fictício
    receber_demo = [
        {"ID": f"R{i}", "Descrição": f"NF {1000+i}", "Pessoa": rng.choice(clientes_demo),
         "Valor": round(rng.uniform(500,5000),2), "Valor Pago": 0.0,
         "Saldo": round(rng.uniform(500,5000),2),
         "Vencimento": (datetime.now() + timedelta(days=rng.randint(-30,60))),
         "Pagamento": None, "Status": rng.choice(["ABERTO","VENCIDO","RECEBIDO"]),
         "Categoria": rng.choice(cats_demo), "Centro Custo": ""}
        for i in range(30)
    ]
    pagar_demo = [
        {"ID": f"P{i}", "Descrição": f"Fornecedor {i}", "Pessoa": f"Fornecedor {i+1}",
         "Valor": round(rng.uniform(200,3000),2), "Valor Pago": round(rng.uniform(0,200),2),
         "Saldo": round(rng.uniform(0,3000),2),
         "Vencimento": (datetime.now() + timedelta(days=rng.randint(-15,45))),
         "Pagamento": None, "Status": rng.choice(["ABERTO","PAGO","VENCIDO"]),
         "Categoria": rng.choice(["PESSOAL","INFRA","FORNECEDOR","IMPOSTOS"]), "Centro Custo": ""}
        for i in range(20)
    ]

    # OS fictícias
    os_demo = [
        {"ID": f"OS{i:03d}", "Número": f"OS{i:03d}",
         "Data": d1 + timedelta(days=rng.randint(0, max(delta,1))),
         "Cliente": rng.choice(clientes_demo),
         "Tecnico": rng.choice(["Carlos", "Ana", "Marcos"]),
         "Descricao": rng.choice(["Manutenção preventiva","Instalação","Suporte remoto","Atualização"]),
         "Status": rng.choice(["ABERTA","CONCLUÍDO","AGUARDANDO","CANCELADA"]),
         "Prioridade": rng.choice(["ALTA","NORMAL","BAIXA"]),
         "Valor": round(rng.uniform(100,800),2),
         "Fechamento": None}
        for i in range(50)
    ]

    # Contratos fictícios
    contratos_demo = [
        {"ID": f"C{i:03d}", "Número": f"C{i:03d}",
         "Cliente": rng.choice(clientes_demo),
         "Descricao": "Contrato de manutenção e suporte",
         "Valor": rng.choice([500,800,1200,1500,2000,3000]),
         "Periodicidade": "MENSAL",
         "Inicio": d1 - timedelta(days=rng.randint(30,365)),
         "Fim": d1 + timedelta(days=rng.randint(180,730)),
         "Status": rng.choice(["ATIVO","ATIVO","ATIVO","VENCIDO","CANCELADO"])}
        for i in range(15)
    ]

    html_content = gerar_dashboard_html(
        df_vendas=df, receber=receber_demo, pagar=pagar_demo,
        os_list=os_demo, contratos=contratos_demo,
        caminho=h_path, data_ini=d_ini, data_fim=d_fim,
    )
    print(f"\n  📊 Dashboard DEMO gerado: {h_path}")
    return html_content


# ══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import webbrowser
    html = main()
    # Abre automaticamente no navegador padrão
    if os.path.exists(SAIDA_HTML):
        webbrowser.open(f"file:///{SAIDA_HTML.replace(chr(92),'/')}")
